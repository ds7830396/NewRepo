import os
import asyncio
import base64
import qrcode
import sqlite3
import threading
import time
import queue
import logging
import hashlib
import functools
from io import BytesIO
from datetime import datetime
from telethon.errors import AuthKeyUnregisteredError
from flask import (
    Flask, session, request, render_template_string,
    jsonify, g, redirect, url_for, send_from_directory
)
from telethon import TelegramClient, events
import json
from flask_socketio import SocketIO, emit
from telethon.events import MessageDeleted

import secrets
import math
import pandas as pd
import tempfile
import pdfplumber
sqlite3.register_adapter(datetime, lambda d: d.strftime("%Y-%m-%d %H:%M:%S"))

# Настройка логирования
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.secret_key = 'supersecretkey_for_demo'
app.config['DATABASE'] = 'app.db'
app.json.ensure_ascii = False
# ПЕРЕНЕСТИ СЮДА (строки, которые сейчас перед шаблонами)
socketio = SocketIO(app, cors_allowed_origins="*") 

def notify_clients(data_type="general"):
    socketio.emit('db_updated', {'type': data_type})
# ---------- Глобальные структуры ----------
background_tasks = {}      # session_id -> thread
user_clients = {}          # session_id -> (thread, client, user_id)
qr_sessions = {}           # session_id -> данные для QR
listener_locks = {}        # блокировки для каждого session_id
user_loops = {}            # session_id -> asyncio event loop


import asyncio
import logging

# --- ПОДАВЛЕНИЕ СИСТЕМНЫХ ОШИБОК TELETHON ПРИ ЗАКРЫТИИ ПОТОКОВ ---
def custom_exception_handler(loop, context):
    exception = context.get("exception")
    msg_str = str(exception) if exception else str(context.get("message", ""))
    
    # 1. Игнорируем ошибки закрытого цикла
    if isinstance(exception, (RuntimeError, GeneratorExit)) or "Task was destroyed" in msg_str or "Event loop is closed" in msg_str:
        return
    
    # 2. Игнорируем внутренний баг Telethon при отмене задач (AttributeError)
    if isinstance(exception, AttributeError) and "'NoneType' object has no attribute" in msg_str and ("recv" in msg_str or "disconnect" in msg_str):
        return
        
    # Для остальных, реальных ошибок используем стандартный обработчик
    loop.default_exception_handler(context)

# Применяем этот обработчик ко всем новым потокам
# (Вам нужно добавить `loop.set_exception_handler(custom_exception_handler)` 
# в функцию, где вы создаете новый цикл `loop = asyncio.new_event_loop()`)

import functools
from flask import session, redirect, url_for
from datetime import datetime, timedelta, timezone
def login_required(f):
    @functools.wraps(f)
    def wrapped(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login_page'))
            
        # --- НОВАЯ СТРОГАЯ ПРОВЕРКА СЕССИИ ---
        user_id = session['user_id']
        db = get_db()
        user = db.execute("SELECT session_token FROM users WHERE id = ?", (user_id,)).fetchone()
        
        # Если в базе есть токен, а в текущем браузере он не совпадает — разлогиниваем
        if user and 'session_token' in user.keys() and user['session_token']:
            if session.get('session_token') != user['session_token']:
                session.clear()
                return redirect(url_for('login_page'))
        # --------------------------------------
        
        return f(*args, **kwargs)
    return wrapped

def ensure_event_loop():
    try:
        return asyncio.get_running_loop()
    except RuntimeError:
        loop = asyncio.new_event_loop()
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        return loop

# ---------- Работа с БД ----------
def get_db():
    db = getattr(g, '_database', None)
    if db is None:
        # 1. Добавляем timeout=20. Если база занята, поток подождет 20 секунд, а не упадет сразу
        db = g._database = sqlite3.connect(app.config['DATABASE'], timeout=20)
        
        # 2. Включаем режим WAL для параллельной работы
        db.execute("PRAGMA journal_mode=WAL;")
        
        # 3. Синхронизация (опционально для ускорения записи)
        db.execute("PRAGMA synchronous=NORMAL;")
        
        db.row_factory = sqlite3.Row
    return db

def bot_interaction_scheduler():
    """Фоновый поток для автоматической отправки команд ботам по расписанию"""
    while True:
        time.sleep(60)  # Проверяем каждую минуту
        try:
            with app.app_context():
                db = get_db()
                active_interactions = db.execute("SELECT * FROM interaction_bots WHERE status = 'active'").fetchall()
                now = datetime.now()
                
                for interaction in active_interactions:
                    # 1. СНАЧАЛА достаем ID юзербота
                    userbot_id = interaction['userbot_id']
                    
                    # 2. ЗАТЕМ проверяем, рабочее ли у него сейчас время
                    if not is_parsing_allowed(userbot_id):
                        continue 
                        
                    last_run = interaction['last_run']
                    if last_run:
                        # Учитываем формат времени SQLite
                        try:
                            last_run_dt = datetime.strptime(last_run, '%Y-%m-%d %H:%M:%S.%f')
                        except ValueError:
                            last_run_dt = datetime.strptime(last_run, '%Y-%m-%d %H:%M:%S')
                            
                        diff_minutes = (now - last_run_dt).total_seconds() / 60
                        if diff_minutes < interaction['interval_minutes']:
                            continue  # Время еще не пришло
                    
                    # Если время пришло, отправляем команды
                    # Если время пришло, отправляем команды
                    if userbot_id in user_clients:
                        _, client, _ = user_clients[userbot_id]
                        loop = user_loops.get(userbot_id)
                        
                        # ДОБАВЛЕНА ПРОВЕРКА client.is_connected()
                        if loop and loop.is_running() and client.is_connected():
                            commands = json.loads(interaction['commands'])
                            bot_username = interaction['bot_username']
                            
                            # Отправляем команды по очереди
                            for cmd in commands:
                                asyncio.run_coroutine_threadsafe(client.send_message(bot_username, cmd), loop)
                                time.sleep(2) # Небольшая пауза между командами
                            
                            # Обновляем время последнего запуска
                            db.execute("UPDATE interaction_bots SET last_run = ? WHERE id = ?", (now, interaction['id']))
                            db.commit()
                            
                            if 'notify_clients' in globals():
                                notify_clients()
                             
        except Exception as e:
            logger.error(f"Ошибка в планировщике ботов: {e}")





# --- Маршруты для Товаров ---
@app.before_request
def check_session_token():
    # Пропускаем проверку для страниц логина и логаута
    if request.endpoint in ['login_page', 'static', 'logout']:
        return
        
    if 'user_id' in session:
        try:
            db = get_db()
            user = db.execute("SELECT session_token FROM users WHERE id = ?", (session['user_id'],)).fetchone()
            
            if user and 'session_token' in user.keys() and user['session_token']:
                if session.get('session_token') != user['session_token']:
                    session.clear()
                    
                    # Если запрос пришел от фонового скрипта (fetch/ajax)
                    if request.path.startswith('/api/'):
                        return jsonify({'error': 'Сеанс завершен'}), 401
                        
                    # Если это обычная загрузка страницы
                    return redirect(url_for('login_page'))
        except Exception:
            pass





@app.route('/api/products/<int:prod_id>/messages')
@login_required
def get_product_messages(prod_id):
    user_id = session['user_id']
    db = get_db()
    
    # 1. Получаем синонимы товара
    prod = db.execute("SELECT synonyms FROM products WHERE id = ? AND user_id = ?", (prod_id, user_id)).fetchone()
    if not prod or not prod['synonyms']:
        return jsonify({'proposed': [], 'confirmed': []})
    
    # --- УМНАЯ ОЧИСТКА (превращает строку в набор слов) ---
    # --- УЛУЧШЕННАЯ ОЧИСТКА ---
    import re
    def clean_for_search(text):
        if not text: return set()
        
        # ДОБАВЛЯЕМ '*' (звездочку) в список удаляемых символов
        # Также добавили '.', если вдруг поставщик пишет "17.512"
        t = re.sub(r'[()*.[\]{},;!|/+\-]', ' ', text)
        
        # Убираем лишние пробелы и неразрывные пробелы
        t = re.sub(r'\s+', ' ', t.replace('\xa0', ' ')).strip().lower()
        
        # Решаем проблему кириллицы/латиницы
        t = t.translate(str.maketrans('асеокрх', 'aceokpx'))
        
        # Возвращаем набор чистых слов
        return set(t.split())

    # Подготовка синонимов (разбиваем на сеты слов)
    synonyms_as_word_sets = []
    for s in prod['synonyms'].split(','):
        if s.strip():
            words = clean_for_search(s)
            if words:
                synonyms_as_word_sets.append(words)

    # --- ИСПРАВЛЕНИЕ: Получаем все последние сообщения (all_msgs) ---
    # Без этого блока была ошибка "all_msgs is not defined"
    all_msgs_rows = db.execute("""
        SELECT m.id, m.text, m.date, m.chat_id, m.sender_name, m.chat_title, m.type, tc.custom_name
        FROM messages m
        LEFT JOIN tracked_chats tc ON m.chat_id = tc.chat_id AND m.user_id = tc.user_id
        WHERE m.user_id = ? 
        ORDER BY m.date DESC LIMIT 1000
    """, (user_id,)).fetchall()
    all_msgs = [dict(r) for r in all_msgs_rows]

    # --- ИСПРАВЛЕНИЕ: Собираем актуальные отпечатки (latest_fingerprints) ---
    # Без этого блока была ошибка "latest_fingerprints is not defined"
    latest_fingerprints = {}
    for msg in all_msgs:
        if not msg['text']: continue
        lines = msg['text'].split('\n')
        for i, line in enumerate(lines):
            if not line.strip(): continue
            fp = get_fingerprint(msg['chat_id'], msg['sender_name'], line)
            if fp not in latest_fingerprints or msg['date'] > latest_fingerprints[fp]['date']:
                # Ищем цену в текущей строке или на 1-2 строки ниже
                price_val = extract_price(line)
                if not price_val and i + 1 < len(lines):
                    price_val = extract_price(lines[i+1])
                if not price_val and i + 2 < len(lines):
                    price_val = extract_price(lines[i+2])

                latest_fingerprints[fp] = {
                    'text': line,
                    'price': price_val,
                    'date': msg['date'],
                    'message_id': msg['id'],
                    'line_index': i
                }

    # 2. Обработка ПОДТВЕРЖДЕННЫХ
    confirmed_rows = db.execute("""
        SELECT pm.id as pm_id, pm.group_id, pm.message_id, m.chat_title, m.chat_id, m.sender_name, 
               pm.extracted_price, m.text, m.date, pm.line_index, tc.custom_name, m.type
        FROM product_messages pm
        JOIN messages m ON pm.message_id = m.id
        LEFT JOIN tracked_chats tc ON m.chat_id = tc.chat_id AND m.user_id = tc.user_id
        WHERE pm.product_id = ? AND pm.status = 'confirmed'
    """, (prod_id,)).fetchall()
    
    confirmed_raw = []
    confirmed_fingerprints = set()
    has_updates = False 

    for r in confirmed_rows:
        d = dict(r)
        if d['line_index'] != -1 and d['text']:
            lines = d['text'].split('\n')
            if 0 <= d['line_index'] < len(lines):
                orig_line = lines[d['line_index']]
                fp = get_fingerprint(d['chat_id'], d['sender_name'], orig_line)
                confirmed_fingerprints.add(fp)
                
                latest = latest_fingerprints.get(fp)
                if latest and latest['date'] > d['date']:
                    d['text'] = latest['text']
                    d['extracted_price'] = latest['price']
                    d['date'] = latest['date'] 
                    current_id = d.get('pm_id')
                    try:
                        db.execute("UPDATE product_messages SET message_id=?, line_index=?, extracted_price=? WHERE id=?",
                                   (latest['message_id'], latest['line_index'], latest['price'], current_id))
                        has_updates = True
                    except: pass
                else:
                    d['text'] = orig_line
        confirmed_raw.append(d)
        
    if has_updates:
        db.commit()
        notify_clients()
        
    grouped_confirmed = {}
    for d in confirmed_raw:
        g_id = d['group_id'] if d['group_id'] else d['pm_id']
        if g_id not in grouped_confirmed or d['date'] > grouped_confirmed[g_id]['date']:
            grouped_confirmed[g_id] = d
    
    confirmed = sorted(list(grouped_confirmed.values()), key=lambda x: x['date'], reverse=True)
    
    # 3. Обработка ПРЕДЛОЖЕННЫХ (склеивание дубликатов и поиск по словам)
    proposed_list = []
    seen_fps = set()
    
    for row in all_msgs:
        lines = row['text'].split('\n') if row['text'] else []
        for i, line in enumerate(lines):
            if not line.strip(): continue
            fp = get_fingerprint(row['chat_id'], row['sender_name'], line)
            
            if fp in seen_fps or fp in confirmed_fingerprints:
                continue
            
            # --- ЛОГИКА ЖЕСТКОГО ПОИСКА ПО СЛОВАМ ---
            line_words = clean_for_search(line)
            is_match = False
            for syn_set in synonyms_as_word_sets:
                # Если ВСЕ слова из синонима есть в строке поставщика
                if syn_set.issubset(line_words):
                    is_match = True
                    break
            
            if is_match:
                seen_fps.add(fp)
                
                # Ищем цену в текущей строке или на 1-2 строки ниже
                s_price = extract_price(line)
                if not s_price and i + 1 < len(lines):
                    s_price = extract_price(lines[i+1])
                if not s_price and i + 2 < len(lines):
                    s_price = extract_price(lines[i+2])
                
                proposed_list.append({
                    'message_id': row['id'],
                    'line_index': i,
                    'chat_title': row['chat_title'],
                    'sender_name': row['sender_name'],
                    'custom_name': row.get('custom_name'),
                    'type': row['type'],
                    'match_line': line,
                    'suggested_price': s_price,
                    'date': row['date']
                })
    
    proposed_list.sort(key=lambda x: x['date'], reverse=True)
    return jsonify({'proposed': proposed_list, 'confirmed': confirmed})

import re

def parse_number(s):
    clean = re.sub(r'[^\d]', '', s)
    return float(clean) if clean else 0.0

def extract_price(line):
    if not line: return None
    # Добавил K/R для обработки "KR" как на вашем скриншоте
    line_clean = re.sub(r'[₽$€рRkKкК]', '', line) 
    number_pattern = r'(\d[\d\s.,]*)'
    all_numbers = list(re.finditer(number_pattern, line_clean))
    
    if not all_numbers: return None
    
    for match in reversed(all_numbers):
        number_str = match.group(1).strip()
        try:
            price_value = parse_number(number_str)
            if price_value < 3000: continue
        except: continue
        
        end_pos = match.end()
        text_after = line_clean[end_pos:].strip().lower()
        if re.match(r'^[-]?\s*(шт|шт\.|штук|pcs|pc|item|ед|единиц)', text_after): continue
        
        text_before = line_clean[:match.start()].strip()
        if text_before.endswith('-'): return price_value
        if end_pos >= len(line_clean) or line_clean[end_pos] in '()': return price_value
        if end_pos == len(line_clean) or not line_clean[end_pos].isdigit(): return price_value
        
    for match in reversed(all_numbers):
        number_str = match.group(1).strip()
        try:
            price_value = parse_number(number_str)
            if price_value >= 1500:
                end_pos = match.end()
                text_after = line_clean[end_pos:].strip().lower()
                if not re.match(r'^[-]?\s*(шт|шт\.|штук|pcs|pc|item|ед|единиц)', text_after):
                    return price_value
        except: continue
    return None

def get_fingerprint(chat_id, sender_name, line):
    """Создает 'слепок' строки, удаляя цены для сравнения товаров"""
    # Заменяем все числа от 1000 и выше (с точками и пробелами) на тег [PRICE]
    norm_line = re.sub(r'\b\d{1,3}(?:[ .,]\d{3})+\b|\b\d{4,}\b', '[PRICE]', line.lower())
    norm_line = re.sub(r'[₽$€рRkKкК]', '', norm_line).strip()
    return f"{chat_id}_{sender_name}_{norm_line}"



@app.route('/api/product_messages/confirm', methods=['POST'])
@login_required
def confirm_product_message():
    data = request.get_json()
    db = get_db()
    try:
        line_idx = data.get('line_index', -1)
        price = data.get('price')
        if price == '': 
            price = None
            
        db.execute("""
            INSERT INTO product_messages (product_id, message_id, line_index, status, extracted_price) 
            VALUES (?, ?, ?, 'confirmed', ?)
            ON CONFLICT(product_id, message_id, line_index) DO UPDATE SET status='confirmed', extracted_price=excluded.extracted_price
        """, (data['product_id'], data['message_id'], line_idx, price))
        db.commit()
        notify_clients()
         
        return jsonify({'success': True})
    except Exception as e:
        logger.error(f"Ошибка привязки: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/product_messages/detach', methods=['POST'])
@login_required
def detach_product_message():
    data = request.get_json()
    db = get_db()
    # Ищем, состоит ли сообщение в группе. Если да — удаляем всю группу
    pm = db.execute("SELECT id, group_id FROM product_messages WHERE product_id=? AND message_id=?", 
                    (data['product_id'], data['message_id'])).fetchone()
    if pm:
        if pm['group_id']:
            db.execute("DELETE FROM product_messages WHERE group_id = ?", (pm['group_id'],))
        else:
            db.execute("DELETE FROM product_messages WHERE id = ?", (pm['id'],))
        db.commit()
        notify_clients()
         
    return jsonify({'success': True})

@app.route('/api/product_messages/update_price', methods=['POST'])
@login_required
def update_product_price():
    data = request.get_json()
    db = get_db()
    db.execute("UPDATE product_messages SET extracted_price = ? WHERE id = ?", 
               (data['price'], data['pm_id']))
    db.commit()
    notify_clients()
     
    return jsonify({'success': True})



@app.route('/api/products', methods=['GET', 'POST'])
@login_required
def manage_products():
    user_id = session['user_id']
    db = get_db()
    if request.method == 'GET':
        products = db.execute("SELECT * FROM products WHERE user_id=?", (user_id,)).fetchall()
        return jsonify([dict(p) for p in products])
    
    # POST
    data = request.get_json()
    folder_id = data.get('folder_id') # Получаем ID папки из запроса
    
    # ДОБАВЛЕНО: сохранение folder_id в базу данных
    db.execute("INSERT INTO products (user_id, name, synonyms, price, folder_id) VALUES (?, ?, ?, ?, ?)",
               (user_id, data.get('name'), data.get('synonyms', ''), data.get('price', 0.0), folder_id))
    db.commit()
    notify_clients()
     
    return jsonify({'success': True})

@app.route('/api/products/<int:prod_id>', methods=['DELETE'])
@login_required
def delete_product(prod_id):
    db = get_db()
    db.execute("DELETE FROM products WHERE id=? AND user_id=?", (prod_id, session['user_id']))
    db.commit()
    notify_clients()
    return jsonify({'success': True})


@app.route('/api/reports/confirmed')
@login_required
def get_all_confirmed():
    user_id = session['user_id']
    db = get_db()
    
    query = """
        SELECT pm.id as binding_id, pm.group_id, p.id as product_id, p.name as product_name, pm.extracted_price, m.text, 
               m.chat_id, m.chat_title, m.sender_name, m.date, pm.line_index, pm.message_id, pm.is_actual,
               m.telegram_message_id, m.type, ec.sheet_url, tc.chat_title as tc_chat_title, tc.custom_name
        FROM product_messages pm
        JOIN products p ON pm.product_id = p.id
        JOIN messages m ON pm.message_id = m.id
        LEFT JOIN tracked_chats tc ON m.chat_id = tc.chat_id AND m.user_id = tc.user_id
        LEFT JOIN excel_configs ec ON m.chat_id = ec.chat_id AND m.type LIKE 'excel%'
        WHERE p.user_id = ?
        AND pm.status = 'confirmed'
    """
    rows = db.execute(query, (user_id,)).fetchall()
    
    # Собираем актуальные версии строк (с добавленными telegram_message_id и type)
    all_msgs = db.execute("SELECT id, text, chat_id, sender_name, date, telegram_message_id, type FROM messages WHERE user_id = ? AND (is_blocked IS NULL OR is_blocked = 0) AND (is_delayed IS NULL OR is_delayed = 0) ORDER BY date DESC", (user_id,)).fetchall()
    latest_fingerprints = {}
    for row in all_msgs:
        if not row['text']: continue
        lines = row['text'].split('\n')
        for i, line in enumerate(lines):
            if not line.strip(): continue
            fp = get_fingerprint(row['chat_id'], row['sender_name'], line)
            if fp not in latest_fingerprints:
                price_val = extract_price(line)
                if not price_val and i + 1 < len(lines):
                    price_val = extract_price(lines[i+1])
                if not price_val and i + 2 < len(lines):
                    price_val = extract_price(lines[i+2])

                latest_fingerprints[fp] = {
                    'message_id': row['id'],
                    'line_index': i,
                    'text': line,
                    'date': row['date'],
                    'price': price_val,
                    'telegram_message_id': row['telegram_message_id'],
                    'type': row['type']
                }

    result_raw = []
    has_updates = False 
    for r in rows:
        d = dict(r)
        if d['line_index'] != -1 and d['text']:
            lines = d['text'].split('\n')
            if 0 <= d['line_index'] < len(lines):
                orig_line = lines[d['line_index']]
                fp = get_fingerprint(d['chat_id'], d['sender_name'], orig_line)
                
                latest = latest_fingerprints.get(fp)
                if latest and latest['date'] > d['date']:
                    d['text'] = latest['text']
                    d['extracted_price'] = latest['price']
                    d['date'] = latest['date']
                    d['telegram_message_id'] = latest['telegram_message_id']
                    d['type'] = latest['type']
                    
                    current_id = d.get('pm_id') or d.get('id')

                    try:
                        db.execute("UPDATE product_messages SET message_id=?, line_index=?, extracted_price=? WHERE id=?",
                                   (latest['message_id'], latest['line_index'], latest['price'], current_id))
                        has_updates = True
                    except sqlite3.IntegrityError:
                        if current_id:
                            db.execute("DELETE FROM product_messages WHERE id=?", (current_id,))
                            has_updates = True
                else:
                    d['text'] = orig_line
        result_raw.append(d)
        
    if has_updates:
        db.commit()
        notify_clients()

    grouped_reports = {}
    for d in result_raw:
        g_id = d['group_id'] if d['group_id'] else d['binding_id']
        if g_id not in grouped_reports or d['date'] > grouped_reports[g_id]['date']:
            grouped_reports[g_id] = d
            
    final_result = list(grouped_reports.values())
    final_result.sort(key=lambda x: x['date'], reverse=True)
    return jsonify(final_result)


# --- Маршруты для Взаимодействия с ботами ---
@app.route('/api/interaction_bots', methods=['GET', 'POST'])
@login_required
def manage_interaction_bots():
    user_id = session['user_id']
    db = get_db()
    if request.method == 'GET':
        bots = db.execute("SELECT * FROM interaction_bots WHERE user_id=?", (user_id,)).fetchall()
        return jsonify([dict(b) for b in bots])
    
    # POST
    data = request.get_json()
    bot_username = data.get('bot_username').strip()
    custom_name = data.get('custom_name')
    userbot_id = data.get('userbot_id')
    
    # 1. Находим запущенного юзербота (он нужен, чтобы узнать числовой ID бота)
    client_to_use = None
    client_loop_to_use = None
    
    if userbot_id in user_clients:
        _, client_to_use, _ = user_clients[userbot_id]
        client_loop_to_use = user_loops.get(userbot_id)
    else:
        active_sessions = [sid for sid, (_, _, uid) in user_clients.items() if uid == user_id]
        if active_sessions:
            sid = active_sessions[0]
            _, client_to_use, _ = user_clients[sid]
            client_loop_to_use = user_loops.get(sid)

    # 2. Превращаем текстовый @username в ЧИСЛОВОЙ ID для парсера
    numeric_chat_id = None
    if client_to_use and client_loop_to_use:
        async def get_bot_id():
            try:
                entity = await client_to_use.get_entity(bot_username)
                return entity.id
            except Exception as e:
                logging.error(f"Не удалось получить ID для {bot_username}: {e}")
                return None
        
        try:
            future = asyncio.run_coroutine_threadsafe(get_bot_id(), client_loop_to_use)
            numeric_chat_id = future.result(timeout=10)
        except Exception:
            pass

    # Если юзербот выключен, мы не сможем получить ID — выдаем ошибку
    if not numeric_chat_id:
        return jsonify({'error': 'Не удалось определить ID бота. Убедитесь, что юзербот включен и юзернейм указан верно (например, @MyBot).'}), 400

    # 3. Сохраняем автоматизацию (здесь оставляем текст, чтобы бот знал, кому писать)
    db.execute("""INSERT INTO interaction_bots 
                  (user_id, userbot_id, bot_username, custom_name, commands, interval_minutes, status) 
                  VALUES (?, ?, ?, ?, ?, ?, 'active')""",
               (user_id, userbot_id, bot_username, custom_name,
                json.dumps(data.get('commands', [])), data.get('interval_minutes', 60)))
    
    # 4. АВТОМАТИЧЕСКОЕ ДОБАВЛЕНИЕ В ПАРСЕР (здесь используем ЧИСЛОВОЙ ID!)
    try:
        db.execute("INSERT INTO tracked_chats (user_id, chat_id, chat_title, custom_name) VALUES (?, ?, ?, ?)",
                   (user_id, numeric_chat_id, bot_username, custom_name))
    except sqlite3.IntegrityError:
        pass # Уже добавлен

    db.commit()
    notify_clients()

    # 5. ПАРСИНГ ИСТОРИИ (также по числовому ID)
    # 5. ПАРСИНГ ИСТОРИИ 
    chat_title_for_parser = custom_name if custom_name else bot_username
    asyncio.run_coroutine_threadsafe(
        fetch_chat_history(client_to_use, bot_username, chat_title_for_parser, user_id, numeric_chat_id), 
        client_loop_to_use
    )

    return jsonify({'success': True})

@app.route('/api/interaction_bots/<int:id>', methods=['DELETE'])
@login_required
def delete_interaction_bot(id):
    db = get_db()
    db.execute("DELETE FROM interaction_bots WHERE id=? AND user_id=?", (id, session['user_id']))
    db.commit()
    notify_clients()
    return jsonify({'success': True})

@app.teardown_appcontext
def close_connection(exception):
    db = getattr(g, '_database', None)
    if db is not None:
        db.close()

def init_db():
    if not os.path.exists('sessions'):
        os.makedirs('sessions')
    with app.app_context():
        db = get_db()
        db.execute('''
        CREATE TABLE IF NOT EXISTS pub_markups (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    pub_id INTEGER NOT NULL,
                    folder_id INTEGER DEFAULT 0,
                    markup_type TEXT DEFAULT 'percent',
                    markup_value REAL DEFAULT 0,
                    rounding INTEGER DEFAULT 100,
                    FOREIGN KEY(pub_id) REFERENCES publications(id) ON DELETE CASCADE,
                    UNIQUE(pub_id, folder_id)
                )
    ''')
        db.execute('''
        CREATE TABLE IF NOT EXISTS publications (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT,
            is_active INTEGER DEFAULT 0,
            interval_min INTEGER DEFAULT 60,
            chat_id TEXT,
            message_id TEXT,
            userbot_id INTEGER,
            template TEXT,
            allowed_items TEXT -- Здесь будет храниться JSON с выбранными галочками (как в API)
        )
    ''')
        db.execute('''
        CREATE TABLE IF NOT EXISTS api_sources (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            name TEXT,
            url TEXT,
            token TEXT,
            interval_min INTEGER DEFAULT 60,
            is_active INTEGER DEFAULT 1
        )
    ''')
        db.executescript('''
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                login TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                role TEXT DEFAULT 'user'
            );
            CREATE TABLE IF NOT EXISTS user_telegram_sessions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                api_id TEXT,
                api_hash TEXT,
                session_file TEXT,
                status TEXT DEFAULT 'inactive',
                FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
            );
            CREATE TABLE IF NOT EXISTS messages (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                telegram_message_id INTEGER UNIQUE,
                type TEXT,
                text TEXT,
                date TIMESTAMP,
                chat_id INTEGER,
                chat_title TEXT,
                FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
            );
            CREATE TABLE IF NOT EXISTS folders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                name TEXT NOT NULL,
                parent_id INTEGER,
                FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE,
                FOREIGN KEY(parent_id) REFERENCES folders(id) ON DELETE CASCADE
            );
            CREATE TABLE IF NOT EXISTS saved_messages (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                message_id INTEGER,
                folder_id INTEGER,
                saved_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(message_id) REFERENCES messages(id) ON DELETE CASCADE,
                FOREIGN KEY(folder_id) REFERENCES folders(id) ON DELETE CASCADE,
                UNIQUE(message_id, folder_id)
            );
            CREATE TABLE IF NOT EXISTS tracked_chats (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                chat_id INTEGER NOT NULL,
                chat_title TEXT,
                added_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE,
                UNIQUE(user_id, chat_id)
            );
            CREATE TABLE IF NOT EXISTS products (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                name TEXT NOT NULL,
                synonyms TEXT,
                price REAL,
                FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
            );
            
            CREATE TABLE IF NOT EXISTS interaction_bots (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                userbot_id INTEGER NOT NULL,
                bot_username TEXT NOT NULL,
                commands TEXT NOT NULL,
                interval_minutes INTEGER NOT NULL DEFAULT 60,
                last_run TIMESTAMP,
                status TEXT DEFAULT 'active',
                FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE,
                FOREIGN KEY(userbot_id) REFERENCES user_telegram_sessions(id) ON DELETE CASCADE
            );
        ''')
        # Добавляем новые колонки в существующие таблицы (игнорируем ошибку, если они уже есть)
        # 1. Безопасное добавление колонок
        columns_to_add = [
            ("product_messages", "line_index INTEGER DEFAULT -1"),
            ("product_messages", "group_id INTEGER"),
            ("product_messages", "is_actual INTEGER DEFAULT 1"), 
            ("messages", "is_blocked INTEGER DEFAULT 0"),
            ("tracked_chats", "custom_name TEXT"),
            ("interaction_bots", "custom_name TEXT"),
            ("messages", "sender_name TEXT"),
            ("products", "folder_id INTEGER"),
            ("user_telegram_sessions", "account_name TEXT"),
            ("user_telegram_sessions", "time_start TEXT DEFAULT '00:00'"), 
            ("user_telegram_sessions", "time_end TEXT DEFAULT '23:59'"),   
            ("user_telegram_sessions", "schedule_enabled INTEGER DEFAULT 0"),
            # --- НОВЫЕ КОЛОНКИ ДЛЯ API КЛИЕНТОВ ---
            ("api_clients", "time_start TEXT DEFAULT '00:00'"),
            ("api_clients", "time_end TEXT DEFAULT '23:59'"),
            ("api_clients", "schedule_enabled INTEGER DEFAULT 0"),
            ("api_clients", "access_rules TEXT"),
            ("api_clients", "allowed_folders TEXT DEFAULT 'all'"), # <--- ДОБАВИТЬ ЭТО
            ("api_clients", "allowed_chats TEXT DEFAULT 'all'"),
            ("api_clients", "publish_chat_id INTEGER"),
            ("api_clients", "publish_message_id INTEGER"),
            ("api_clients", "publish_enabled INTEGER DEFAULT 0"),
            ("api_clients", "userbot_id INTEGER"),
            ("api_clients", "publish_template TEXT"),
            ("publications", "user_id INTEGER"),
            ("api_clients", "folders TEXT"),
            ("api_clients", "publish_interval INTEGER DEFAULT 60"),
            ("users", "session_token TEXT"),
            ("excel_configs", "is_grouped INTEGER DEFAULT 0"),
            ("publications", "markup_type TEXT DEFAULT 'percent'"), # <--- НОВОЕ
            ("publications", "markup_value REAL DEFAULT 0"),        # <--- НОВОЕ
            ("publications", "rounding INTEGER DEFAULT 100"),
            ("excel_configs", "sku_col INTEGER DEFAULT -1"),
            ("excel_configs", "source_type TEXT DEFAULT 'file'"),       # 'file' или 'google_sheet'
            ("excel_configs", "sheet_url TEXT"),                        # Ссылка на таблицу
            ("excel_configs", "parse_interval INTEGER DEFAULT 60")
        ]
        for table, col in columns_to_add:
            try:
                db.execute(f"ALTER TABLE {table} ADD COLUMN {col};")
            except sqlite3.OperationalError:
                pass # Если колонка уже есть, просто идем к следующей

        # 2. Обновление таблицы привязок (чтобы разные строки одного сообщения не затирали друг друга)
        # 2. Обновление таблицы привязок

        # --- ТАБЛИЦА ДЛЯ НАСТРОЕК EXCEL (С ПОДДЕРЖКОЙ ЛИСТОВ) ---
        db.executescript('''
            CREATE TABLE IF NOT EXISTS excel_configs_new (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                chat_id INTEGER NOT NULL,
                sheet_name TEXT NOT NULL,
                name_col INTEGER NOT NULL,
                name_row_offset INTEGER NOT NULL,
                price_col INTEGER NOT NULL,
                price_row_offset INTEGER NOT NULL,
                block_step INTEGER NOT NULL,
                start_row INTEGER NOT NULL,
                FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE,
                UNIQUE(chat_id, sheet_name)
            );
        ''')


        # Безопасный перенос старых данных (если они были) и переименование
        try:
            db.executescript('''
                INSERT OR IGNORE INTO excel_configs_new (id, user_id, chat_id, sheet_name, name_col, name_row_offset, price_col, price_row_offset, block_step, start_row)
                SELECT id, user_id, chat_id, 'Лист1', name_col, name_row_offset, price_col, price_row_offset, block_step, start_row FROM excel_configs;
                DROP TABLE excel_configs;
            ''')
        except Exception:
            pass
        try:
            db.executescript('ALTER TABLE excel_configs_new RENAME TO excel_configs;')
        except Exception:
            pass



        try:
            # --- ТАБЛИЦА ДЛЯ НАСТРОЕК EXCEL ---
            db.executescript('''
                CREATE TABLE IF NOT EXISTS excel_missing_sheets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                chat_id INTEGER,
                chat_title TEXT,
                sheet_name TEXT,
                FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE,
                UNIQUE(chat_id, sheet_name)
            );
            ''')
            db.executescript('''
                CREATE TABLE IF NOT EXISTS excel_configs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL,
                    chat_id INTEGER NOT NULL,
                    name_col INTEGER NOT NULL,
                    name_row_offset INTEGER NOT NULL,
                    price_col INTEGER NOT NULL,
                    price_row_offset INTEGER NOT NULL,
                    block_step INTEGER NOT NULL,
                    start_row INTEGER NOT NULL,
                    FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE,
                    UNIQUE(chat_id)
                );
            ''')
            # --- ТАБЛИЦЫ ДЛЯ API И НАЦЕНОК ---
            db.executescript('''
                CREATE TABLE IF NOT EXISTS api_clients (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL,
                    name TEXT NOT NULL,
                    token TEXT UNIQUE NOT NULL,
                    FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
                );
                CREATE TABLE IF NOT EXISTS api_markups (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    client_id INTEGER NOT NULL,
                    folder_id INTEGER DEFAULT 0, -- 0 означает "Для всех остальных папок"
                    markup_type TEXT DEFAULT 'percent', -- 'percent' или 'fixed'
                    markup_value REAL DEFAULT 0,
                    rounding INTEGER DEFAULT 100, -- 100, 500, 1000 и тд.
                    FOREIGN KEY(client_id) REFERENCES api_clients(id) ON DELETE CASCADE,
                    UNIQUE(client_id, folder_id)
                );
            ''')
            db.executescript('''
                CREATE TABLE IF NOT EXISTS pm_new (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    product_id INTEGER NOT NULL,
                    message_id INTEGER NOT NULL,
                    line_index INTEGER DEFAULT -1,
                    group_id INTEGER,
                    status TEXT DEFAULT 'proposed',
                    extracted_price REAL,
                    is_actual INTEGER DEFAULT 1, -- <--- ДОБАВЛЕНА ЭТА СТРОКА
                    FOREIGN KEY(product_id) REFERENCES products(id) ON DELETE CASCADE,
                    FOREIGN KEY(message_id) REFERENCES messages(id) ON DELETE CASCADE,
                    UNIQUE(product_id, message_id, line_index)
                );
                -- Переносим старые данные (добавляем 1 по умолчанию для is_actual)
                INSERT OR IGNORE INTO pm_new (id, product_id, message_id, line_index, group_id, status, extracted_price, is_actual) 
                SELECT id, product_id, message_id, line_index, group_id, status, extracted_price, 1 FROM product_messages;
                
                -- Заменяем старую таблицу на новую
                DROP TABLE IF EXISTS product_messages;
                ALTER TABLE pm_new RENAME TO product_messages;
            ''')
        except Exception as e:
            pass


        # --- ИСПРАВЛЕНИЕ БАГА "5 СООБЩЕНИЙ" (Снятие глобального UNIQUE) ---
        try:
            db.executescript('''
                CREATE TABLE IF NOT EXISTS messages_new (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER,
                    telegram_message_id INTEGER,
                    type TEXT,
                    text TEXT,
                    date TIMESTAMP,
                    chat_id INTEGER,
                    chat_title TEXT,
                    is_blocked INTEGER DEFAULT 0,
                    sender_name TEXT,
                    FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE,
                    UNIQUE(chat_id, telegram_message_id) -- Теперь уникальность только внутри одного чата!
                );
                
                -- Переносим все старые сообщения в новую правильную таблицу
                INSERT OR IGNORE INTO messages_new (id, user_id, telegram_message_id, type, text, date, chat_id, chat_title, is_blocked, sender_name)
                SELECT id, user_id, telegram_message_id, type, text, date, chat_id, chat_title, is_blocked, sender_name FROM messages;
                
                -- Удаляем сломанную таблицу и ставим на её место новую
                DROP TABLE messages;
                ALTER TABLE messages_new RENAME TO messages;
            ''')
        except Exception as e:
            logger.error(f"Ошибка миграции сообщений: {e}")

        # 3. Создание учетки администратора по умолчанию
        cursor = db.execute("SELECT COUNT(*) FROM users")
        if cursor.fetchone()[0] == 0:
            pwd_hash = hashlib.sha256('admin'.encode()).hexdigest()
            db.execute("INSERT INTO users (login, password_hash, role) VALUES (?, ?, ?)",
                       ('admin', pwd_hash, 'admin'))
            db.execute("INSERT INTO folders (user_id, name, parent_id) VALUES (1, 'По умолчанию', NULL)")
            
        db.commit()
        notify_clients()

init_db()


@app.route('/api/products/<int:prod_id>/link_folder', methods=['POST'])
@login_required
def link_product_folder(prod_id):
    data = request.get_json()
    folder_id = data.get('folder_id') # Может быть None (отвязка)
    db = get_db()
    db.execute("UPDATE products SET folder_id=? WHERE id=? AND user_id=?", (folder_id, prod_id, session['user_id']))
    db.commit()
    notify_clients()
    return jsonify({'success': True})



@app.route('/api/products/<int:prod_id>', methods=['PUT'])
@login_required
def edit_product(prod_id):
    data = request.get_json()
    db = get_db()
    db.execute("UPDATE products SET name=?, synonyms=? WHERE id=? AND user_id=?", 
               (data['name'], data['synonyms'], prod_id, session['user_id']))
    db.commit()
    notify_clients()
    return jsonify({'success': True})

@app.route('/api/access_tree')
@login_required
def get_access_tree():
    user_id = session['user_id']
    db = get_db()
    folders = [dict(f) for f in db.execute("SELECT id, name, parent_id FROM folders WHERE user_id=?", (user_id,)).fetchall()]
    
    # Гениальный запрос: вытаскиваем товары сразу с их подтвержденными поставщиками
    prods_raw = db.execute("""
        SELECT p.id as product_id, p.name as product_name, p.folder_id,
               m.chat_id, tc.custom_name, tc.chat_title
        FROM products p
        LEFT JOIN product_messages pm ON p.id = pm.product_id AND pm.status = 'confirmed'
        LEFT JOIN messages m ON pm.message_id = m.id
        LEFT JOIN tracked_chats tc ON m.chat_id = tc.chat_id AND m.user_id = tc.user_id
        WHERE p.user_id = ?
        GROUP BY p.id, m.chat_id
        ORDER BY p.name
    """, (user_id,)).fetchall()
    
    products = {}
    for row in prods_raw:
        pid = row['product_id']
        if pid not in products:
            products[pid] = {'id': pid, 'name': row['product_name'], 'folder_id': row['folder_id'], 'chats': []}
        if row['chat_id']:
            c_name = row['custom_name'] if row['custom_name'] else (row['chat_title'] or 'Неизвестный поставщик')
            if not any(c['chat_id'] == row['chat_id'] for c in products[pid]['chats']):
                products[pid]['chats'].append({'chat_id': row['chat_id'], 'name': c_name})
            
    return jsonify({'folders': folders, 'products': list(products.values())})



def fetch_google_sheet_as_df(sheet_url):
    """
    Превращает обычную ссылку на Google Таблицу в прямую ссылку на скачивание XLSX
    и сразу читает её через pandas.
    Пример: https://docs.google.com/spreadsheets/d/1ABC123.../edit#gid=0
    """
    try:
        if "/edit" in sheet_url:
            # Заменяем концовку ссылки на команду экспорта
            export_url = sheet_url.split('/edit')[0] + '/export?format=xlsx'
        else:
            export_url = sheet_url
            
        # Читаем напрямую из интернета в DataFrame
        df = pd.read_excel(export_url)
        return df
    except Exception as e:
        logger.error(f"Ошибка при загрузке Google Таблицы: {e}")
        return None

@app.route('/api/excel/google_sheet/preview', methods=['POST'])
@login_required
def preview_google_sheet_route():
    data = request.json
    sheet_url = data.get('sheet_url')
    try:
        import requests
        from io import BytesIO
        from openpyxl import load_workbook

        if "/edit" in sheet_url:
            export_url = sheet_url.split('/edit')[0] + '/export?format=xlsx'
        else:
            export_url = sheet_url
            
        resp = requests.get(export_url, timeout=30)
        if resp.status_code != 200:
            return jsonify({'error': 'Не удалось скачать таблицу'}), 400

        # Загружаем книгу ПРАВИЛЬНО (не в режиме read_only)
        xls_file = BytesIO(resp.content)
        wb = load_workbook(xls_file, data_only=True) # data_only=True читает значения, а не формулы
        
        sheets_data = {}
        # Оставляем только видимые листы
        visible_sheets = [s.title for s in wb.worksheets if s.sheet_state == 'visible']
        
        for sn in visible_sheets:
            ws = wb[sn]
            rows_list = []
            count = 0
            
            # Итерируемся по строкам напрямую через openpyxl
            for row in ws.iter_rows(values_only=True):
                # Проверяем, не скрыта ли строка
                row_idx = count + 1
                if row_idx in ws.row_dimensions and ws.row_dimensions[row_idx].hidden:
                    count += 1
                    continue
                
                # Очищаем данные от None
                clean_row = [str(cell) if cell is not None else "" for cell in row]
                rows_list.append(clean_row)
                count += 1
                
                # Ограничиваем превью 15 видимыми строками
                if len(rows_list) >= 15:
                    break
            
            sheets_data[sn] = rows_list
            
        return jsonify({'success': True, 'sheets_data': sheets_data})
    except Exception as e:
        logger.error(f"GS Preview Error: {e}")
        return jsonify({'error': str(e)}), 500



import pdfplumber

@app.route('/api/pdf/preview', methods=['POST'])
@login_required
def preview_pdf_route():
    if 'file' not in request.files:
        return jsonify({'error': 'Нет файла'}), 400
    
    file = request.files['file']
    try:
        with pdfplumber.open(file) as pdf:
            if not pdf.pages:
                return jsonify({'error': 'PDF пустой'}), 400
            
            # Извлекаем таблицу с первой страницы
            table = pdf.pages[0].extract_table()
            if not table:
                return jsonify({'error': 'Не удалось найти чёткую таблицу в PDF (возможно, нет линий сетки).'}), 400
            
            # Очищаем таблицу от пустых значений (None) и переносов строк
            cleaned_table = []
            for row in table:
                cleaned_table.append([str(cell).replace('\n', ' ') if cell is not None else "" for cell in row])
                
        # Возвращаем в формате "1 лист" (т.к. у PDF нет листов как в Excel)
        return jsonify({'success': True, 'sheets_data': {'Страница 1': cleaned_table[:30]}}) # берем первые 30 строк для превью
    except Exception as e:
        logger.error(f"PDF Preview Error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/pdf/save_config', methods=['POST'])
@login_required
def save_pdf_config():
    data = request.json
    db = get_db()
    try:
        # Сохраняем в ту же таблицу, но с меткой source_type = 'pdf'
        db.execute("""
            INSERT INTO excel_configs 
            (user_id, chat_id, sheet_name, name_col, name_row_offset, price_col, price_row_offset, block_step, start_row, is_grouped, sku_col, source_type, parse_interval) 
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'pdf', ?)
            ON CONFLICT(chat_id, sheet_name) DO UPDATE SET 
            name_col=excluded.name_col, price_col=excluded.price_col, 
            block_step=excluded.block_step, start_row=excluded.start_row,
            is_grouped=excluded.is_grouped, sku_col=excluded.sku_col, source_type='pdf'
        """, (
            session['user_id'], data['chat_id'], data['sheet_name'], 
            data['name_col'], 0, data['price_col'], data.get('price_row_offset', 0), 
            data.get('block_step', 1), data['start_row'], data.get('is_grouped', 0), 
            data.get('sku_col', -1), data.get('parse_interval', 60)
        ))
        db.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/excel/google_sheet/save_config', methods=['POST'])
@login_required
def save_gs_config():
    data = request.json
    chat_id = data.get('chat_id')
    if not chat_id:
        return jsonify({'error': 'Ошибка: Выбор чата поставщика обязателен!'}), 400
    db = get_db()
    try:
        db.execute("""
            INSERT INTO excel_configs 
            (user_id, chat_id, sheet_name, name_col, name_row_offset, price_col, price_row_offset, block_step, start_row, is_grouped, sku_col, source_type, sheet_url, parse_interval) 
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'google_sheet', ?, ?)
            ON CONFLICT(chat_id, sheet_name) DO UPDATE SET 
            name_col=excluded.name_col, name_row_offset=excluded.name_row_offset, price_col=excluded.price_col, 
            price_row_offset=excluded.price_row_offset, block_step=excluded.block_step, start_row=excluded.start_row,
            is_grouped=excluded.is_grouped, sku_col=excluded.sku_col, source_type=excluded.source_type, 
            sheet_url=excluded.sheet_url, parse_interval=excluded.parse_interval
        """, (
            session['user_id'], data['chat_id'], data['sheet_name'], 
            data['name_col'], 0, data['price_col'], data.get('price_row_offset', 0), 
            data.get('block_step', 1), data['start_row'], data.get('is_grouped', 0), 
            data.get('sku_col', -1), data['sheet_url'], data['parse_interval']
        ))
        
        # Удаляем из missing_sheets (убираем плашку "Не настроен лист")
        cid_str = str(data['chat_id'])
        core_id = cid_str[4:] if cid_str.startswith('-100') else (cid_str[1:] if cid_str.startswith('-') else cid_str)
        possible_ids = [core_id, f"-{core_id}", f"-100{core_id}"]
        for pid in possible_ids:
            if data['sheet_name'] == '*':
                db.execute("DELETE FROM excel_missing_sheets WHERE chat_id = ? AND user_id = ?", (pid, session['user_id']))
            else:
                db.execute("DELETE FROM excel_missing_sheets WHERE chat_id = ? AND sheet_name = ? AND user_id = ?", (pid, data['sheet_name'], session['user_id']))
        
        db.commit()
        return jsonify({'success': True})
    except Exception as e:
        logger.error(f"Ошибка сохранения конфига Google Sheet: {e}")
        return jsonify({'error': str(e)}), 500


def format_price_list(template, products):
    """Вставляет товары в шаблон между маркерами START_PRICE_LIST и END_PRICE_LIST."""
    lines = []
    for p in products:
        lines.append(f"• {p['name']} — {p['price']} руб.")
    price_list = "\n".join(lines)
    
    if "<!-- START_PRICE_LIST -->" in template and "<!-- END_PRICE_LIST -->" in template:
        # Заменяем только часть между маркерами
        parts = template.split("<!-- START_PRICE_LIST -->")
        before = parts[0]
        after = parts[1].split("<!-- END_PRICE_LIST -->")[1]
        return before + "<!-- START_PRICE_LIST -->\n" + price_list + "\n<!-- END_PRICE_LIST -->" + after
    else:
        # Иначе заменяем всё сообщение
        return price_list

# Словарь в памяти для отслеживания времени последней публикации
# Словарь в памяти для отслеживания времени последней публикации
last_publish_times = {}

def publish_scheduler():
    """Фоновый поток для публикации каталогов в Telegram."""
    time.sleep(10)
    with app.app_context():
        while True:
            try:
                db = get_db()
                # Берем все активные публикации
                pubs = db.execute("SELECT * FROM publications WHERE is_active = 1").fetchall()
                now = datetime.now()
                
                for pub in pubs:
                    pub_id = pub['id']
                    interval_minutes = pub['interval_min'] or 60
                    
                    # 1. ПРОВЕРКА ИНТЕРВАЛА
                    last_run = last_publish_times.get(pub_id)
                    if last_run:
                        diff_minutes = (now - last_run).total_seconds() / 60
                        if diff_minutes < interval_minutes:
                            continue  # Время публикации еще не пришло
                    
                    # 2. ПОЛУЧЕНИЕ ЦЕН И ТОВАРОВ
                    userbot_id = pub['userbot_id']
                    if userbot_id not in user_clients: continue
                    
                    # Извлекаем user_id из юзербота
                    _, tg_client, user_id = user_clients[userbot_id] 
                    
                    allowed_items = {}
                    try: allowed_items = json.loads(pub['allowed_items']) 
                    except: pass
                    
                    if not allowed_items: continue
                    
                    # ---- НОВОЕ: ДОСТАЕМ НАЦЕНКИ ----
                    markups_raw = db.execute("SELECT folder_id, markup_type, markup_value, rounding FROM pub_markups WHERE pub_id=?", (pub_id,)).fetchall()
                    markups = {}
                    default_markup = {'markup_type': 'percent', 'markup_value': 0, 'rounding': 100}
                    for m in markups_raw:
                        if m['folder_id'] == 0: default_markup = dict(m)
                        else: markups[m['folder_id']] = dict(m)
                    # --------------------------------
                    
                    final_products = []
                    
                    # Перебираем товары и разрешенных поставщиков из JSON
                    for p_id_str, allowed_suppliers in allowed_items.items():
                        if not allowed_suppliers: 
                            continue
                        
                        placeholders = ','.join(['?'] * len(allowed_suppliers))
                        query_params = [int(p_id_str)] + allowed_suppliers
                        
                        # ВАЖНО: Добавлено p.folder_id в SELECT
                        price_row = db.execute(f"""
                            SELECT pm.extracted_price, p.name, p.folder_id 
                            FROM product_messages pm 
                            JOIN messages m ON pm.message_id = m.id 
                            JOIN products p ON p.id = pm.product_id
                            WHERE pm.product_id = ? 
                              AND pm.status = 'confirmed' 
                              AND pm.is_actual = 1 
                              AND m.chat_id IN ({placeholders})
                            ORDER BY m.date DESC LIMIT 1
                        """, query_params).fetchone()
                        
                        if price_row and price_row['extracted_price'] is not None:
                            # -------- ПРИМЕНЯЕМ НАЦЕНКУ ПО ПАПКАМ ---------
                            base_price = float(price_row['extracted_price'])
                            f_id = price_row['folder_id'] or 0
                            mk = markups.get(f_id, default_markup)
                            
                            if mk['markup_type'] == 'percent':
                                final_price = base_price * (1 + mk['markup_value'] / 100)
                            else:
                                final_price = base_price + mk['markup_value']
                                
                            rnd = mk['rounding']
                            if rnd > 0:
                                final_price = math.ceil(final_price / rnd) * rnd

                            final_products.append({
                                'name': price_row['name'], 
                                'price': int(final_price)
                            })
                    
                    if not final_products: continue

                    template = pub['template'] or "🔥 Актуальные цены:\n\n{prices}\n\nДля заказа пишите менеджеру!"
                    
                    # 1. Красиво собираем товары в текстовый список
                    prices_text = "\n".join([f"▪️ {p['name']} — {p['price']} ₽" for p in final_products])
                    
                    # 2. Подставляем список вместо метки {prices}
                    if "{prices}" in template:
                        message_text = template.replace("{prices}", prices_text)
                    else:
                        # Если вы случайно забыли написать {prices} в шаблоне, 
                        # бот просто аккуратно приклеит прайс-лист в самый низ сообщения
                        message_text = f"{template}\n\n{prices_text}"

                    loop = user_loops.get(userbot_id)
                    if not loop or not loop.is_running(): continue

                    chat_id = pub['chat_id']
                    message_id = pub['message_id']
                    
                    # 3. ОТПРАВКА В ТЕЛЕГРАМ
                    # 3. ОТПРАВКА В ТЕЛЕГРАМ
                    # 3. ОТПРАВКА В ТЕЛЕГРАМ
                    async def update_or_create(cid, mid, txt, p_id):
                        try:
                            cid = int(cid)
                        except ValueError:
                            pass

                        # --- ЛОГИКА РАЗДЕЛЕНИЯ ТЕКСТА ---
                        parts = []
                        max_len = 4000 # Оставляем небольшой запас до 4096
                        temp_txt = txt
                        
                        while len(temp_txt) > max_len:
                            # Ищем последний перенос строки в пределах лимита, чтобы не рвать строку пополам
                            split_idx = temp_txt.rfind('\n', 0, max_len)
                            if split_idx == -1:
                                split_idx = max_len # Если нет переносов, режем жестко
                                
                            parts.append(temp_txt[:split_idx])
                            temp_txt = temp_txt[split_idx:].strip()
                            
                        if temp_txt:
                            parts.append(temp_txt)
                        # --------------------------------

                        if mid:
                            try:
                                # ВАЖНЫЙ НЮАНС ПРИ РЕДАКТИРОВАНИИ: 
                                # Telegram не позволяет превратить 1 старое сообщение в 2 новых.
                                # Поэтому при редактировании обновится только первая часть, а остаток обрежется.
                                await tg_client.edit_message(cid, int(mid), parts[0])
                                if len(parts) > 1:
                                    logger.warning(f"Прайс слишком большой для обновления одного сообщения (ID {mid}). Хвост обрезан.")
                            except Exception as e:
                                logger.error(f"Ошибка редактирования: {e}")
                        else:
                            try:
                                # В режиме "отправки нового" просто шлем все куски друг за другом
                                first_msg = None
                                for part in parts:
                                    msg = await tg_client.send_message(cid, part)
                                    if not first_msg:
                                        first_msg = msg # Запоминаем только самое первое сообщение

                                # Сохраняем в базу ID первого сообщения (если захотите его потом удалять/редактировать)
                                if first_msg:
                                    with app.app_context():
                                        local_db = get_db()
                                        local_db.execute("UPDATE publications SET message_id = ? WHERE id = ?", (str(first_msg.id), p_id))
                                        local_db.commit()
                            except Exception as e:
                                logger.error(f"Ошибка отправки разделенного сообщения: {e}")
                                
                    asyncio.run_coroutine_threadsafe(update_or_create(chat_id, message_id, message_text, pub_id), loop)
                    last_publish_times[pub_id] = now

            except Exception as e:
                logging.error(f"Ошибка в планировщике публикации: {e}")
            time.sleep(60)



def get_catalog_for_client(client_id):
    """Возвращает каталог для указанного клиента (как в /api/v1/catalog)."""
    db = get_db()
    client = db.execute("SELECT * FROM api_clients WHERE id=?", (client_id,)).fetchone()
    if not client:
        return None

    user_id = client['user_id']
    db = get_db()
    # Получаем клиента (теперь забираем всю строку со всеми настройками)
    
        
    if client['schedule_enabled']:
        from datetime import timedelta # На всякий случай импортируем, если не импортировано
        now = (datetime.utcnow() + timedelta(hours=3)).time()
        start_time = datetime.strptime(client['time_start'], '%H:%M').time()
        end_time = datetime.strptime(client['time_end'], '%H:%M').time()
        
        is_active = start_time <= now <= end_time if start_time <= end_time else (now >= start_time or now <= end_time)
        if not is_active: return {'categories': [], 'products': []}
        
    user_id = client['user_id']
    client_id = client['id']
    
    # Читаем доступы
    allowed_folders = client['allowed_folders'] if 'allowed_folders' in client.keys() else 'all'
    allowed_chats = client['allowed_chats'] if 'allowed_chats' in client.keys() else 'all'

    folders_list = allowed_folders.split(',') if allowed_folders not in ('all', '', None) else allowed_folders
    chats_list = allowed_chats.split(',') if allowed_chats not in ('all', '', None) else allowed_chats

    markups_raw = db.execute("SELECT folder_id, markup_type, markup_value, rounding FROM api_markups WHERE client_id=?", (client_id,)).fetchall()
    markups = {}
    default_markup = {'markup_type': 'percent', 'markup_value': 0, 'rounding': 100}
    for m in markups_raw:
        if m['folder_id'] == 0: default_markup = dict(m)
        else: markups[m['folder_id']] = dict(m)
            
    # ДОБАВЛЕНО: latest_chat_id (извлекаем чат, из которого была подтверждена последняя цена)
    products_raw = db.execute("""
        SELECT p.id, p.name, p.folder_id, p.price as manual_price,
               (SELECT pm.extracted_price FROM product_messages pm JOIN messages m ON pm.message_id = m.id WHERE pm.product_id = p.id AND pm.status = 'confirmed' ORDER BY m.date DESC LIMIT 1) as parsed_price,
               (SELECT pm.is_actual FROM product_messages pm JOIN messages m ON pm.message_id = m.id WHERE pm.product_id = p.id AND pm.status = 'confirmed' ORDER BY m.date DESC LIMIT 1) as is_actual,
               (SELECT m.chat_id FROM product_messages pm JOIN messages m ON pm.message_id = m.id WHERE pm.product_id = p.id AND pm.status = 'confirmed' ORDER BY m.date DESC LIMIT 1) as latest_chat_id
        FROM products p
        WHERE p.user_id = ?
    """, (user_id,)).fetchall()
    
    products = []
    active_folder_ids = set()
    
   # Перед циклом распаковываем JSON
    try:
        access_rules = json.loads(client['access_rules']) if 'access_rules' in client.keys() and client['access_rules'] else {}
    except:
        access_rules = {}
        
    for p in products_raw:
        if p['is_actual'] == 0: continue
            
        # --- ТОЧЕЧНЫЙ ФИЛЬТР ПО ДЕРЕВУ ---
        if not access_rules: continue # Если клиенту ничего не настроено - отдаем пустой прайс
        
        pid_str = str(p['id'])
        if pid_str not in access_rules: continue # Товар галочкой не отмечен
        
        allowed_chats = access_rules[pid_str]
        chat_id_str = str(p['latest_chat_id'])
        
        # Если latest_chat_id не пустой, проверяем, разрешен ли этот поставщик
        if chat_id_str != 'None':
            if allowed_chats != ['all'] and chat_id_str not in allowed_chats: 
                continue # Цена от этого поставщика запрещена
        # ---------------------------------
        
        base_price = float(p['parsed_price'] if p['parsed_price'] is not None else (p['manual_price'] or 0))
        
   
            
        mk = markups.get(p['folder_id'], default_markup)
        
        if mk['markup_type'] == 'percent':
            final_price = base_price * (1 + mk['markup_value'] / 100)
        else:
            final_price = base_price + mk['markup_value']
            
        rnd = mk['rounding']
        if rnd > 0:
            final_price = math.ceil(final_price / rnd) * rnd
            
        products.append({
            'id': p['id'],
            'name': p['name'],
            'category_id': p['folder_id'],
            'price': int(final_price)
        })
        
        if p['folder_id']:
            active_folder_ids.add(p['folder_id'])
            
    # 3. Получаем категории и убираем пустые (Требование 1)
    folders_raw = db.execute("SELECT id, name, parent_id FROM folders WHERE user_id=?", (user_id,)).fetchall()
    
    # Собираем всех родителей для активных папок, чтобы не сломать дерево
    parent_map = {f['id']: f['parent_id'] for f in folders_raw}
    folders_to_keep = set(active_folder_ids)
    
    for fid in active_folder_ids:
        current = fid
        while current in parent_map and parent_map[current] is not None:
            current = parent_map[current]
            folders_to_keep.add(current)
            
    categories = [{'id': f['id'], 'name': f['name'], 'parent_id': f['parent_id']} 
                  for f in folders_raw if f['id'] in folders_to_keep]
        
    return {
    'categories': categories,
    'products': products
}
  

def google_sheets_scheduler():
    """Фоновый поток для парсинга Гугл Таблиц по расписанию с учетом скрытых строк"""
    import requests
    from io import BytesIO
    from openpyxl import load_workbook
    
    time.sleep(5)
    last_parsed = {}
    
    with app.app_context():
        while True:
            try:
                db = get_db()
                # Берем все конфиги для Google Таблиц
                configs = db.execute("SELECT * FROM excel_configs WHERE source_type = 'google_sheet'").fetchall()
                current_time = time.time()
                
                # Группируем по URL, чтобы не скачивать один и тот же файл многократно
                url_configs = {}
                for c in configs:
                    url = c['sheet_url']
                    if url not in url_configs: url_configs[url] = []
                    url_configs[url].append(c)

                for url, conf_list in url_configs.items():
                    # Проверяем интервал по первому конфигу в группе
                    first_conf = conf_list[0]
                    cid = first_conf['id']
                    interval_sec = int(first_conf['parse_interval'] or 60) * 60
                    
                    if cid not in last_parsed or (current_time - last_parsed[cid]) >= interval_sec:
                        logger.info(f"Парсинг Гугл Таблицы {url} по расписанию...")
                        
                        try:
                            # Формируем ссылку на экспорт
                            export_url = url.split('/edit')[0] + '/export?format=xlsx' if "/edit" in url else url
                            resp = requests.get(export_url, timeout=30)
                            if resp.status_code != 200:
                                logger.error(f"Не удалось скачать таблицу: {resp.status_code}")
                                continue
                            
                            # Загружаем книгу через openpyxl (БЕЗ read_only, чтобы видеть скрытые строки)
                            xls_file = BytesIO(resp.content)
                            wb = load_workbook(xls_file, data_only=True)
                        except Exception as e:
                            logger.error(f"Ошибка загрузки книги: {e}")
                            continue

                        # Определяем видимые листы
                        visible_sheets = [s.title for s in wb.worksheets if s.sheet_state == 'visible']

                        for config in conf_list:
                            c_dict = dict(config) # Конвертируем Row в dict
                            sheet_target = c_dict['sheet_name']
                            user_id = c_dict['user_id']
                            chat_id = c_dict['chat_id']
                            
                            # Пропускаем, если конкретный лист скрыт
                            if sheet_target != '*' and sheet_target not in visible_sheets:
                                continue 
                                
                            sheets_to_parse = visible_sheets if sheet_target == '*' else [sheet_target]
                            
                            parsed_lines = []
                            for sn in sheets_to_parse:
                                if sn not in wb.sheetnames: continue
                                ws = wb[sn]
                                
                                # Превращаем данные листа в список строк для удобства
                                # Индексация в all_rows будет с 0, а в Excel (row_dimensions) с 1
                                all_rows = list(ws.iter_rows(values_only=True))
                                
                                step = int(c_dict.get('block_step', 1))
                                start = int(c_dict.get('start_row', 0))
                                n_col = int(c_dict.get('name_col', 0))
                                p_col = int(c_dict.get('price_col', 1))
                                n_off = int(c_dict.get('name_row_offset', 0))
                                p_off = int(c_dict.get('price_row_offset', 0))
                                is_grouped = c_dict.get('is_grouped', 0)
                                sku_col = c_dict.get('sku_col', -1)

                                if is_grouped == 1:
                                    # --- ЛОГИКА ДЛЯ СГРУППИРОВАННЫХ ТАБЛИЦ ---
                                    current_group = ""
                                    for i in range(start, len(all_rows)):
                                        # ПРОВЕРКА: Скрыта ли строка? (openpyxl индексирует с 1)
                                        if i + 1 in ws.row_dimensions and ws.row_dimensions[i + 1].hidden:
                                            continue
                                            
                                        row = all_rows[i]
                                        col_n_val = str(row[n_col]).strip() if n_col < len(row) and row[n_col] is not None else ""
                                        col_sku_val = str(row[sku_col]).strip() if sku_col != -1 and sku_col < len(row) and row[sku_col] is not None else ""
                                        col_p_val = str(row[p_col]).strip() if p_col < len(row) and row[p_col] is not None else ""
                                        
                                        has_price = (col_p_val and col_p_val.lower() != 'none' and col_p_val.lower() != 'nan')
                                        
                                        # Проверка цены
                                        extracted = extract_price(col_p_val) if has_price else None
                                        if extracted is not None and extracted <= 0: continue

                                        if n_col == sku_col:
                                            if col_n_val and not has_price: current_group = col_n_val
                                            elif col_n_val and has_price and current_group:
                                                parsed_lines.append(f"{current_group} {col_n_val} - {col_p_val}")
                                        else:
                                            if col_n_val and col_n_val.lower() != 'none' and not has_price: 
                                                current_group = col_n_val
                                            if col_sku_val and has_price and current_group:
                                                parsed_lines.append(f"{current_group} {col_sku_val} - {col_p_val}")
                                else:
                                    # --- ЛОГИКА ДЛЯ ОБЫЧНЫХ ТАБЛИЦ ---
                                    for i in range(start, len(all_rows), step):
                                        # Проверка на скрытость основной строки
                                        if i + 1 in ws.row_dimensions and ws.row_dimensions[i + 1].hidden:
                                            continue
                                            
                                        name_idx = i + n_off
                                        price_idx = i + p_off
                                        
                                        if name_idx < len(all_rows) and price_idx < len(all_rows):
                                            # Проверяем на скрытость строки со смещением
                                            if (name_idx + 1 in ws.row_dimensions and ws.row_dimensions[name_idx + 1].hidden) or \
                                               (price_idx + 1 in ws.row_dimensions and ws.row_dimensions[price_idx + 1].hidden):
                                                continue
                                                
                                            row_n = all_rows[name_idx]
                                            row_p = all_rows[price_idx]
                                            
                                            n_val = str(row_n[n_col]).strip() if n_col < len(row_n) and row_n[n_col] is not None else ""
                                            p_val = str(row_p[p_col]).strip() if p_col < len(row_p) and row_p[p_col] is not None else ""
                                            
                                            if n_val and p_val and n_val.lower() != 'none' and p_val.lower() != 'none':
                                                extracted = extract_price(p_val)
                                                if extracted is not None and extracted > 0:
                                                    parsed_lines.append(f"{n_val} - {p_val}")

                            if parsed_lines:
                                msg_text = "📊 [GOOGLE ТАБЛИЦА]:\n" + "\n".join(parsed_lines)
                                tc = db.execute("SELECT custom_name, chat_title FROM tracked_chats WHERE chat_id = ? AND user_id = ?", (chat_id, user_id)).fetchone()
                                chat_title = tc['custom_name'] if tc and tc['custom_name'] else (tc['chat_title'] if tc else str(chat_id))
                                
                                now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                                msg_type = f"excel_{c_dict['id']}" 
                                
                                # Сохраняем сообщение в базу
                                cursor = db.execute(
                                    "INSERT INTO messages (user_id, type, text, date, chat_id, chat_title, sender_name) VALUES (?, ?, ?, ?, ?, ?, ?)",
                                    (user_id, msg_type, msg_text, now_str, chat_id, chat_title, f"Лист: {sheet_target}")
                                )
                                new_msg_id = cursor.lastrowid
                                
                                # --- МИГРАЦИЯ ПРИВЯЗОК ---
                                old_bindings = db.execute("""
                                    SELECT pm.id, pm.line_index, p.synonyms
                                    FROM product_messages pm
                                    JOIN products p ON pm.product_id = p.id
                                    JOIN messages m ON pm.message_id = m.id
                                    WHERE m.chat_id = ? AND m.user_id = ? AND m.id != ? AND m.type = ?
                                """, (chat_id, user_id, new_msg_id, msg_type)).fetchall()

                                if old_bindings:
                                    lines = msg_text.split('\n')
                                    for b in old_bindings:
                                        synonyms = [s.strip().lower() for s in (b['synonyms'] or "").split(',') if s.strip()]
                                        match_found = False
                                        new_price = None
                                        new_line_idx = 0
                                        
                                        for i, line in enumerate(lines):
                                            if any(syn in line.lower() for syn in synonyms):
                                                match_found = True
                                                new_price = extract_price(line)
                                                new_line_idx = i
                                                break
                                        
                                        if match_found and new_price and new_price > 0:
                                            db.execute("UPDATE product_messages SET message_id = ?, line_index = ?, extracted_price = ?, is_actual = 1 WHERE id = ?", (new_msg_id, new_line_idx, new_price, b['id']))
                                        else:
                                            db.execute("UPDATE product_messages SET is_actual = 0 WHERE id = ?", (b['id'],))
                                            
                                # Чистим старые сообщения этого листа
                                db.execute("DELETE FROM messages WHERE chat_id = ? AND user_id = ? AND id != ? AND type = ?", (chat_id, user_id, new_msg_id, msg_type))
                                
                                # --- АВТО-ПРИВЯЗКА НОВЫХ ---
                                all_prods = db.execute("SELECT id, synonyms FROM products WHERE user_id=?", (user_id,)).fetchall()
                                for idx, line in enumerate(parsed_lines):
                                    ext_p = extract_price(line)
                                    if not ext_p or ext_p <= 0: continue
                                    line_l = line.lower()
                                    for prod in all_prods:
                                        syns = [s.strip().lower() for s in (prod['synonyms'] or '').split(',') if s.strip()]
                                        if syns and any(syn in line_l for syn in syns):
                                            exists = db.execute("SELECT id FROM product_messages WHERE product_id=? AND message_id=?", (prod['id'], new_msg_id)).fetchone()
                                            if not exists:
                                                db.execute("INSERT INTO product_messages (product_id, message_id, extracted_price, status, is_actual, line_index) VALUES (?, ?, ?, 'pending', 1, ?)", (prod['id'], new_msg_id, ext_p, idx + 1))
                                            break
                                            
                                db.commit()
                                notify_clients()
                                logger.info(f"✅ Успешно: Лист {sheet_target} из {url}")
                        
                        # Запоминаем время последнего парсинга
                        for config in conf_list:
                            last_parsed[config['id']] = current_time

            except Exception as e:
                logger.error(f"Глобальная ошибка планировщика GS: {e}", exc_info=True)
                
            time.sleep(60) # Проверка раз в минуту

# Запускаем поток вместе с сервером
threading.Thread(target=google_sheets_scheduler, daemon=True).start()



@app.route('/api/messages/<int:msg_id>/block', methods=['POST'])
@login_required
def block_message(msg_id):
    db = get_db()
    # Устанавливаем флаг блокировки для сообщения
    db.execute("UPDATE messages SET is_blocked = 1 WHERE id = ? AND user_id = ?", (msg_id, session['user_id']))
    db.commit()
    notify_clients()
    return jsonify({'success': True})


@app.route('/api/product_messages/merge', methods=['POST'])
@login_required
def merge_product_messages():
    data = request.get_json()
    pm_ids = data.get('pm_ids', [])
    if len(pm_ids) < 2:
        return jsonify({'error': 'Выберите минимум 2 сообщения'}), 400
    
    db = get_db()
    # Берем минимальный ID как главный номер группы
    group_id = min(map(int, pm_ids))
    placeholders = ','.join('?' for _ in pm_ids)
    
    db.execute(f"UPDATE product_messages SET group_id = ? WHERE id IN ({placeholders})", [group_id] + pm_ids)
    db.commit()
    notify_clients()
    return jsonify({'success': True})


@app.route('/api/products/<int:source_id>/merge', methods=['POST'])
@login_required
def merge_product(source_id):
    target_id = request.get_json()['target_id']
    user_id = session['user_id']
    db = get_db()
    
    # Получаем синонимы обоих товаров
    src = db.execute("SELECT synonyms FROM products WHERE id=? AND user_id=?", (source_id, user_id)).fetchone()
    tgt = db.execute("SELECT synonyms FROM products WHERE id=? AND user_id=?", (target_id, user_id)).fetchone()
    
    if src and tgt:
        src_syns = [s.strip() for s in src['synonyms'].split(',') if s.strip()]
        tgt_syns = [s.strip() for s in tgt['synonyms'].split(',') if s.strip()]
        combined = list(set(src_syns + tgt_syns)) # Объединяем без дубликатов
        
        # Обновляем синонимы целевого товара
        db.execute("UPDATE products SET synonyms=? WHERE id=?", (', '.join(combined), target_id))
        
        # Переносим привязанные сообщения (игнорируем ошибку, если такое сообщение уже привязано)
        try:
            db.execute("UPDATE OR IGNORE product_messages SET product_id=? WHERE product_id=?", (target_id, source_id))
        except:
            pass
            
        # Удаляем старый товар
        db.execute("DELETE FROM products WHERE id=?", (source_id,))
        db.commit()
        notify_clients()
    return jsonify({'success': True})

import re
@app.route('/api/profile/change_password', methods=['POST'])
@login_required
def change_password():
    user_id = session['user_id']
    data = request.get_json()
    
    old_password = data.get('old_password')
    new_password = data.get('new_password')

    if not old_password or not new_password:
        return jsonify({'error': 'Заполните все поля'}), 400

    db = get_db()
    
    # Хэшируем введённый старый пароль и проверяем его
    old_pwd_hash = hashlib.sha256(old_password.encode()).hexdigest()
    user = db.execute("SELECT id FROM users WHERE id = ? AND password_hash = ?", (user_id, old_pwd_hash)).fetchone()

    if not user:
        return jsonify({'error': 'Неверный старый пароль'}), 400

    # Если всё верно, хэшируем новый пароль и записываем в БД 
    new_pwd_hash = hashlib.sha256(new_password.encode()).hexdigest()
    db.execute("UPDATE users SET password_hash = ? WHERE id = ?", (new_pwd_hash, user_id))
    db.commit()

    return jsonify({'success': True})


async def fetch_chat_history(client, target_entity, chat_title, user_id, db_chat_id=None):
    """Парсинг старых сообщений с подробной статистикой"""
    if db_chat_id is None:
        db_chat_id = getattr(target_entity, 'id', target_entity)

    try:
        with app.app_context():
            db = get_db()
            added = 0
            duplicates = 0
            no_text = 0
            errors = 0
            
            async for message in client.iter_messages(target_entity): 
                # Пропускаем пустые (например, просто фото без подписи)
                if (not message.text and not message.document):
                    no_text += 1
                    continue
                
                try:
                    sender = await message.get_sender()
                    sender_name = f"{getattr(sender, 'first_name', '')} {getattr(sender, 'last_name', '')}".strip() or "Unknown"
                    msg_type = 'channel' if message.is_channel else 'group' if message.is_group else 'private'
                    
                    parsed_text = await parse_excel_message(client, message, db_chat_id, user_id)
                    if not parsed_text:
                        no_text += 1
                        continue

                    # ИСПРАВЛЕНИЕ: Форматируем дату вручную, чтобы не было DeprecationWarning
                    msg_date = message.date.strftime("%Y-%m-%d %H:%M:%S")

                    db.execute(
                        'INSERT INTO messages (user_id, telegram_message_id, type, text, date, chat_id, chat_title, sender_name) VALUES (?, ?, ?, ?, ?, ?, ?, ?)',
                        (user_id, message.id, msg_type, parsed_text, msg_date, db_chat_id, chat_title, sender_name)
                    )
                    added += 1
                    
                    if added % 50 == 0:
                        db.commit()
                        notify_clients()
                        
                except sqlite3.IntegrityError:
                    duplicates += 1  # Сообщение уже есть в базе
                except Exception as msg_e:
                    errors += 1
                    logger.warning(f"Ошибка в сообщении {message.id}: {msg_e}")
                    
            db.commit()
            notify_clients()
            
            # ВЫВОДИМ ПОДРОБНЫЙ ОТЧЕТ В КОНСОЛЬ
            logger.info(f"Парсинг [{chat_title}]: Добавлено {added} | Уже были в базе: {duplicates} | Без текста/фото: {no_text} | Ошибок: {errors}")
            
    except Exception as e:
        logger.error(f"Глобальная ошибка парсинга истории для {chat_title}: {e}")

@app.route('/api/api_clients/<int:client_id>/publish', methods=['POST'])
@login_required
def update_publish_settings(client_id):
    data = request.json
    db = get_db()
    # Проверяем, что клиент принадлежит текущему пользователю
    client = db.execute("SELECT id FROM api_clients WHERE id=? AND user_id=?", 
                        (client_id, session['user_id'])).fetchone()
    if not client:
        return jsonify({'error': 'Клиент не найден'}), 404

    db.execute("""
        UPDATE api_clients 
        SET publish_enabled=?, publish_chat_id=?, userbot_id=?, publish_template=?, folders=?, publish_interval=?
        WHERE id=?
    """, (data.get('publish_enabled', 0), data.get('publish_chat_id'), data.get('userbot_id'),
          data.get('publish_template'), data.get('folders'), data.get('publish_interval', 60), client_id))
    db.commit()
    return jsonify({'success': True})

# ---------- Декораторы авторизации ----------
def login_required(f):
    @functools.wraps(f)
    def wrapped(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return wrapped

def admin_required(f):
    @functools.wraps(f)
    def wrapped(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login_page'))
        user_id = session['user_id']
        db = get_db()
        cursor = db.execute("SELECT role FROM users WHERE id = ?", (user_id,))
        row = cursor.fetchone()
        if not row or row['role'] != 'admin':
            return "Access denied", 403
        return f(*args, **kwargs)
    return wrapped

# ---------- Фоновый слушатель сообщений ----------
def stop_message_listener(session_id):
    client_info = user_clients.get(session_id)
    if client_info:
        _, client, _ = client_info
        loop = user_loops.get(session_id)
        if loop and loop.is_running():
            async def safe_disconnect():
                try:
                    await client.disconnect()
                except:
                    pass
            try:
                asyncio.run_coroutine_threadsafe(safe_disconnect(), loop)
            except:
                pass

    # ВАЖНО: Мы БОЛЬШЕ НЕ удаляем user_clients.pop() здесь!
    # Это сделает сам фоновый поток перед своей смертью.

    try:
        with app.app_context():
            db = get_db()
            db.execute("UPDATE user_telegram_sessions SET status = 'inactive' WHERE id = ?", (session_id,))
            db.commit()
            if 'notify_clients' in globals(): notify_clients()
    except Exception as e:
        logging.error(f"Ошибка БД при остановке сессии {session_id}: {e}")

@app.route('/logo.svg')
def serve_logo():
    # Отдаем файл logo.svg прямо из корня проекта
    return send_from_directory(os.getcwd(), 'logo.svg')

# Было: @app.route('/api/excel/configs/chat/<int:chat_id>', methods=['DELETE'])
@app.route('/api/excel/configs/chat/<chat_id>', methods=['DELETE'])
def delete_excel_chat_configs(chat_id):
    try:
        chat_id = int(chat_id)
        user_id = session.get('user_id')
        db = get_db()
        
        # 1. Удаляем привязки к товарам
        db.execute("""
            DELETE FROM product_messages 
            WHERE message_id IN (SELECT id FROM messages WHERE chat_id = ? AND user_id = ? AND type LIKE 'excel%')
        """, (chat_id, user_id))
        
        # 2. Удаляем сами сообщения (прайс-листы)
        db.execute("DELETE FROM messages WHERE chat_id = ? AND user_id = ? AND type LIKE 'excel%'", (chat_id, user_id))
        
        # 3. Удаляем конфигурацию парсера
        db.execute("DELETE FROM excel_configs WHERE chat_id = ? AND user_id = ?", (chat_id, user_id))
        
        db.commit()
        return jsonify({'success': True})
    except Exception as e:
        print(f"Ошибка при удалении файла: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/publications/<int:pub_id>/markups', methods=['GET', 'POST'])
@login_required
def manage_pub_markups(pub_id):
    db = get_db()
    if request.method == 'GET':
        markups = db.execute("""
            SELECT pm.*, f.name as folder_name 
            FROM pub_markups pm 
            LEFT JOIN folders f ON pm.folder_id = f.id 
            WHERE pm.pub_id=?
        """, (pub_id,)).fetchall()
        return jsonify([dict(m) for m in markups])
    
    data = request.json
    folder_id = data.get('folder_id') or 0
    db.execute("""
        INSERT INTO pub_markups (pub_id, folder_id, markup_type, markup_value, rounding) 
        VALUES (?, ?, ?, ?, ?)
        ON CONFLICT(pub_id, folder_id) DO UPDATE SET 
        markup_type=excluded.markup_type, markup_value=excluded.markup_value, rounding=excluded.rounding
    """, (pub_id, folder_id, data['markup_type'], data['markup_value'], data['rounding']))
    db.commit()
    return jsonify({'success': True})

@app.route('/api/pub_markups/<int:markup_id>', methods=['DELETE'])
@login_required
def delete_pub_markup(markup_id):
    db = get_db()
    db.execute("DELETE FROM pub_markups WHERE id=? AND folder_id != 0", (markup_id,))
    db.commit()
    return jsonify({'success': True})


def start_message_listener(session_id, user_id, api_id, api_hash):
    with listener_locks.setdefault(session_id, threading.Lock()):
        if session_id in background_tasks:
            logging.info(f"Слушатель для сессии {session_id} уже запущен")
            return

        def listener():
            loop = asyncio.new_event_loop()
         
            asyncio.set_event_loop(loop)
            session_file = f'sessions/user_{user_id}_{session_id}'
            client = TelegramClient(session_file, int(api_id), api_hash)
            
            user_clients[session_id] = (threading.current_thread(), client, user_id)
            user_loops[session_id] = loop

           

            # --- НОВЫЙ БЛОК: СЛУШАЕМ ИЗМЕНЕНИЯ В СООБЩЕНИЯХ ---
           # --- НОВЫЙ БЛОК: СЛУШАЕМ ИЗМЕНЕНИЯ В СООБЩЕНИЯХ ---
           

# --- НОВЫЙ БЛОК: СЛУШАЕМ УДАЛЕНИЯ ---
            @client.on(MessageDeleted)
            async def deleted_handler(event):
                with app.app_context():
                    db = get_db()
                    for msg_id in event.deleted_ids:
                        # УБРАЛИ проверку chat_id, так как в Telethon при удалении он часто равен None
                        db.execute("""
                            UPDATE product_messages 
                            SET is_actual = 0 
                            WHERE message_id IN (SELECT id FROM messages WHERE telegram_message_id = ?)
                        """, (msg_id,))
                    db.commit()
                    notify_clients()

            @client.on(events.MessageEdited)
            async def edit_handler(event):
                message = event.message
                if (not message.text and not message.document):
                    return
                
                chat = await event.get_chat()
                
                # --- ИГНОРИРУЕМ ИСХОДЯЩИЕ ТОЛЬКО ДЛЯ БОТОВ ---
                if getattr(message, 'out', False) and event.is_private and getattr(chat, 'bot', False):
                    return
                # ---------------------------------------------
                
                chat_id = chat.id
                
                # --- ЖЕСТКАЯ ИЗОЛЯЦИЯ ПАРСИНГА БОТОВ (ДЛЯ РЕДАКТИРОВАНИЯ) ---
                if event.is_private:
                    sender_entity = await event.get_sender()
                    if getattr(sender_entity, 'bot', False):
                        bot_un = getattr(sender_entity, 'username', '')
                        if not bot_un:
                            try:
                                entity = await client.get_entity(chat_id)
                                bot_un = getattr(entity, 'username', '')
                            except: pass
                        
                        bot_un_clean = bot_un.lower().replace('@', '') if bot_un else ''
                        bot_un_at = f"@{bot_un_clean}" if bot_un_clean else ''
                        
                        with app.app_context():
                            local_db = get_db()
                            check_bot = local_db.execute(
                                "SELECT id FROM interaction_bots WHERE user_id = ? AND userbot_id = ? AND (LOWER(bot_username) IN (?, ?))",
                                (user_id, session_id, bot_un_clean, bot_un_at)
                            ).fetchone()
                            
                            if not check_bot:
                                return # Бот не привязан к этому аккаунту! Игнорируем.
                # -----------------------------------------------------------------
                
                # --- МАГИЯ EXCEL ---
                parsed_text = await parse_excel_message(client, message, chat_id, user_id)
                if not parsed_text:
                    return

                with app.app_context():
                    db = get_db()
                    # 1. Обновляем текст в базе
                    db.execute("UPDATE messages SET text = ? WHERE telegram_message_id = ? AND chat_id = ? AND user_id = ?", 
                               (parsed_text, message.id, chat_id, user_id))
        
                    bindings = db.execute("""
                        SELECT pm.id, pm.line_index, p.synonyms 
                        FROM product_messages pm 
                        JOIN products p ON pm.product_id = p.id 
                        JOIN messages m ON pm.message_id = m.id
                        WHERE m.telegram_message_id = ? AND m.chat_id = ? AND m.user_id = ?
                    """, (message.id, chat_id, user_id)).fetchall()
        
                    lines = parsed_text.split('\n') if parsed_text else []

                    for b in bindings:
                        line_idx = b['line_index']
                        synonyms = [s.strip().lower() for s in b['synonyms'].split(',') if s.strip()]
                        
                        is_actual = 0
                        new_price = None
                        
                        # Если привязка идет ко всему сообщению (line_index = -1)
                        if line_idx == -1:
                            text_lower = parsed_text.lower()
                            # Проверяем, остался ли товар в этой строке
                            if any(syn in text_lower for syn in synonyms):
                                is_actual = 1
                                new_price = extract_price(edited_line)
                                if not new_price and line_idx + 1 < len(lines):
                                    new_price = extract_price(lines[line_idx+1])
                                if not new_price and line_idx + 2 < len(lines):
                                    new_price = extract_price(lines[line_idx+2])
                                
                        # Если привязка идет к конкретной строке
                        else:
                            if 0 <= line_idx < len(lines):
                                edited_line = lines[line_idx]
                                text_lower = edited_line.lower()
                                
                                # Проверяем, остался ли товар в этой строке
                                if any(syn in text_lower for syn in synonyms):
                                    is_actual = 1
                                    # ДОСТАЕМ НОВУЮ ЦЕНУ ИЗ ОТРЕДАКТИРОВАННОЙ СТРОКИ
                                    new_price = extract_price(edited_line)
                        
                        # Обновляем и статус актуальности, и НОВУЮ ЦЕНУ
                        db.execute("""
                            UPDATE product_messages 
                            SET is_actual = ?, extracted_price = ? 
                            WHERE id = ?
                        """, (is_actual, new_price, b['id']))
            
                    db.commit()
                    notify_clients()

            # --------------------------------------------------
            # --------------------------------------------------

            

            @client.on(events.NewMessage)
            async def handler(event):
                message = event.message
                
                if (not message.text and not message.document):
                    return

                chat = await event.get_chat()

                # --- ИГНОРИРУЕМ ИСХОДЯЩИЕ ТОЛЬКО ДЛЯ БОТОВ ---
                if getattr(message, 'out', False) and event.is_private and getattr(chat, 'bot', False):
                    return
                # ---------------------------------------------

                allowed_now = is_parsing_allowed(session_id)
                is_delayed = 0 if allowed_now else 1
                
                chat_id = chat.id

                # Проверка: отслеживаем ли мы этот чат?
                with app.app_context():
                    db = get_db()
                    cursor = db.execute(
                        "SELECT COUNT(*) FROM tracked_chats WHERE user_id = ? AND chat_id = ?",
                        (user_id, chat_id)
                    )
                    count = cursor.fetchone()[0]
                    
                    # СТРОГОЕ ПРАВИЛО: Если чат не добавлен в список - игнорируем сообщение!
                    if count == 0:
                        return

                # Определяем тип и название чата
                if event.is_private:
                    first = getattr(chat, 'first_name', '')
                    last = getattr(chat, 'last_name', '')
                    chat_title = f"{first} {last}".strip() or "Unknown"
                    msg_type = 'private'
                elif event.is_group:
                    chat_title = getattr(chat, 'title', 'Untitled Group')
                    msg_type = 'group'
                elif event.is_channel:
                    chat_title = getattr(chat, 'title', 'Untitled Channel')
                    msg_type = 'channel'
                else:
                    chat_title = "Unknown"
                    msg_type = 'unknown'

                sender = await event.get_sender()
                sender_name = f"{getattr(sender, 'first_name', '')} {getattr(sender, 'last_name', '')}".strip() or "Unknown"

                # --- ЖЕСТКАЯ ИЗОЛЯЦИЯ ПАРСИНГА БОТОВ (ДЛЯ НОВЫХ СООБЩЕНИЙ) ---
                if event.is_private and getattr(sender, 'bot', False):
                    bot_un = getattr(sender, 'username', '')
                    if not bot_un:
                        try:
                            entity = await client.get_entity(chat_id)
                            bot_un = getattr(entity, 'username', '')
                        except: pass
                        
                    bot_un_clean = bot_un.lower().replace('@', '') if bot_un else ''
                    bot_un_at = f"@{bot_un_clean}" if bot_un_clean else ''
                    
                    with app.app_context():
                        local_db = get_db()
                        check_bot = local_db.execute(
                            "SELECT id FROM interaction_bots WHERE user_id = ? AND userbot_id = ? AND (LOWER(bot_username) IN (?, ?))",
                            (user_id, session_id, bot_un_clean, bot_un_at)
                        ).fetchone()
                        
                        if not check_bot:
                            return # Бот не привязан к этому аккаунту! Игнорируем.
                # -------------------------------------------------------------

                parsed_text = await parse_excel_message(client, message, chat_id, user_id)
                
                # Если после парсинга текста нет (например, пустой Excel или не настроен), прерываем
                if not parsed_text:
                    return

                with app.app_context():
                    db = get_db()
                    try:
                        # 1. Записываем новое сообщение
              
                        msg_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        cursor = db.execute(
                            'INSERT INTO messages (user_id, telegram_message_id, type, text, date, chat_id, chat_title, sender_name, is_delayed) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)',
                            (user_id, message.id, msg_type, parsed_text, msg_date, chat_id, chat_title, sender_name, is_delayed)
                        )
                        new_msg_id = cursor.lastrowid
                        db.commit()

                        # --- ЧИСТИМ ДУБЛИ СРАЗУ, ДАЖЕ НОЧЬЮ ---
                        

                        # ТАМОЖНЯ: Если время нерабочее - выходим (не обновляем цены)
                        if not allowed_now:
                            return
                      
                        # === НОВЫЙ БЛОК: Умная очистка старых сообщений и миграция привязок ===
                        
                        
                        # А. Ищем привязки товаров к старым сообщениям из этого же чата
                        old_bindings = db.execute("""
                            SELECT pm.id, pm.line_index, p.synonyms
                            FROM product_messages pm
                            JOIN products p ON pm.product_id = p.id
                            JOIN messages m ON pm.message_id = m.id
                            WHERE m.chat_id = ? AND m.user_id = ? AND m.id != ?
                        """, (chat_id, user_id, new_msg_id)).fetchall()

                        if old_bindings and parsed_text:
                            lines = parsed_text.split('\n')
                            text_lower = parsed_text.lower()
                            
                            for b in old_bindings:
                                synonyms = [s.strip().lower() for s in b['synonyms'].split(',') if s.strip()]
                                line_idx = b['line_index']
                                match_found = False
                                new_price = None
                                
                                if line_idx == -1:
                                    if any(syn in text_lower for syn in synonyms):
                                        match_found = True
                                        new_price = extract_price(parsed_text)
                                else:
                                    for i, line in enumerate(lines):
                                        if any(syn in line.lower() for syn in synonyms):
                                            match_found = True
                                            new_price = extract_price(line)
                                            if not new_price and i + 1 < len(lines):
                                                new_price = extract_price(lines[i+1])
                                            if not new_price and i + 2 < len(lines):
                                                new_price = extract_price(lines[i+2])
                                            line_idx = i
                                            break
                                
                                # Если товар все еще есть в новом прайсе — переносим привязку
                                if match_found:
                                    db.execute("""
                                        UPDATE product_messages 
                                        SET message_id = ?, line_index = ?, extracted_price = ?, is_actual = 1 
                                        WHERE id = ?
                                    """, (new_msg_id, line_idx, new_price, b['id']))
                        
                        # Б. Безопасное удаление старых сообщений (разгружаем базу)
                        # Оставляем зазор в 5 минут, чтобы не сломать прайсы, которые бот отправляет несколькими сообщениями подряд
                        # Зазор 0 минут: как только пришел новый прайс, старый удаляется МГНОВЕННО
                        cutoff_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        db.execute("""
                            DELETE FROM messages 
                            WHERE chat_id = ? 
                              AND user_id = ? 
                              AND id != ? 
                              AND date < ?
                              AND id NOT IN (SELECT message_id FROM product_messages)
                        """, (chat_id, user_id, new_msg_id, cutoff_time))
                        
                        db.commit()
                        # =======================================================================

                        notify_clients()
                        
                        if chat_title:
                            db.execute(
                                "UPDATE tracked_chats SET chat_title = ? WHERE user_id = ? AND chat_id = ? AND chat_title IS NULL",
                                (chat_title, user_id, chat_id)
                            )
                            db.commit()
                            notify_clients()
                    except sqlite3.IntegrityError:
                        pass

            async def main():
                await client.connect()
            
                # --- ОБРАБОТКА СЛЕТЕВШЕЙ АВТОРИЗАЦИИ ---
                if not await client.is_user_authorized():
                    logging.error(f"Сессия {session_id}: Клиент не авторизован!")
                    try:
                        await client.disconnect()
                    except Exception:
                        pass
                
                    with app.app_context():
                        db = get_db()
                        # Устанавливаем статус unauthorized
                        db.execute("UPDATE user_telegram_sessions SET status = 'unauthorized' WHERE id = ?", (session_id,))
                        db.commit()
                        if 'notify_clients' in globals(): notify_clients()
                    return
                # ---------------------------------------

                me = await client.get_me()
            
                # --- Получаем @username или Имя юзербота ---
                account_name = getattr(me, 'username', None)
                if account_name:
                    account_name = f"@{account_name}"
                else:
                    account_name = getattr(me, 'first_name', f"ID: {me.id}")
                # --------------------------------------------------
            
                logging.info(f"Сессия {session_id}: слушатель запущен для {me.first_name}")
            
                with app.app_context():
                    db = get_db()
                    try:
                        db.execute(
                            "UPDATE user_telegram_sessions SET status = 'active', session_file = ?, account_name = ? WHERE id = ?",
                            (session_file, account_name, session_id)
                        )
                    except sqlite3.OperationalError:
                        db.execute(
                            "UPDATE user_telegram_sessions SET status = 'active', session_file = ? WHERE id = ?",
                            (session_file, session_id)
                        )
                    db.commit()
                    if 'notify_clients' in globals(): notify_clients()
                
                # Блокирующий вызов: клиент слушает события до отключения
                # Блокирующий вызов: клиент слушает события до отключения
                try:
                    await client.run_until_disconnected()
                except AuthKeyUnregisteredError:
                    logging.error(f"Сессия {session_id}: Ключ авторизации отозван сервером Telegram (слетела сессия).")
                    with app.app_context():
                        db = get_db()
                        db.execute("UPDATE user_telegram_sessions SET status = 'unauthorized' WHERE id = ?", (session_id,))
                        db.commit()
                        if 'notify_clients' in globals(): notify_clients()


            # --- ЗАПУСК И БЕЗОПАСНАЯ ОЧИСТКА В ЦИКЛЕ ---
            try:
                loop.run_until_complete(main())
            except Exception as e:
                logging.exception(f"Сессия {session_id}: ошибка в слушателе - {e}")
            finally:
                # 1. Корректное отключение от серверов Telegram
                try:
                    if client and client.is_connected():
                        loop.run_until_complete(client.disconnect())
                except Exception:
                    pass
            
                # 2. Очищаем глобальные словари ВНУТРИ родного потока.
                # Это полностью исключает панику сборщика мусора (GC) в Flask!
                user_clients.pop(session_id, None)
                user_loops.pop(session_id, None)
                background_tasks.pop(session_id, None)

                # 3. Отменяем оставшиеся фоновые задачи Telethon и закрываем цикл
                try:
                    tasks = asyncio.all_tasks(loop)
                    for t in tasks:
                        t.cancel()
                
                    if tasks:
                        # Даем задачам шанс чисто закрыться (игнорируя ошибки отмены)
                        loop.run_until_complete(asyncio.gather(*tasks, return_exceptions=True))
                
                    loop.close()
                except Exception:
                    pass

                logging.info(f"Сессия {session_id}: слушатель полностью остановлен")

    # --- СОЗДАНИЕ И ЗАПУСК ПОТОКА ---
    # Даем потоку имя для удобства дебага
    thread = threading.Thread(target=listener, daemon=True, name=f"UserbotThread-{session_id}")
    thread.start()
    background_tasks[session_id] = thread
    logging.info(f"Поток слушателя для сессии {session_id} создан")

# ---------- Маршруты Flask ----------
@app.route('/')
@login_required
def index():
    return render_template_string(TEMPLATE)

@app.route('/api/profile/logout_others', methods=['POST'])
@login_required
def logout_others():
    user_id = session['user_id']
    new_token = secrets.token_hex(16)
    
    db = get_db()
    db.execute("UPDATE users SET session_token = ? WHERE id = ?", (new_token, user_id))
    db.commit()
    
    session['session_token'] = new_token
    
    # --- НОВАЯ СТРОКА: Рассылаем сигнал всем подключенным клиентам ---
    socketio.emit('force_logout')
    
    return jsonify({'success': True})

@app.route('/login', methods=['GET', 'POST'])
def login_page():
    if request.method == 'POST':
        login = request.form.get('login')
        password = request.form.get('password')
        pwd_hash = hashlib.sha256(password.encode()).hexdigest()
        
        db = get_db()
        user = db.execute("SELECT * FROM users WHERE login = ? AND password_hash = ?", (login, pwd_hash)).fetchone()
        
        if user:
            # --- ГЕНЕРАЦИЯ И СОХРАНЕНИЕ ТОКЕНА ---
            token = user['session_token'] if 'session_token' in user.keys() else None
            if not token:
                token = secrets.token_hex(16)
                try:
                    db.execute("UPDATE users SET session_token = ? WHERE id = ?", (token, user['id']))
                    db.commit()
                except sqlite3.OperationalError:
                    pass # На случай, если колонка еще не создалась
            
            session['user_id'] = user['id']
            session['role'] = user['role']
            session['login'] = user['login']
            session['session_token'] = token # Обязательно записываем в браузер
            # -------------------------------------
            
            return redirect(url_for('index'))
        return render_template_string(LOGIN_TEMPLATE, error="Неверный логин или пароль")
    return render_template_string(LOGIN_TEMPLATE)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login_page'))

@app.route('/api/publish_tree_data', methods=['GET'])
@login_required
def get_publish_tree_data():
    db = get_db()
    user_id = session.get('user_id')
    
    folders = db.execute("SELECT id, name, parent_id FROM folders WHERE user_id=?", (user_id,)).fetchall()
    products = db.execute("SELECT id, name, folder_id FROM products WHERE user_id=?", (user_id,)).fetchall()
    
    # Получаем поставщиков, у которых есть актуальные цены на товары
    # ВАЖНО: Я использую m.chat_id и m.chat_title. Если в вашей базе таблица messages 
    # использует sender_id или channel_name, просто замените эти слова в запросе ниже!
    suppliers = db.execute("""
        SELECT pm.product_id, m.chat_id as supplier_id, m.chat_title as supplier_name, pm.extracted_price
        FROM product_messages pm
        JOIN messages m ON pm.message_id = m.id
        WHERE pm.status = 'confirmed' AND pm.is_actual = 1 AND m.user_id = ?
        GROUP BY pm.product_id, m.chat_id
    """, (user_id,)).fetchall()
    
    return jsonify({
        'folders': [dict(f) for f in folders],
        'products': [dict(p) for p in products],
        'suppliers': [dict(s) for s in suppliers]
    })

@app.route('/api/tracked_chats/<int:chat_id>', methods=['PUT'])
@login_required
def update_tracked_chat(chat_id):
    data = request.get_json()
    custom_name = data.get('custom_name')
    user_id = session['user_id']
    db = get_db()
    # Проверяем, принадлежит ли чат пользователю
    cur = db.execute("SELECT id FROM tracked_chats WHERE id = ? AND user_id = ?", (chat_id, user_id))
    if not cur.fetchone():
        return jsonify({'error': 'Not found'}), 404
    # Разрешаем пустую строку, сохраняем как NULL
    if custom_name == '':
        custom_name = None
    db.execute("UPDATE tracked_chats SET custom_name = ? WHERE id = ?", (custom_name, chat_id))
    db.commit()
    notify_clients()
    return jsonify({'success': True})


@app.route('/api/get_qr', methods=['POST'])
@login_required
def get_qr():
    ensure_event_loop()
    data = request.get_json()
    api_id = data.get('api_id')
    api_hash = data.get('api_hash')
    session_id = data.get('session_id') # <--- ВОТ ЗДЕСЬ ПРИНИМАЕМ ID СТАРОЙ СЕССИИ
    user_id = session['user_id']

    if not api_id or not api_hash:
        return jsonify({'success': False, 'error': 'Missing fields'}), 400

    db = get_db()
    
    # --- ИСПРАВЛЕННАЯ ЛОГИКА ---
    if session_id:
        # Если передали ID старой сессии, просто обновляем её статус
        db.execute("UPDATE user_telegram_sessions SET status = 'pending' WHERE id = ? AND user_id = ?", (session_id, user_id))
    else:
        # Иначе создаем новую
        cursor = db.execute(
            "INSERT INTO user_telegram_sessions (user_id, api_id, api_hash, status) VALUES (?, ?, ?, 'pending')",
            (user_id, api_id, api_hash)
        )
        session_id = cursor.lastrowid
    # ---------------------------
        
    db.commit()
    if 'notify_clients' in globals(): notify_clients()
    
    session_file = f'sessions/user_{user_id}_{session_id}'
    q = queue.Queue()

    def qr_worker():
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        client = TelegramClient(session_file, int(api_id), api_hash)

        async def _run():
            await client.connect()
            if await client.is_user_authorized():
                await client.disconnect()
                q.put({'already_authorized': True})
                return

            try:
                qr_login = await client.qr_login()
                q.put({'qr_url': qr_login.url})
                
                await qr_login.wait(timeout=300)
                await client.disconnect()
                
                with app.app_context():
                    local_db = get_db()
                    local_db.execute("UPDATE user_telegram_sessions SET status = 'active' WHERE id = ?", (session_id,))
                    local_db.commit()
                    if 'notify_clients' in globals(): notify_clients()
                logging.info(f"Сессия {session_id}: статус обновлён на active")
                start_message_listener(session_id, user_id, api_id, api_hash)

            except asyncio.TimeoutError:
                logging.error(f"Сессия {session_id}: таймаут ожидания QR")
                await client.disconnect()
                with app.app_context():
                    local_db = get_db()
                    # Если ошибка, возвращаем статус unauthorized вместо удаления
                    local_db.execute("UPDATE user_telegram_sessions SET status = 'unauthorized' WHERE id = ?", (session_id,))
                    local_db.commit()
                    if 'notify_clients' in globals(): notify_clients()
            except Exception as e:
                logging.exception(f"Сессия {session_id}: ошибка при ожидании QR")
                await client.disconnect()
                with app.app_context():
                    local_db = get_db()
                    local_db.execute("UPDATE user_telegram_sessions SET status = 'unauthorized' WHERE id = ?", (session_id,))
                    local_db.commit()
                    if 'notify_clients' in globals(): notify_clients()
            finally:
                qr_sessions.pop(session_id, None)

        loop.run_until_complete(_run())
        loop.close()

    thread = threading.Thread(target=qr_worker, daemon=True)
    thread.start()

    try:
        res = q.get(timeout=15)
    except queue.Empty:
        db.execute("UPDATE user_telegram_sessions SET status = 'unauthorized' WHERE id = ?", (session_id,))
        db.commit()
        if 'notify_clients' in globals(): notify_clients()
        return jsonify({'success': False, 'error': 'Таймаут генерации QR'}), 500

    if res.get('already_authorized'):
        start_message_listener(session_id, user_id, api_id, api_hash)
        db.execute("UPDATE user_telegram_sessions SET status = 'active' WHERE id = ?", (session_id,))
        db.commit()
        if 'notify_clients' in globals(): notify_clients()
        return jsonify({'success': True, 'already_authorized': True})

    qr_url = res.get('qr_url')
    if not qr_url:
        db.execute("UPDATE user_telegram_sessions SET status = 'unauthorized' WHERE id = ?", (session_id,))
        db.commit()
        if 'notify_clients' in globals(): notify_clients()
        return jsonify({'success': False, 'error': 'Не удалось сгенерировать QR'}), 500

    qr_sessions[session_id] = True 

    img = qrcode.make(qr_url)
    buffer = BytesIO()
    img.save(buffer, format='PNG')
    buffer.seek(0)
    img_base64 = base64.b64encode(buffer.getvalue()).decode()

    return jsonify({'success': True, 'qr': img_base64, 'session_id': session_id})

@app.route('/api/check_login')
@login_required
def check_login():
    session_id = request.args.get('session_id')
    if not session_id:
        return jsonify({'success': False, 'error': 'Missing session_id'}), 400
    session_id = int(session_id)

    db = get_db()
    row = db.execute("SELECT status FROM user_telegram_sessions WHERE id = ?", (session_id,)).fetchone()
    if row and row['status'] == 'active':
        return jsonify({'success': True})
    return jsonify({'success': False})

@app.route('/api/check_session')
@login_required
def check_session():
    user_id = session['user_id']
    active = any(uid == user_id for _, _, uid in user_clients.values())
    return jsonify({'authorized': active})

import requests

# --- РОУТЫ ДЛЯ НАСТРОЙКИ API-ПАРСЕРА ---
@app.route('/api/api_sources', methods=['GET', 'POST'])
@login_required
def handle_api_sources():
    db = get_db()
    user_id = session.get('user_id')
    
    if request.method == 'GET':
        sources = db.execute("SELECT * FROM api_sources WHERE user_id=?", (user_id,)).fetchall()
        return jsonify([dict(s) for s in sources])
    else:
        data = request.json
        db.execute("""
            INSERT INTO api_sources (user_id, name, url, token, interval_min)
            VALUES (?, ?, ?, ?, ?)
        """, (user_id, data.get('name'), data.get('url'), data.get('token'), data.get('interval_min', 60)))
        db.commit()
        return jsonify({'success': True})

@app.route('/api/api_sources/<int:source_id>', methods=['DELETE'])
def delete_api_source(source_id):
    try:
        db = get_db()
        user_id = session.get('user_id')
        dummy_chat_id = f"api_src_{source_id}"
        
        # 1. Удаляем привязки цен к товарам
        db.execute("DELETE FROM product_messages WHERE message_id IN (SELECT id FROM messages WHERE chat_id=? AND user_id=?)", (dummy_chat_id, user_id))
        
        # 2. Удаляем сами сохраненные прайсы/сообщения
        db.execute("DELETE FROM messages WHERE chat_id=? AND user_id=?", (dummy_chat_id, user_id))
        
        # 3. Удаляем само подключение API
        db.execute("DELETE FROM api_sources WHERE id=? AND user_id=?", (source_id, user_id))
        
        db.commit()
        return jsonify({'success': True})
    except Exception as e:
        print(f"Ошибка при удалении API: {e}")
        return jsonify({'error': str(e)}), 500

# --- ФОНОВЫЙ ПАРСЕР ЧУЖИХ API ---
def api_parser_scheduler():
    """Раз в минуту проверяет чужие API и заносит товары в базу как поставщик"""
    time.sleep(5)
    last_sync_times = {}
    
    with app.app_context():
        while True:
            try:
                db = get_db()
                sources = db.execute("SELECT * FROM api_sources WHERE is_active=1").fetchall()
                now = datetime.now()
                
                for src in sources:
                    src_id = src['id']
                    user_id = src['user_id']
                    interval = src['interval_min']
                    
                    # Проверяем, пришло ли время опрашивать этот API
                    last_run = last_sync_times.get(src_id)
                    if last_run and (now - last_run).total_seconds() / 60 < interval:
                        continue
                        
                    # Формируем URL с токеном (как в вашей системе)
                    url = f"{src['url']}?token={src['token']}"
                    
                    try:
                        resp = requests.get(url, timeout=15)
                        if resp.status_code == 200:
                            data = resp.json()
                            products = data.get('products', [])
                            
                            # Отбираем только валидные товары
                            valid_products = [p for p in products if p.get('name') and p.get('price') is not None]
                            
                            # Создаем "фейковый чат", чтобы API выглядел как поставщик
                            dummy_chat_id = f"api_src_{src_id}"
                            dummy_chat_title = f"🌐 API: {src['name']}"
                            
                            # 1. Собираем все товары в единый текст (построчно)
                            lines = [f"Авто-обновление из API ({len(valid_products)} товаров)"]
                            for p in valid_products:
                                lines.append(f"{p['name']} — {p['price']}")
                                
                            msg_text = "\n".join(lines)
                            sender_name = f"🌐 API: {src['name']}"
                            
                            # Вставляем "сообщение" от этого поставщика с полным списком
                            cursor = db.execute(
                                "INSERT INTO messages (user_id, chat_id, chat_title, text, date, sender_name) VALUES (?, ?, ?, ?, ?, ?) RETURNING id",
                                (user_id, dummy_chat_id, dummy_chat_title, msg_text, now, sender_name)
                            )
                            dummy_msg_id = cursor.fetchone()['id']
                            
                            # 2. Привязываем товары с указанием конкретной строки (line_index)
                            for i, p in enumerate(valid_products):
                                name = p.get('name')
                                price = p.get('price')
                                
                                # Ищем товар в базе или создаем новый (без папки)
                                prod_row = db.execute("SELECT id FROM products WHERE name=? AND user_id=?", (name, user_id)).fetchone()
                                if prod_row:
                                    prod_id = prod_row['id']
                                else:
                                    cur = db.execute("INSERT INTO products (user_id, name) VALUES (?, ?) RETURNING id", (user_id, name))
                                    prod_id = cur.fetchone()['id']
                                    
                                # line_index = i + 1, так как 0-я строка - это заголовок "Авто-обновление..."
                                db.execute("""
                                    INSERT INTO product_messages (product_id, message_id, extracted_price, status, is_actual, line_index)
                                    VALUES (?, ?, ?, 'pending', 1, ?)
                                """, (prod_id, dummy_msg_id, price, i + 1))
                                
                            db.commit()
                            notify_clients()
                            print(f"✅ Спарсили {len(valid_products)} товаров из {src['name']}")
                    except Exception as req_err:
                        print(f"❌ Ошибка запроса к API {src['name']}: {req_err}")
                        
                    last_sync_times[src_id] = now
            except Exception as e:
                print(f"Ошибка в планировщике парсинга API: {e}")
                
            time.sleep(60) # Проверяем таймеры каждую минуту



import math
from flask import request, jsonify, session

import math
from flask import request, jsonify, session

@app.route('/api/messages')
def get_messages():
    # Проверка авторизации
    if 'user_id' not in session:
        return jsonify({'messages': [], 'error': 'Not authorized'}), 401
        
    user_id = session['user_id']
    msg_type = request.args.get('type', 'all')
    search = request.args.get('search', '').lower()
    exclude = request.args.get('exclude', '').lower()
    
    # НОВОЕ: Получаем данные из выпадающих списков
    sender = request.args.get('sender', '')
    msg_source_type = request.args.get('msg_type', '')
    
    # Обработка пагинации
    try:
        page = int(request.args.get('page', 1))
        if page < 1: page = 1
    except ValueError:
        page = 1
        
    try:
        limit = int(request.args.get('limit', 20))
        if limit < 20: limit = 20
        elif limit > 100: limit = 100
    except ValueError:
        limit = 20
        
    offset = (page - 1) * limit

    db = get_db()
    
    # ДОБАВЛЕНО: Скрываем сообщения, которые пришли вне тайминга (is_delayed = 1)
    where_clauses = ["m.user_id = ?", "(m.is_blocked IS NULL OR m.is_blocked = 0)", "(m.is_delayed IS NULL OR m.is_delayed = 0)"]
    params = [user_id]

    if msg_type == 'favorites':
        where_clauses.append("m.is_favorite = 1")
    elif msg_type.startswith('folder_'):
        folder_id = msg_type.replace('folder_', '')
        # Связываем сообщения с папкой через таблицу отслеживаемых чатов (tc)
        where_clauses.append("tc.folder_id = ?")
        params.append(folder_id)
    elif msg_type.startswith('chat_'):
        chat_id = msg_type.replace('chat_', '')
        where_clauses.append("m.chat_id = ?")
        params.append(chat_id)

    if search:
        where_clauses.append("LOWER(m.text) LIKE ?")
        params.append(f"%{search}%")
        
    if exclude:
        where_clauses.append("LOWER(m.text) NOT LIKE ?")
        params.append(f"%{exclude}%")

    # НОВОЕ: Фильтрация по "От кого"
    # НОВОЕ: Фильтрация по "От кого"
    if sender and sender != 'all':
        # Используем REPLACE, чтобы база данных игнорировала баг со словом ' None' при поиске
        where_clauses.append("(tc.custom_name = ? OR m.chat_title = ? OR REPLACE(m.sender_name, ' None', '') = ?)")
        params.extend([sender, sender, sender])

    # НОВОЕ: Фильтрация по "Тип сообщений"
    if msg_source_type and msg_source_type != 'all':
        where_clauses.append("m.type = ?")
        params.append(msg_source_type)

    where_str = " AND ".join(where_clauses)
    
    try:
        # --- ШАГ 1: Считаем уникальные связки (Чат + Заголовок прайса) ---
        count_query = f"""
            SELECT COUNT(*) as total FROM (
                SELECT 1
                FROM messages m
                LEFT JOIN tracked_chats tc ON CAST(m.chat_id AS TEXT) = CAST(tc.chat_id AS TEXT) AND m.user_id = tc.user_id
                WHERE {where_str}
                GROUP BY m.chat_id, SUBSTR(m.text, 1, 50)
            )
        """
        total_count_row = db.execute(count_query, params).fetchone()
        total_count = total_count_row['total'] if total_count_row else 0
        
        # Вычисляем общее количество страниц
        total_pages = math.ceil(total_count / limit) if total_count > 0 else 1

        # --- ШАГ 2: Запрашиваем последние сообщения для каждого типа прайса ---
        data_query = f"""
            SELECT m.*, tc.custom_name, MAX(m.date) as latest_date
            FROM messages m
            LEFT JOIN tracked_chats tc 
                   ON CAST(m.chat_id AS TEXT) = CAST(tc.chat_id AS TEXT) 
                  AND m.user_id = tc.user_id
            WHERE {where_str} 
            GROUP BY m.chat_id, SUBSTR(m.text, 1, 30)  -- Группируем по чату и первым 50 символам
            ORDER BY latest_date DESC 
            LIMIT ? OFFSET ?
        """
        
        # Добавляем limit и offset в конец списка параметров
        data_params = params.copy()
        data_params.extend([limit, offset])
        
        cursor = db.execute(data_query, data_params)
        messages = [dict(row) for row in cursor.fetchall()]

        # Возвращаем полный объект на фронтенд
        return jsonify({
            'messages': messages,
            'total': total_count,
            'page': page,
            'limit': limit,
            'total_pages': total_pages
        })

    except Exception as e:
        print(f"Ошибка в API сообщений: {e}")
        return jsonify({'messages': [], 'total': 0, 'page': page, 'limit': limit, 'total_pages': 1, 'error': str(e)})


# --- МАРШРУТЫ ДЛЯ АВТОПУБЛИКАЦИЙ ---

@app.route('/api/publications', methods=['GET'])
@login_required
def get_publications():
    db = get_db()
    user_id = session['user_id'] # Получаем ID текущего пользователя
    pubs = db.execute("SELECT * FROM publications WHERE user_id = ?", (user_id,)).fetchall() # Фильтруем
    return jsonify([dict(p) for p in pubs])

@app.route('/api/publications', methods=['POST'])
@login_required
def save_publication():
    data = request.json
    db = get_db()
    pub_id = data.get('id')
    
    allowed_items_json = json.dumps(data.get('allowed_items', {}))
    markup_type = data.get('markup_type', 'percent')
    markup_value = float(data.get('markup_value', 0))
    rounding = int(data.get('rounding', 100))
    
    if pub_id:
        db.execute("""
            UPDATE publications 
            SET name=?, is_active=?, interval_min=?, chat_id=?, message_id=?, userbot_id=?, template=?, allowed_items=?, markup_type=?, markup_value=?, rounding=?
            WHERE id=?
        """, (data.get('name'), data.get('is_active', 0), data.get('interval_min', 60), 
              data.get('chat_id'), data.get('message_id'), data.get('userbot_id'), 
              data.get('template'), allowed_items_json, markup_type, markup_value, rounding, pub_id))
    else:
        db.execute("""
    INSERT INTO publications (name, is_active, interval_min, chat_id, message_id, userbot_id, template, allowed_items, markup_type, markup_value, rounding, user_id)
    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (data.get('name'), data.get('is_active', 0), data.get('interval_min', 60), 
              data.get('chat_id'), data.get('message_id'), data.get('userbot_id'), 
              data.get('template'), allowed_items_json, markup_type, markup_value, rounding,session['user_id']))
        # Получаем ID только что созданной публикации
        pub_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
        
    # --- СОХРАНЕНИЕ НАЦЕНОК, КОТОРЫЕ БЫЛИ ДОБАВЛЕНЫ ДО СОЗДАНИЯ ПУБЛИКАЦИИ ---
    temp_markups = data.get('temp_markups', [])
    for m in temp_markups:
        db.execute("""
            INSERT INTO pub_markups (pub_id, folder_id, markup_type, markup_value, rounding) 
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(pub_id, folder_id) DO UPDATE SET 
            markup_type=excluded.markup_type, markup_value=excluded.markup_value, rounding=excluded.rounding
        """, (pub_id, m['folder_id'], m['markup_type'], m['markup_value'], m['rounding']))
        
    db.commit()
    notify_clients()
    return jsonify({'success': True})

@app.route('/api/publications/<int:pub_id>', methods=['DELETE'])
@login_required
def delete_publication(pub_id):
    db = get_db()
    db.execute("DELETE FROM publications WHERE id=?", (pub_id,))
    db.commit()
    notify_clients()
    return jsonify({'success': True})
    
@app.route('/api/publications/<int:pub_id>/toggle', methods=['POST'])
@login_required
def toggle_publication(pub_id):
    data = request.json
    db = get_db()
    db.execute("UPDATE publications SET is_active=? WHERE id=?", (data.get('is_active', 0), pub_id))
    db.commit()
    notify_clients()
    return jsonify({'success': True})

@app.route('/api/folders/tree')
@login_required
def get_folder_tree():
    user_id = session['user_id']
    db = get_db()
    cursor = db.execute("SELECT id, name, parent_id FROM folders WHERE user_id = ? ORDER BY parent_id, name", (user_id,))
    rows = cursor.fetchall()
    folders = {}
    roots = []
    for row in rows:
        folders[row['id']] = {
            'id': row['id'],
            'name': row['name'],
            'parent_id': row['parent_id'],
            'children': []
        }
    for fid, f in folders.items():
        if f['parent_id'] is None:
            roots.append(f)
        else:
            if f['parent_id'] in folders:
                folders[f['parent_id']]['children'].append(f)
    return jsonify(roots)

@app.route('/api/folders', methods=['POST'])
@login_required
def create_folder():
    user_id = session['user_id']
    data = request.get_json()
    name = data.get('name')
    parent_id = data.get('parent_id')
    if not name:
        return jsonify({'error': 'Name required'}), 400
    db = get_db()
    try:
        db.execute("INSERT INTO folders (user_id, name, parent_id) VALUES (?, ?, ?)",
                   (user_id, name, parent_id))
        db.commit()
        notify_clients()
        return jsonify({'success': True})
    except sqlite3.IntegrityError:
        return jsonify({'error': 'Folder exists'}), 400

@app.route('/api/folders/<int:folder_id>/rename', methods=['POST'])
@login_required
def rename_folder_route(folder_id):
    user_id = session['user_id']
    data = request.get_json()
    new_name = data.get('name')
    
    if not new_name:
        return jsonify({'error': 'Имя не может быть пустым'}), 400
        
    db = get_db()
    
    row = db.execute("SELECT name FROM folders WHERE id = ? AND user_id = ?", (folder_id, user_id)).fetchone()
    if not row:
        return jsonify({'error': 'Папка не найдена'}), 404
    if row['name'] == 'По умолчанию':
        return jsonify({'error': 'Нельзя переименовать папку по умолчанию'}), 400
        
    try:
        db.execute("UPDATE folders SET name = ? WHERE id = ? AND user_id = ?", (new_name, folder_id, user_id))
        db.commit()
        notify_clients() # Уведомляем клиентов для автообновления
        return jsonify({'success': True})
    except Exception as e:
        logger.error(f"Ошибка переименования: {e}")
        return jsonify({'error': 'Системная ошибка при переименовании'}), 500

@app.route('/api/folders/<int:folder_id>', methods=['DELETE'])
@login_required
def delete_folder(folder_id):
    user_id = session['user_id']
    db = get_db()
    row = db.execute("SELECT name FROM folders WHERE id = ? AND user_id = ?", (folder_id, user_id)).fetchone()
    if not row:
        return jsonify({'error': 'Folder not found'}), 404
    if row['name'] == 'По умолчанию':
        return jsonify({'error': 'Cannot delete default folder'}), 400
        
    db.execute("DELETE FROM folders WHERE id = ?", (folder_id,))
    db.commit()
    notify_clients()
    return jsonify({'success': True})

@app.route('/api/folder/<int:folder_id>/messages')
@login_required
def get_folder_messages(folder_id):
    user_id = session['user_id']
    db = get_db()
    cursor = db.execute("""
        SELECT m.*, tc.custom_name, tc.chat_title as tc_chat_title
        FROM messages m
        JOIN saved_messages sm ON m.id = sm.message_id
        LEFT JOIN tracked_chats tc ON m.chat_id = tc.chat_id AND m.user_id = tc.user_id
        WHERE sm.folder_id = ? AND m.user_id = ?
        AND (m.is_blocked IS NULL OR m.is_blocked = 0)
        ORDER BY sm.saved_date DESC
    """, (folder_id, user_id))
    # ...
    rows = cursor.fetchall()
    messages = [{
        'id': row['id'],
        'telegram_message_id': row['telegram_message_id'], 
        'chat_id': row['chat_id'],                         
        'type': row['type'],
        'text': row['text'],
        'date': row['date'],
        'chat_title': row['chat_title'],
        'sender_name': row['sender_name'] if 'sender_name' in row.keys() else 'Unknown',
        'custom_name': row['custom_name'] if 'custom_name' in row.keys() else None,
        'tc_chat_title': row['tc_chat_title'] if 'tc_chat_title' in row.keys() else None # <--- ДОБАВИТЬ ЭТУ СТРОКУ
    } for row in rows]
    return jsonify(messages)


@app.route('/api/save_messages', methods=['POST'])
@login_required
def save_messages():
    user_id = session['user_id']
    data = request.get_json()
    message_ids = data.get('message_ids', [])
    folder_id = data.get('folder_id')
    if not message_ids or not folder_id:
        return jsonify({'error': 'Missing data'}), 400
    db = get_db()
    cursor = db.execute("SELECT id FROM folders WHERE id = ? AND user_id = ?", (folder_id, user_id))
    if not cursor.fetchone():
        return jsonify({'error': 'Folder not found'}), 404
    for mid in message_ids:
        try:
            db.execute("INSERT INTO saved_messages (message_id, folder_id) VALUES (?, ?)", (mid, folder_id))
        except sqlite3.IntegrityError:
            pass
    db.commit()
    notify_clients()
    return jsonify({'success': True})

@app.route('/api/userbots', methods=['GET'])
@login_required
def get_userbots():
    user_id = session['user_id']
    db = get_db()
    # ДОБАВИЛИ time_start, time_end, schedule_enabled в SQL-запрос
    try:
        cursor = db.execute("SELECT id, api_id, api_hash, status, account_name, time_start, time_end, schedule_enabled FROM user_telegram_sessions WHERE user_id=?", (user_id,))
    except sqlite3.OperationalError:
        cursor = db.execute("SELECT id, api_id, api_hash, status, time_start, time_end, schedule_enabled FROM user_telegram_sessions WHERE user_id=?", (user_id,))
        
    bots = [dict(row) for row in cursor.fetchall()]
    return jsonify(bots)

# --- НОВЫЙ МАРШРУТ ДЛЯ СОХРАНЕНИЯ РАСПИСАНИЯ ---
@app.route('/api/userbots/<int:bot_id>/schedule', methods=['POST'])
@login_required
def update_userbot_schedule(bot_id):
    user_id = session['user_id']
    data = request.get_json()
    db = get_db()
    db.execute("""
        UPDATE user_telegram_sessions 
        SET time_start = ?, time_end = ?, schedule_enabled = ? 
        WHERE id = ? AND user_id = ?
    """, (data['time_start'], data['time_end'], int(data['schedule_enabled']), bot_id, user_id))
    db.commit()
    notify_clients()
    return jsonify({'success': True})
# -----------------------------------------------

@app.route('/api/userbots/<int:bot_id>/toggle', methods=['POST'])
@login_required
def toggle_userbot(bot_id):
    user_id = session['user_id']
    db = get_db()
    cursor = db.execute("SELECT user_id, api_id, api_hash, status FROM user_telegram_sessions WHERE id=?", (bot_id,))
    row = cursor.fetchone()
    if not row or row['user_id'] != user_id:
        return jsonify({'error': 'Not found'}), 404

    if row['status'] == 'active':
        stop_message_listener(bot_id)
        new_status = 'inactive'
    else:
        start_message_listener(bot_id, user_id, row['api_id'], row['api_hash'])
        new_status = 'active'

    db.execute("UPDATE user_telegram_sessions SET status = ? WHERE id = ?", (new_status, bot_id))
    db.commit()
    notify_clients()
    return jsonify({'success': True, 'status': new_status})

@app.route('/api/userbots/<int:bot_id>', methods=['DELETE'])
@login_required
def delete_userbot(bot_id):
    user_id = session['user_id']
    db = get_db()
    cursor = db.execute("SELECT user_id, session_file FROM user_telegram_sessions WHERE id=?", (bot_id,))
    row = cursor.fetchone()
    if not row or row['user_id'] != user_id:
        return jsonify({'error': 'Not found'}), 404

    if bot_id in user_clients:
        stop_message_listener(bot_id)

    session_file = row['session_file']
    if session_file and os.path.exists(session_file):
        os.remove(session_file)

    db.execute("DELETE FROM user_telegram_sessions WHERE id = ?", (bot_id,))
    db.commit()
    notify_clients()
    return jsonify({'success': True})



@app.route('/api/excel/missing_sheets', methods=['GET'])
@login_required
def get_missing_sheets():
    db = get_db()
    missing = db.execute("SELECT * FROM excel_missing_sheets WHERE user_id = ?", (session['user_id'],)).fetchall()
    return jsonify([dict(m) for m in missing])

@app.route('/api/excel/missing_sheets/<int:m_id>', methods=['DELETE'])
@login_required
def delete_missing_sheet(m_id):
    db = get_db()
    db.execute("DELETE FROM excel_missing_sheets WHERE id = ? AND user_id = ?", (m_id, session['user_id']))
    db.commit()
    return jsonify({'success': True})

@app.route('/api/excel/parse_latest/<chat_id>', methods=['POST'])
@login_required
def parse_latest_excel_route(chat_id):
    chat_id = int(chat_id) # Добавляем преобразование здесь
    user_id = session['user_id']    
    # Ищем активную сессию юзербота
    client_to_use = None
    client_loop_to_use = None
    for sid, (thread, client, uid) in user_clients.items():
        if uid == user_id:
            client_to_use = client
            client_loop_to_use = user_loops.get(sid)
            break
            
    if not client_to_use:
        return jsonify({'error': 'Юзербот не запущен. Запустите его во вкладке "Юзерботы".'}), 400

    async def process_latest_excel():
        try:
            with app.app_context():
                db = get_db()
                tc = db.execute("SELECT chat_title, custom_name FROM tracked_chats WHERE chat_id = ? AND user_id = ?", (chat_id, user_id)).fetchone()
                chat_title = tc['custom_name'] if tc and tc['custom_name'] else (tc['chat_title'] if tc else str(chat_id))

            entity = await client_to_use.get_entity(chat_id)
            
            # Ищем последние 100 сообщений, пока не наткнемся на Excel
            async for message in client_to_use.iter_messages(entity, limit=100):
                if message.document and getattr(message.file, 'ext', '').lower() in ['.xlsx', '.xls', '.pdf']:
                    
                    parsed_text = await parse_excel_message(client_to_use, message, chat_id, user_id)
                    
                    if parsed_text:
                        msg_type = 'channel' if message.is_channel else 'group' if message.is_group else 'private'
                        sender = await message.get_sender()
                        sender_name = f"{getattr(sender, 'first_name', '')} {getattr(sender, 'last_name', '')}".strip() or "Unknown"
                        msg_date = message.date.strftime("%Y-%m-%d %H:%M:%S")

                        with app.app_context():
                            db = get_db()
                            try:
                                db.execute(
                                    'INSERT INTO messages (user_id, telegram_message_id, type, text, date, chat_id, chat_title, sender_name) VALUES (?, ?, ?, ?, ?, ?, ?, ?)',
                                    (user_id, message.id, msg_type, parsed_text, msg_date, chat_id, chat_title, sender_name)
                                )
                            except sqlite3.IntegrityError:
                                # ИСПРАВЛЕНИЕ: Убираем chat_id из WHERE, так как он может отличаться из-за приставки -100
                                db.execute(
                                    "UPDATE messages SET text = ? WHERE telegram_message_id = ? AND user_id = ?", 
                                    (parsed_text, message.id, user_id)
                                )
                            db.commit()
                            notify_clients()
                            logger.info(f"✅ Успешно спарсили последний Excel файл из {chat_title}")
                    
                    break # Останавливаем поиск после первого найденного файла
                    
        except Exception as e:
            logger.error(f"Ошибка ручного парсинга Excel: {e}")

    # Запускаем асинхронную задачу в фоне
    asyncio.run_coroutine_threadsafe(process_latest_excel(), client_loop_to_use)
    
    return jsonify({'success': True})



@app.route('/api/tracked_chats', methods=['GET'])
@login_required
def get_tracked_chats():
    user_id = session['user_id']
    db = get_db()
    # ДОБАВЛЕНО: custom_name в выборку
    cursor = db.execute("SELECT id, chat_id, chat_title, custom_name FROM tracked_chats WHERE user_id=?", (user_id,))
    chats = [dict(row) for row in cursor.fetchall()]
    return jsonify(chats)

@app.route('/api/tracked_chats', methods=['POST'])
@login_required
def add_tracked_chat():
    data = request.get_json()
    raw_chat_id = data.get('chat_id')
    custom_name = data.get('custom_name') # Получаем кастомное название
    user_id = session['user_id']
  
    if not raw_chat_id:
        return jsonify({'error': 'chat_id required'}), 400

    chat_title = custom_name # Изначально ставим кастомное имя (если оно пустое, ниже получим из ТГ)
    numeric_chat_id = None
    client_to_use = None
    client_loop_to_use = None

    try:
        entity_query = int(raw_chat_id)
    except ValueError:
        entity_query = raw_chat_id

    # Ищем активную сессию юзербота для текущего пользователя
    # Ищем активные сессии юзерботов для текущего пользователя
    active_sessions = [sid for sid, (_, _, uid) in user_clients.items() if uid == user_id]
    
    # ПЕРЕБИРАЕМ ВСЕХ БОТОВ, ПОКА НЕ НАЙДЕМ ТОГО, У КОТОРОГО ЕСТЬ ДОСТУП
    for sid in active_sessions:
        thread, client, uid = user_clients[sid]
        client_loop = user_loops.get(sid)
        
        async def fetch_chat():
            return await client.get_entity(entity_query)
            
        if client_loop:
            try:
                future = asyncio.run_coroutine_threadsafe(fetch_chat(), client_loop) 
                entity = future.result(timeout=5)
                numeric_chat_id = entity.id 
                
                # Если доступ есть, запоминаем этого бота как рабочего для данного чата
                client_to_use = client
                client_loop_to_use = client_loop
                
                # Записываем название чата
                if not chat_title:
                    if hasattr(entity, 'title'):
                        chat_title = entity.title
                    else:
                        chat_title = f"{entity.first_name or ''} {entity.last_name or ''}".strip()
                        
                break # Успешно нашли нужного бота, прерываем цикл перебора!
                
            except Exception as e:
                logging.warning(f"Бот сессии {sid} не имеет доступа к чату или произошла ошибка: {e}")
                continue # Пробуем следующего бота
                
                # Если кастомное имя не ввели, берем оригинальное название из Telegram
                if not chat_title:
                    if hasattr(entity, 'title'):
                        chat_title = entity.title
                    else:
                        chat_title = f"{entity.first_name or ''} {entity.last_name or ''}".strip()
            except Exception as e:
                logging.warning(f"Не удалось получить информацию о чате: {e}")

    # Фолбэк, если юзербот не запущен или произошла ошибка
    if not numeric_chat_id:
        try:
            numeric_chat_id = int(raw_chat_id)
        except ValueError:
            return jsonify({'error': 'Запустите юзербота, чтобы добавлять чаты по ссылке, или введите точный числовой ID'}), 400

    # Если название всё ещё пустое (например, юзербот выключен и ввели просто ID без кастомного имени)
    if not chat_title:
        chat_title = str(numeric_chat_id)

    db = get_db()
    try:
        db.execute(
            "INSERT INTO tracked_chats (user_id, chat_id, chat_title, custom_name) VALUES (?, ?, ?, ?)",
            (user_id, numeric_chat_id, chat_title, custom_name)
        )
        db.commit()
        notify_clients()
        
        # Запускаем парсинг истории сообщений в фоне, если юзербот активен
        # Запускаем парсинг истории сообщений в фоне, если юзербот активен
        if client_to_use and client_loop_to_use and numeric_chat_id:
            target_to_parse = entity if 'entity' in locals() else numeric_chat_id
            asyncio.run_coroutine_threadsafe(
                fetch_chat_history(client_to_use, target_to_parse, chat_title, user_id, numeric_chat_id), 
                client_loop_to_use
            )

        return jsonify({'success': True, 'chat_title': chat_title})
    except sqlite3.IntegrityError:
        return jsonify({'error': 'Этот чат уже добавлен в список отслеживаемых'}), 400

@app.route('/api/tracked_chats/<item_id>', methods=['DELETE'])
def delete_tracked_chat(item_id):
    try:
        db = get_db()
        user_id = session.get('user_id')
        
        # 1. Узнаем настоящий Telegram ID этого чата (например, -1001234567) по его ID в таблице (например, 43)
        row = db.execute("SELECT chat_id FROM tracked_chats WHERE id=? AND user_id=?", (item_id, user_id)).fetchone()
        
        if row:
            real_chat_id = row['chat_id']
            
            # 2. Удаляем привязки цен к товарам, которые пришли из этого чата
            db.execute("DELETE FROM product_messages WHERE message_id IN (SELECT id FROM messages WHERE chat_id=? AND user_id=?)", (real_chat_id, user_id))
            
            # 3. Удаляем сами спарсенные сообщения из этого чата
            db.execute("DELETE FROM messages WHERE chat_id=? AND user_id=?", (real_chat_id, user_id))
            
        # 4. Удаляем сам чат из списка отслеживаемых (по его внутреннему номеру)
        db.execute("DELETE FROM tracked_chats WHERE id=? AND user_id=?", (item_id, user_id))
        
        db.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.rollback()
        print(f"Ошибка при удалении чата: {e}")
        return jsonify({'error': str(e)}), 500

# ---------- Административные маршруты ----------
@app.route('/api/admin/users')
@admin_required
def admin_users():
    db = get_db()
    cursor = db.execute("SELECT id, login, role FROM users")
    users = [{'id': row['id'], 'login': row['login'], 'role': row['role']} for row in cursor.fetchall()]
    return jsonify(users)

@app.route('/api/admin/add_user', methods=['POST'])
@admin_required
def admin_add_user():
    data = request.get_json()
    login = data.get('login')
    password = data.get('password')
    if not login or not password:
        return jsonify({'error': 'Login and password required'}), 400
    pwd_hash = hashlib.sha256(password.encode()).hexdigest()
    db = get_db()
    try:
        db.execute("INSERT INTO users (login, password_hash, role) VALUES (?, ?, 'user')", (login, pwd_hash))
        new_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
        db.execute("INSERT INTO folders (user_id, name, parent_id) VALUES (?, 'По умолчанию', NULL)", (new_id,))
        db.commit()
        notify_clients()
        return jsonify({'success': True})
    except sqlite3.IntegrityError:
        return jsonify({'error': 'Login already exists'}), 400




@app.route('/api/admin/clear_messages', methods=['POST'])
@admin_required
def admin_clear_messages():
    data = request.get_json()
    period = data.get('period', 'all')
    db = get_db()
    
    try:
        if period == 'all':
            db.execute("DELETE FROM messages")
        else:
            days = int(period)
            cutoff_date = datetime.now() - timedelta(days=days)
            cutoff_str = cutoff_date.strftime("%Y-%m-%d %H:%M:%S")
            db.execute("DELETE FROM messages WHERE date < ?", (cutoff_str,))
        
        db.commit()
        notify_clients() 
        return jsonify({'success': True})
    except Exception as e:
        logger.error(f"Ошибка при очистке сообщений: {e}")
        return jsonify({'error': str(e)}), 500



# ---------- Шаблоны страниц ----------
LOGIN_TEMPLATE = """
<!DOCTYPE html>
<html lang="ru">
<head>
<link rel="icon" type="image/svg+xml" href="/logo.svg">
<script src="https://cdnjs.cloudflare.com/ajax/libs/socket.io/4.0.1/socket.io.js"></script>
    <meta charset="UTF-8">
    <title>Вход в систему</title>
    <style>
        body { background: #1a1a1a; color: #e0e0e0; font-family: 'Segoe UI', Tahoma, Arial, sans-serif; display: flex; justify-content: center; align-items: center; height: 100vh; margin: 0; }
        .login-box { background: #2a2a2a; padding: 30px; border-radius: 8px; border: 1px solid #3a3a3a; text-align: center; width: 100%; max-width: 320px; box-shadow: 0 4px 15px rgba(0,0,0,0.5); }
        .login-box h2 { margin-bottom: 20px; color: #fff; }
        input { width: 100%; padding: 12px; margin: 8px 0 16px 0; background: #1e1e1e; border: 1px solid #3a3a3a; color: #fff; border-radius: 4px; box-sizing: border-box; font-size: 14px; }
        input:focus { outline: none; border-color: #4a90e2; }
        button { width: 100%; padding: 12px; background: #4a90e2; color: white; border: none; border-radius: 4px; cursor: pointer; font-weight: 600; font-size: 15px; transition: 0.2s; }
        button:hover { background: #357abd; }
        .error { color: #ff5555; background: rgba(255, 85, 85, 0.1); padding: 10px; border-radius: 4px; margin-bottom: 15px; font-size: 14px; border: 1px solid rgba(255, 85, 85, 0.3); }
        
        /* Новые стили для поля с паролем */
        .password-container { position: relative; margin: 8px 0 16px 0; }
        .password-container input { margin: 0; padding-right: 40px; } /* Отступ справа, чтобы текст не залезал под иконку */
        .toggle-password { position: absolute; right: 12px; top: 50%; transform: translateY(-50%); cursor: pointer; color: #888; font-size: 18px; user-select: none; }
    </style>
</head>
<body>
    <div class="login-box">
        <h2>Вход в систему</h2>
        {% if error %}
            <div class="error">{{ error }}</div>
        {% endif %}
        <form action="/login" method="POST">
            <input type="text" name="login" placeholder="Логин" required>
            
            <div class="password-container">
                <input type="password" id="login-password" name="password" placeholder="Пароль" required>
                <span class="toggle-password" onclick="togglePassword(this)"></span>
            </div>

            <button type="submit">Войти</button>
        </form>
    </div>

    <script>
        function togglePassword(iconElement) {
            const passwordInput = document.getElementById('login-password');
            if (passwordInput.type === 'password') {
                passwordInput.type = 'text';
                iconElement.innerText = '🙈'; // Меняем иконку на закрытый глаз
            } else {
                passwordInput.type = 'password';
                iconElement.innerText = '👁️'; // Меняем иконку на открытый глаз
            }
        }
    </script>
</body>
</html>
"""

TEMPLATE = """
<!DOCTYPE html>
<html lang="ru">
<head>
<link rel="icon" type="image/svg+xml" href="/logo.svg">
<script src="https://cdnjs.cloudflare.com/ajax/libs/socket.io/4.0.1/socket.io.js"></script>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Настройка ценообразования</title>
    <style>


.api-folder-header {
        display: flex;
        align-items: center;
        background: #2a2a2a;
        padding: 10px;
        border-radius: 6px;
        cursor: pointer;
        margin-bottom: 5px;
        border: 1px solid #3a3a3a;
        transition: background 0.2s;
    }
    .api-folder-header:hover {
        background: #333;
    }
    .api-folder-content {
        padding-left: 30px;
        margin-bottom: 15px;
        border-left: 2px solid #2a2a2a;
        margin-left: 15px;
    }
    .api-folder-toggle-icon {
        margin-left: auto;
        color: #aaa;
        font-size: 12px;
    }
    /* Стили для дерева папок */
.folder-tree ul {
    display: none; /* По умолчанию скрываем вложенные папки */
    list-style: none;
    padding-left: 20px;
    margin: 5px 0;
}

.folder-tree li.expanded > ul {
    display: block; /* Показываем, если у родителя есть класс expanded */
}

.folder-toggle {
    cursor: pointer;
    display: inline-block;
    width: 20px;
    transition: transform 0.2s;
    color: #666;
}

.folder-tree li.expanded > div .folder-toggle {
    transform: rotate(90deg); /* Поворачиваем стрелочку при раскрытии */
}

.folder-name:hover {
    background: #3d3d3d;
}

.folder-tree li.active > div .folder-name {
    color: #4a90e2;
    background: #34495e;
}


        * { margin: 0; padding: 0; box-sizing: border-box; font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; }
        body { 
    background: linear-gradient(135deg, #14161a 0%, #222630 100%); /* Новый темный градиент */
    color: #e0e0e0; 
    min-height: 100vh; 
    display: flex; 
    flex-direction: column; 
    margin: 0;
}
        .top-menu-wrapper {
    background: #2a2a2a; 
    border-bottom: 1px solid #3a3a3a; 
    position: sticky; top: 0; z-index: 100;
}
.top-menu { 
    padding: 0 20px; 
    display: flex; 
    align-items: center; 
    justify-content: space-between; 
    flex-wrap: wrap; 
    max-width: 1400px; /* Такая же ширина, как у контента */
    margin: 0 auto;    /* Центрируем */
    width: 100%;
    box-sizing: border-box;
}
        .menu-items { display: flex; list-style: none; gap: 8px; flex-wrap: wrap; }
        .menu-items li { padding: 16px 12px; font-weight: 500; color: #b0b0b0; cursor: pointer; border-bottom: 3px solid transparent; transition: 0.2s; white-space: nowrap; }
        .menu-items li.active { color: #fff; border-bottom-color: #4a90e2; }
        .menu-items li:not(.active):hover { background: #333; border-radius: 4px; }
        .logout-btn { padding: 8px 16px; background: #333; border-radius: 20px; font-weight: 600; cursor: pointer; color: #ccc; border: 1px solid #4a4a4a; }
        .logout-btn:hover { background: #3a3a3a; }
        .content-area { 
    padding: 20px; 
    background: transparent; /* Убрал сплошной фон, чтобы было видно градиент страницы */
    flex: 1; 
    display: flex; 
    gap: 20px; 
    max-width: 1400px; /* Ограничиваем ширину контейнера */
    margin: 0 auto;    /* Центрируем по горизонтали */
    width: 100%; 
    box-sizing: border-box;
}
.sidebar { 
            width: max-content; /* АВТОМАТИЧЕСКИ растягивается под самый длинный текст */
            min-width: 200px; 
            max-width: 50vw; /* Защита: не больше половины экрана */
            background: #2a2a2a; border-radius: 8px; padding: 16px;
            border: 1px solid #3a3a3a; height: fit-content; 
            overflow-x: auto; /* Добавляет скролл, если название ну очень длинное */
        }
        .sidebar h3 { color: #ccc; margin-bottom: 12px; font-size: 16px; }
        .folder-tree { list-style: none; margin-bottom: 16px; 
            white-space: nowrap; /* ЗАПРЕЩАЕТ ПЕРЕНОС ДЛИННОГО ТЕКСТА */
        }
        .folder-tree li { padding: 6px 0 6px 16px; cursor: pointer; color: #ddd; position: relative; }
        .folder-tree li .folder-name { display: inline-block; padding: 4px 8px; border-radius: 4px; }
        .folder-tree li .folder-name:hover { background: #3a3a3a; }
        .folder-tree li.active > div > .folder-name { background: #4a90e2; color: white; }
        .folder-tree .children { list-style: none; padding-left: 16px; }
        .folder-actions { display: inline-block; margin-left: 8px; }
        .folder-actions span { cursor: pointer; padding: 2px 6px; border-radius: 4px; font-size: 12px; }
        .folder-actions span:hover { background: #4a4a4a; }
        .new-folder-btn { width: 100%; padding: 10px; background: #333; border: 1px solid #4a4a4a; color: #ccc; border-radius: 6px; cursor: pointer; font-weight: 600; margin-top: 8px; }
        .new-folder-btn:hover { background: #3a3a3a; }
        .main-content { flex: 1; background: #2a2a2a; border-radius: 8px; padding: 20px; border: 1px solid #3a3a3a; }
        .header-info { display: flex; justify-content: space-between; align-items: center; margin-bottom: 20px; color: #aaa; font-size: 14px; }
        .add-bot-btn { background: #333; color: #fff; border: 1px solid #4a4a4a; padding: 8px 16px; border-radius: 20px; font-weight: 600; cursor: pointer; font-size: 14px; }
        .add-bot-btn:hover { background: #3a3a3a; }
        .bot-form-container { 
    position: fixed; 
    top: 50%; 
    left: 50%; 
    transform: translate(-50%, -50%); 
    background: #2a2a2a; 
    border: 1px solid #4a90e2; 
    border-radius: 8px; 
    padding: 20px; 
    z-index: 1000; 
    width: 320px; 
    box-shadow: 0 10px 30px rgba(0,0,0,0.7); 
    display: none; 
}
/* Делаем поля ввода в столбик для красоты */
.bot-form { display: flex; flex-direction: column; gap: 12px; }
.bot-form .form-group { width: 100%; }
.bot-form input { width: 100%; box-sizing: border-box; }
        .bot-form { display: flex; flex-wrap: wrap; gap: 12px; align-items: flex-end; }
        .form-group { display: flex; flex-direction: column; gap: 4px; }
        .form-group label { font-size: 12px; color: #aaa; }
        .form-group input { padding: 8px; background: #1e1e1e; border: 1px solid #3a3a3a; border-radius: 4px; color: #fff; font-size: 14px; min-width: 180px; }
        .qr-btn { background: #333; color: white; border: 1px solid #4a4a4a; padding: 8px 16px; border-radius: 20px; font-weight: 600; cursor: pointer; height: 36px; }
        .qr-btn:hover { background: #3a3a3a; }
        .qr-container { margin-top: 16px; display: none; background: #1e1e1e; padding: 16px; border-radius: 8px; border: 1px solid #3a3a3a; text-align: center; }
        .qr-container img { max-width: 150px; border-radius: 8px; }
        .status-message { margin-top: 8px; color: #ffaa00; }
        .filters-panel { background: #2a2a2a; padding: 16px; border-radius: 8px; margin-bottom: 20px; border: 1px solid #3a3a3a; }
        .filter-row { display: flex; gap: 20px; flex-wrap: wrap; }
        .filter-group { flex: 1; min-width: 200px; }
        .filter-group label { font-size: 13px; font-weight: 600; color: #aaa; display: block; margin-bottom: 4px; }
        .filter-group input, .filter-group select { width: 100%; padding: 8px; background: #1e1e1e; border: 1px solid #3a3a3a; border-radius: 4px; color: #fff; }
        .tags-container { display: flex; flex-wrap: wrap; gap: 6px; margin-top: 8px; }
        .tag { background: #333; color: #ccc; padding: 4px 10px; border-radius: 20px; font-size: 13px; display: inline-flex; align-items: center; gap: 6px; border: 1px solid #4a4a4a; }
        .tag .remove-tag { cursor: pointer; font-weight: bold; color: #888; }
        .tag .remove-tag:hover { color: #ff5555; }
        .apply-btn { background: #333; color: white; border: 1px solid #4a4a4a; padding: 8px 24px; border-radius: 20px; font-weight: 600; cursor: pointer; height: 36px; align-self: flex-end; }
        .apply-btn:hover { background: #3a3a3a; }
        .action-bar { display: flex; gap: 10px; margin-bottom: 16px; align-items: center; }
        .save-btn { background: #333; color: white; border: 1px solid #4a4a4a; padding: 8px 16px; border-radius: 20px; font-weight: 600; cursor: pointer; }
        .save-btn:hover { background: #3a3a3a; }
        .messages-table { width: 100%; border-collapse: collapse; background: #1e1e1e; border-radius: 8px; overflow: hidden; }
        .messages-table th { background: #333; color: #ccc; font-weight: 600; font-size: 13px; padding: 12px 8px; text-align: left; border-bottom: 1px solid #4a4a4a; }
        .messages-table td { padding: 10px 8px; border-bottom: 1px solid #3a3a3a; color: #ddd; }
        .messages-table tr:hover { background: #2a2a2a; }
        .no-messages { text-align: center; padding: 40px; color: #777; background: #1e1e1e; border-radius: 8px; }
        table.simple-table { width: 100%; border-collapse: collapse; background: #2a2a2a; }
        table.simple-table th { background: #333; color: #ccc; padding: 8px; text-align: left; }
        table.simple-table td { padding: 8px; border-bottom: 1px solid #3a3a3a; }
    </style>
</head>
<body>
<div style="display: flex; align-items: center; gap: 10px; margin-bottom: 20px;">
  
    <h1 style="margin: 0; padding-left: 15px;;">Панель управления</h1>
</div>
    <div class="top-menu-wrapper">
    <div class="top-menu">
        <ul class="menu-items" id="main-menu">
    <li data-section="parser" class="active">Склад</li>
    <li data-section="userbots">Подключение</li>
    <li data-section="interaction_bots">Подключение Ботов</li>
    <li data-section="products">Товары </li>
    <li data-section="excel">Подключение файлов</li>
    <li data-section="api">API</li>
    <li data-section="telegram">Telegram</li>
     <li data-section="profile">Мой профиль</li>
    {% if session.role == 'admin' %} 
    <li data-section="admin">Администрирование </li>
    {% endif %} 
</ul>
        <div class="logout-btn" onclick="window.location.href='/logout'">Выйти</div>
    </div>
</div>

    <div class="content-area" id="content-area">
        <div class="sidebar" id="parser-sidebar" style="display: block;">
            <div style="display: flex; align-items: center; justify-content: space-between; margin-bottom: 12px;">
                <h3 style="margin:0;">📁 Папки</h3>
                <span id="toggle-sidebar" style="cursor:pointer; font-size:18px; padding:4px;">◀</span>
            </div>
              <div id="folder-tree-container">
            <ul class="folder-tree" id="folder-tree"></ul>
            <button class="new-folder-btn" id="new-folder-btn">+ Новая папка</button>
        </div>
    </div> <div class="main-content" style="display: flex; flex-direction: column; flex: 1; background: transparent; border: none; padding: 0;">

 <div id="bot-form-container" class="bot-form-container">
    <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 15px;">
        <h3 style="margin: 0; color: #fff; font-size: 18px;">Вход в аккаунт</h3>
        <button onclick="document.getElementById('bot-form-container').style.display='none'" style="background: none; border: none; color: #aaa; cursor: pointer; font-size: 16px;">❌</button>
    </div>
    
    <div class="bot-form">
        <input type="hidden" id="auth_session_id"> <div class="form-group">
            <label>API ID</label>
            <input type="text" id="api_id" placeholder="12345" value="12345">
        </div>
        <div class="form-group">
            <label>API Hash</label>
            <input type="text" id="api_hash" placeholder="abc123..." value="abc123">
        </div>
        <button class="qr-btn" id="qr-btn" style="margin-top: 10px;">Получить QR для входа</button>
    </div>
    <div id="qr-container" class="qr-container">
        <div id="qr-image"></div>
        <div id="qr-status" class="status-message">Отсканируйте QR в Telegram (Настройки -> Устройства -> Сканировать QR)</div>
    </div>
</div>
            <div id="interaction_bots-section" style="display: none;">
                <h2>Настройка ботов (Автоматизация)</h2>
                <div class="card" style="background:#1e1e1e; padding:16px; border-radius:8px; margin-bottom:16px;">
                    <form id="add-interaction-bot-form" style="display:flex; flex-wrap:wrap; gap:10px; align-items:flex-end;">
                        <div class="form-group">
                            <label>Название (для вас)</label>
                            <input type="text" id="ib_custom_name" placeholder="Мой бот 1" required>
                        </div>
                        <div class="form-group">
                            <label>Юзербот (от чьего имени писать)</label>
                            <select id="ib_userbot_id" style="padding:8px; background:#2a2a2a; color:#fff; border:1px solid #3a3a3a;"></select>
                        </div>
                        <div class="form-group">
                            <label>Username бота (например, @MyBot)</label>
                            <input type="text" id="ib_bot_username" required>
                        </div>
                        <div class="form-group">
                            <label>Команды (через запятую: /start, /balance)</label>
                            <input type="text" id="ib_commands" required>
                        </div>
                        <div class="form-group">
                            <label>Интервал (минут)</label>
                            <input type="number" id="ib_interval" value="60" min="1" required style="width:80px;">
                        </div>
                        <button type="submit" class="save-btn">Добавить</button>
                    </form>
                </div>
                <table class="messages-table">
                    <thead><tr><th>Юзербот</th><th>Бот</th><th>Команды</th><th>Интервал (мин)</th><th>Действия</th></tr></thead>
                    <tbody id="interaction-bots-tbody"></tbody>
                </table>
            </div>

        <div id="products-section" style="display: none;">
    <div style="display: flex; gap: 20px; align-items: flex-start;">
        <div class="sidebar" id="products-sidebar" style="width: max-content; min-width: 200px; max-width: 50vw; background: #2a2a2a; border-radius: 8px; padding: 16px; border: 1px solid #3a3a3a; overflow-x: auto;">
            <h3 style="color: #ccc; margin-bottom: 12px; font-size: 16px;">📁 Категории товаров</h3>
            <ul class="folder-tree" id="products-folder-tree"></ul>
        </div>

        <div class="main-content-inner" style="flex: 1; background: #2a2a2a; border-radius: 8px; padding: 20px; border: 1px solid #3a3a3a;">
            <h2 id="products-main-title">Управление товарами</h2>
            <div id="selected-product-folder-title" style="margin-bottom: 15px; color: #4a90e2; font-weight: bold; font-size: 14px;">📍 Папка: Все товары</div>
            
            <input type="text" id="product-search" placeholder="Поиск по названию или синонимам..." style="width:100%; padding:10px; margin-bottom: 16px; background:#1e1e1e; border:1px solid #3a3a3a; color:#fff; border-radius:4px;">
            
            <div class="card" style="background:#1e1e1e; padding:16px; border-radius:8px; margin-bottom:16px; border: 1px solid #333;">
                <form id="add-product-form" style="display:flex; flex-wrap:wrap; gap:10px; align-items:flex-end;">
                    <div class="form-group">
                        <label style="font-size: 12px; color: #aaa;">Название товара</label>
                        <input type="text" id="prod_name" required style="padding: 8px; background: #2a2a2a; border: 1px solid #3a3a3a; color: #fff; border-radius: 4px;">
                    </div>
                    <div class="form-group">
                        <label style="font-size: 12px; color: #aaa;">Синонимы (через запятую)</label>
                        <input type="text" id="prod_synonyms" placeholder="айфон, iPhone" style="padding: 8px; background: #2a2a2a; border: 1px solid #3a3a3a; color: #fff; border-radius: 4px;">
                    </div>
                    <div class="form-group">
                        <label style="font-size: 12px; color: #aaa;">Папка (Каталог)</label>
                        <select id="prod_folder_id" style="padding:8px; background:#2a2a2a; border:1px solid #3a3a3a; color:#fff; border-radius:4px;">
                            <option value="">Без папки (Общие)</option>
                        </select>
                    </div>
                    <button type="submit" class="save-btn" style="height: 36px;">Добавить</button>
                </form>
            </div>

            <table class="messages-table">
                <thead><tr><th>Название</th><th>Синонимы</th><th>Действия</th></tr></thead>
                <tbody id="products-tbody"></tbody>
            </table>

            <div id="product-binding-container" style="display: none; margin-top: 30px; border-top: 2px solid #3498db; padding-top: 20px;">
                <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 15px;">
                    <h3 id="binding-product-title" style="color: #4a90e2; margin: 0;">Привязка сообщений: </h3>
                    <button class="save-btn" onclick="closeProductBinding()" style="background: #444;">❌ Закрыть</button>
                </div>
                <h4 style="color: #ccc; margin-bottom: 10px;">🔍 Предложенные совпадения</h4>
                <table class="messages-table" style="margin-bottom: 30px;">
                    <thead><tr><th>От кого</th><th>Откуда</th><th>Строка</th><th>Цена</th><th>Действие</th></tr></thead>
                    <tbody id="proposed-messages-tbody"></tbody>
                </table>
                <h4 style="color: #ccc; margin-bottom: 10px;">✅ Подтвержденные сообщения</h4>
                <table class="messages-table">
                    <thead><tr>
                        <th style="width: 40px;"><input type="checkbox" id="select-all-conf" onchange="toggleAllConf(this)"></th>
                        <th>От кого</th><th>Откуда</th><th>Текст</th><th>Цена</th><th>Действия</th>
                    </tr></thead>
                    <tbody id="confirmed-messages-tbody"></tbody>
                </table>
                <button class="save-btn" onclick="mergeSelectedConfirmed()" style="margin-top: 15px; background: #3498db; display: none; width: 100%;" id="btn-merge-conf">🔗 Объединить выбранные</button>
            </div>
        </div>
    </div>
</div>

<div id="edit-modal" style="display:none; position:fixed; top:50%; left:50%; transform:translate(-50%, -50%); background:#2a2a2a; padding:20px; border-radius:8px; z-index:1000; border:1px solid #f39c12; width: 400px; box-shadow: 0 10px 30px rgba(0,0,0,0.5);">
<div id="move-folder-modal" style="display:none; position:fixed; top:50%; left:50%; transform:translate(-50%, -50%); background:#2a2a2a; padding:20px; border-radius:8px; z-index:1000; border:1px solid #3498db; width: 400px; box-shadow: 0 10px 30px rgba(0,0,0,0.5);">
    <h3 style="margin-bottom:15px; color:#fff;">Настроить папку: <br><span id="move-prod-name" style="color:#3498db; font-size: 16px;"></span></h3>
    <div class="form-group" style="margin-bottom:15px;">
        <label style="color:#aaa;">Выберите новую папку</label>
        <select id="move-folder-select" style="width:100%; padding:8px; background:#1e1e1e; border:1px solid #3a3a3a; color:#fff; border-radius:4px;">
        </select>
    </div>
    <div style="display:flex; gap:10px;">
        <button onclick="saveProductFolder()" style="flex:1; padding:10px; background:#3498db; color:#fff; border:none; cursor:pointer; border-radius:4px; font-weight:bold;">💾 Сохранить</button>
        <button onclick="document.getElementById('move-folder-modal').style.display='none'" style="flex:1; padding:10px; background:#4a4a4a; color:#fff; border:none; cursor:pointer; border-radius:4px;">❌ Отмена</button>
    </div>
</div>

<h3 style="margin-bottom:15px; color:#fff;">Редактировать товар</h3>
    <div class="form-group" style="margin-bottom:10px;">
        <label style="color:#aaa;">Название</label>
        <input type="text" id="edit-prod-name" style="width:100%; padding:8px; background:#1e1e1e; border:1px solid #3a3a3a; color:#fff; border-radius:4px;">
    </div>
    <div class="form-group" style="margin-bottom:15px;">
        <label style="color:#aaa;">Синонимы (через запятую)</label>
        <input type="text" id="edit-prod-synonyms" style="width:100%; padding:8px; background:#1e1e1e; border:1px solid #3a3a3a; color:#fff; border-radius:4px;">
    </div>
    <div style="display:flex; gap:10px;">
        <button onclick="saveEditProduct()" style="flex:1; padding:10px; background:#27ae60; color:#fff; border:none; cursor:pointer; border-radius:4px; font-weight:bold;">💾 Сохранить</button>
        <button onclick="document.getElementById('edit-modal').style.display='none'" style="flex:1; padding:10px; background:#4a4a4a; color:#fff; border:none; cursor:pointer; border-radius:4px;">❌ Отмена</button>
    </div>
</div>



            
            <div id="parser-section" style="display: block;">
                <div class="header-info">
                    <span id="current-datetime"></span>
                    <button class="add-bot-btn" id="add-bot-btn" style="display: none;">➕ Добавить акаунт</button>
                </div>

                
                <div class="filters-panel">
    <div class="filter-row">
        <div class="filter-group">
            <label>Тип сообщений</label>
            <select id="filter-type">
                <option value="all">Все типы</option>
                <option value="group">Группа</option>
                <option value="channel">Канал</option>
                <option value="private">Боты</option>
            </select>
        </div>
        <div class="filter-group">
            <label>От кого</label>
            <select id="filter-sender">
                <option value="all">Все отправители</option>
            </select>
        </div>
        <div class="filter-group">
            <label>Поиск</label>
            <input type="text" id="filter-search-input" placeholder="Введите слово и нажмите Enter">
            <div id="search-tags" class="tags-container"></div>
        </div>
        <div class="filter-group">
            <label>Слова исключения</label>
            <input type="text" id="filter-exclude-input" placeholder="Введите слово и нажмите Enter">
            <div id="exclude-tags" class="tags-container"></div>
        </div>
        <button class="apply-btn" id="apply-filters">Применить</button>
    </div>
</div>
                

                <table class="messages-table">
                    <thead>
                        <tr>
                             <th>Тип</th>
        <th>Чат / От кого</th>
        <th>Сообщение</th>
        <th>Дата</th>
        <th>Действия</th>
                        </tr>
                    </thead>
                    <tbody id="messages-tbody"></tbody>
                </table>
                <div id="pagination-container" style="margin-top: 20px;"></div>
                <div id="no-messages" class="no-messages" style="display: none;">😕 Нет сообщений</div>
            </div>
            <div id="folder-catalog-section" style="display: none;">
                <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 20px;">
                    <h2 id="fc-title">Папка</h2>
                </div>

                <div id="fc-link-block" class="card" style="background:#1e1e1e; padding:16px; border-radius:8px; margin-bottom:16px;">
                    <h4 style="margin-bottom: 10px; color: #ccc;">🔗 Привязать существующий товар к этой папке</h4>
                    <div style="display:flex; gap:10px;">
                        <select id="fc-unlinked-products" style="flex:1; padding:8px; background:#2a2a2a; border:1px solid #3a3a3a; color:#fff; border-radius:4px;"></select>
                        <button class="save-btn" onclick="linkProductToFolder()">Привязать товар</button>
                    </div>
                </div>

                

                <div id="fc-reports-block" style="display: none; margin-top: 30px;">
                  
                    <div class="filters-panel" style="margin-bottom: 15px;">
                        <div class="form-group" style="width: 250px;">
                            <label style="color:#aaa; font-size:12px; margin-bottom:4px; display:block;">Сортировка цен</label>
                            <select id="fc-report-sort" style="width: 100%; padding: 8px; background: #1e1e1e; border: 1px solid #3a3a3a; color: #fff; border-radius: 4px;" onchange="renderFCReports()">
                                <option value="date_desc">Сначала новые</option>
                                <option value="price_asc">Сначала дешевле (Лучшая цена)</option>
                                <option value="price_desc">Сначала дороже</option>
                                <option value="seller_asc">По продавцу (А-Я)</option>
                                <option value="name_asc">По названию (А-Я)</option>
                            </select>
                        </div>
                    </div>
                    <table class="messages-table">
                        <thead>
                            <tr>
                            <th>ID</th>
                                <th>Товар</th>
                                <th>Цена</th>
                                <th>Продавец</th>
                                <th>Источник</th>
                                <th>Сообщение</th>
                                <th>Дата</th>
                            </tr>
                        </thead>
                        <tbody id="fc-reports-tbody"></tbody>
                    </table>
                </div>
            </div>
            <div id="userbots-section" style="display: none;">
                <h2>Управление акаунтами</h2>
                <button id="add-userbot-btn" class="save-btn">➕ Добавить акаунт</button>
                <div id="userbot-list" style="margin-top:16px;"></div>
                <hr style="border-color:#3a3a3a; margin:20px 0;">
                <h3>Отслеживаемые чаты</h3>
                <div style="display:flex; gap:8px; margin-bottom:16px;">
                    <input type="text" id="new-chat-id" placeholder="ID или ссылка" style="flex:1; background:#1e1e1e; border:1px solid #3a3a3a; color:#fff; padding:8px;">
                    <input type="text" id="new-chat-name" placeholder="Название (для вас)" style="flex:1; background:#1e1e1e; border:1px solid #3a3a3a; color:#fff; padding:8px;">
                    <button id="add-chat-btn" class="save-btn">Добавить</button>
                </div>
                <ul id="tracked-chats-list" style="list-style:none;"></ul>

                <div class="card" style="background:#1e1e1e; padding:16px; border-radius:8px; margin-top:20px;">
    <h3 style="color:#3498db; margin-bottom:10px;">🔗 Подключение к другому API </h3>
    <p style="color:#aaa; font-size:12px; margin-bottom:15px;">Скачивание прайсов по ссылке с других таких же систем.</p>
    
    <form id="api-source-form" style="display:flex; gap:10px; flex-wrap:wrap; margin-bottom: 15px;">
        <input type="text" id="api_source_name" placeholder="Название (например: Партнер 1)" style="flex:1; min-width:150px; padding:8px; background:#2a2a2a; border:1px solid #3a3a3a; color:#fff; border-radius:4px;">
        <input type="text" id="api_source_url" placeholder="Ссылка (http://.../api/catalog)" style="flex:2; min-width:200px; padding:8px; background:#2a2a2a; border:1px solid #3a3a3a; color:#fff; border-radius:4px;">
        <input type="text" id="api_source_token" placeholder="Токен" style="flex:1; min-width:150px; padding:8px; background:#2a2a2a; border:1px solid #3a3a3a; color:#fff; border-radius:4px;">
        <input type="number" id="api_source_interval" placeholder="Мин." value="60" title="Интервал опроса в минутах" style="width:80px; padding:8px; background:#2a2a2a; border:1px solid #3a3a3a; color:#fff; border-radius:4px;">
        <button type="button" class="save-btn" onclick="saveApiSource()" style="background:#3498db;">Добавить</button>
    </form>

    <table class="simple-table">
        <thead><tr><th>Название</th><th>Ссылка</th><th>Интервал</th><th>Действия</th></tr></thead>
        <tbody id="api-sources-tbody"></tbody>
    </table>
</div>
            </div>

            <div id="office-section" style="display: none;"><h2>Офис</h2><p>Информация о компании</p></div>
            <div id="orders-section" style="display: none;"><h2>Заказы</h2><p>Список заказов</p></div>
            <div id="contractors-section" style="display: none;"><h2>Контрагенты</h2><p>Контрагенты</p></div>
            <div id="charts-section" style="display: none;"><h2>Графики</h2><p>Графики</p></div>
            <div id="cash-section" style="display: none;"><h2>Касса</h2><p>Касса</p></div>
            
           <div id="profile-section" style="display: none;">
                <h2>Мой профиль</h2>
                <div class="card" style="background:#1e1e1e; padding:20px; border-radius:8px; max-width: 400px; margin-top: 16px; border: 1px solid #3a3a3a;">
                    <h3 style="margin-bottom: 15px; color: #fff;">Изменение пароля</h3>
                    <form id="change-password-form" style="display: flex; flex-direction: column; gap: 15px;">
                        <input type="text" autocomplete="username" value="текущий_пользователь" style="display:none;">
                        <div class="form-group" style="position: relative;">
                            <label style="color: #aaa; font-size: 12px; margin-bottom: 4px; display: block;">Старый пароль</label>
                            <div style="position: relative;">
                                <input type="password" id="old-password" autocomplete="current-password" required style="width: 100%; padding: 10px; padding-right: 40px; background: #2a2a2a; border: 1px solid #3a3a3a; color: #fff; border-radius: 4px; box-sizing: border-box;">
                                <span class="toggle-password" onclick="togglePasswordVisibility('old-password', this)" style="position: absolute; right: 10px; top: 50%; transform: translateY(-50%); cursor: pointer; color: #888; font-size: 16px;">👁️</span>
                            </div>
                        </div>

                        <div class="form-group" style="position: relative;">
                            <label style="color: #aaa; font-size: 12px; margin-bottom: 4px; display: block;">Новый пароль</label>
                            <div style="position: relative;">
                                <input type="password" id="new-password" autocomplete="new-password"  required minlength="4" style="width: 100%; padding: 10px; padding-right: 40px; background: #2a2a2a; border: 1px solid #3a3a3a; color: #fff; border-radius: 4px; box-sizing: border-box;">
                                <span class="toggle-password" onclick="togglePasswordVisibility('new-password', this)" style="position: absolute; right: 10px; top: 50%; transform: translateY(-50%); cursor: pointer; color: #888; font-size: 16px;">👁️</span>
                            </div>
                        </div>

                        <div class="form-group" style="position: relative;">
                            <label style="color: #aaa; font-size: 12px; margin-bottom: 4px; display: block;">Повторите новый пароль</label>
                            <div style="position: relative;">
                                <input type="password" id="confirm-password" autocomplete="new-password" required minlength="4" style="width: 100%; padding: 10px; padding-right: 40px; background: #2a2a2a; border: 1px solid #3a3a3a; color: #fff; border-radius: 4px; box-sizing: border-box;">
                                <span class="toggle-password" onclick="togglePasswordVisibility('confirm-password', this)" style="position: absolute; right: 10px; top: 50%; transform: translateY(-50%); cursor: pointer; color: #888; font-size: 16px;">👁️</span>
                            </div>
                        </div>

                        <button type="submit" class="save-btn" style="margin-top: 10px; padding: 10px; background: #4a90e2; border-color: #4a90e2; font-size: 14px;">Сохранить пароль</button>
                    </form>
                    <div class="card" style="background:#1e1e1e; padding:20px; border-radius:8px; max-width: 400px; margin-top: 16px; border: 1px solid #3a3a3a;">
                    <h3 style="margin-bottom: 15px; color: #fff;">Управление сеансами</h3>
                    <p style="color: #aaa; font-size: 13px; margin-bottom: 15px;">Вы можете принудительно завершить все активные сеансы на других устройствах (телефонах, других компьютерах). Ваша текущая сессия в этом браузере останется активной.</p>
                    <button onclick="logoutOtherDevices()" class="save-btn" style="background: #e74c3c; border-color: #e74c3c; width: 100%; padding: 10px;">🚪 Выйти на всех других устройствах</button>
                </div>
                </div>
            </div>



          <div id="excel-section" style="display: none;">
    <h2>Подключение файлов и таблиц</h2>
    
    <div style="display:flex; gap:20px; flex-wrap:wrap; margin-top:20px;">
        
        <div class="card" style="background:#1e1e1e; padding:16px; border-radius:8px; flex:1; min-width: 300px; border-left: 4px solid #3498db; display: flex; flex-direction: column;">
            <h3 style="color:#3498db; margin-bottom: 15px; margin-top: 0;">📄 1. Загрузите файл Excel</h3>
            <div style="flex-grow: 1;">
                <label style="color:#aaa; font-size:12px;">Выберите отслеживаемый чат</label>
                <select id="excel-chat-select" style="width:100%; padding:8px; background:#2a2a2a; border:1px solid #3a3a3a; color:#fff; border-radius:4px; margin-bottom:10px;"></select>
                
                <label style="color:#aaa; font-size:12px;">Файл (.xlsx, .xls)</label>
                <input type="file" id="excel-file" accept=".xlsx, .xls" style="width:100%; margin-bottom:15px; padding:7px; background:#2a2a2a; border:1px solid #3a3a3a; color:#fff; border-radius:4px;">
            </div>
            <button type="button" style="width:100%; background:#3498db; border:none; padding:12px; border-radius:6px; color:#fff; font-weight:bold; cursor:pointer; font-size: 14px;" onclick="uploadExcelPreview()">Загрузить и настроить</button>
        </div>

        <div class="card" style="background:#1e1e1e; padding:16px; border-radius:8px; flex:1; min-width: 300px; border-left: 4px solid #2ecc71; display: flex; flex-direction: column;">
            <h3 style="color:#2ecc71; margin-bottom: 15px; margin-top: 0;">🔗 Подключение Google Таблицы</h3>
            <div style="flex-grow: 1;">
                <input type="url" id="gs-url" placeholder="Ссылка на Google Таблицу" required style="width:100%; margin-bottom:10px; padding:8px; background:#2a2a2a; border:1px solid #3a3a3a; color:#fff; border-radius:4px;">
                
                <label style="color:#aaa; font-size:12px;">Привязать к поставщику (каналу):</label>
                <select id="gs-chat-select" required style="width:100%; margin-bottom:10px; padding:8px; background:#2a2a2a; border:1px solid #3a3a3a; color:#fff; border-radius:4px;"></select>
                
                <label style="color:#aaa; font-size:12px;">Парсить каждые (мин)</label>
                <input type="number" id="gs-interval" value="60" required style="width:100%; margin-bottom:15px; padding:8px; background:#2a2a2a; border:1px solid #3a3a3a; color:#fff; border-radius:4px;">
            </div>
            <button type="button" style="width:100%; background:#2ecc71; border:none; padding:12px; border-radius:6px; color:#fff; font-weight:bold; cursor:pointer; font-size: 14px;" onclick="previewGoogleSheet()">Настроить</button>
        </div>

        <div class="card" style="background:#1e1e1e; padding:16px; border-radius:8px; flex:1; min-width: 300px; border-left: 4px solid #e74c3c; display: flex; flex-direction: column;">
            <h3 style="color:#e74c3c; margin-bottom: 15px; margin-top: 0;">📄 Подключение PDF прайсов</h3>
            <div style="flex-grow: 1;">
                <label style="color:#aaa; font-size:12px;">Выберите отслеживаемый чат</label>
                <select id="pdf-chat-select" style="width:100%; padding:8px; background:#2a2a2a; border:1px solid #3a3a3a; color:#fff; border-radius:4px; margin-bottom:10px;"></select>
                
                <label style="color:#aaa; font-size:12px;">Файл (.pdf)</label>
                <input type="file" id="pdf-file" accept=".pdf" style="width:100%; margin-bottom:15px; padding:7px; background:#2a2a2a; border:1px solid #3a3a3a; color:#fff; border-radius:4px;">
            </div>
            <button type="button" style="width:100%; background:#e74c3c; border:none; padding:12px; border-radius:6px; color:#fff; font-weight:bold; cursor:pointer; font-size: 14px;" onclick="uploadPdfPreview()">Загрузить и настроить</button>
        </div>
        
    </div> <div class="card" id="gs-preview-card" style="display:none; background:#1e1e1e; padding:16px; border-radius:8px; margin-top:20px; border-left:4px solid #2ecc71;">
    </div>

    <div class="card" id="pdf-preview-card" style="display:none; background:#1e1e1e; padding:16px; border-radius:8px; margin-top:20px; border-left:4px solid #e74c3c;">
    </div>
    
    <div class="card" style="background:#1e1e1e; padding:16px; border-radius:8px; margin-top:20px;">
        <h3>Активные настройки</h3>
        <div style="overflow-x: auto;">
            <table class="simple-table" style="margin-top:10px; width: 100%;">
                <thead>
                    <tr>
                        <th>Тип источника</th>
                        <th>Чат поставщика</th>
                        <th>Папка</th>
                        <th>Колонки</th>
                        <th>Файл / Ссылка</th>
                        <th>Действия</th>
                    </tr>
                </thead>
                <tbody id="excel-configs-tbody"></tbody>
            </table>
        </div>
    </div>

    <div class="card" id="excel-preview-card" style="display:none; background:#1e1e1e; padding:16px; border-radius:8px; margin-top:20px; border-left:4px solid #f39c12;">
        <h3 id="excel-status" style="color:#f39c12;">Шаг 1: Кликните на ячейку с ПЕРВЫМ НАЗВАНИЕМ товара</h3>
        <div id="excel-table-container" style="margin-top:15px; overflow-x:auto;"></div>
    </div>

</div>

            <div id="api-section" style="display: none;">
                <h2>Управление API (Для выдачи каталога)</h2>
                <div style="display:flex; gap:20px; align-items:flex-start; margin-top:16px; flex-wrap:wrap;">
                    
                    <div class="card" style="background:#1e1e1e; padding:16px; border-radius:8px; flex:1; min-width:300px;">
                        <h3>👥 Мои клиенты</h3>
                        <div style="display:flex; gap:10px; margin: 15px 0;">
                            <input type="text" id="new-api-client-name" placeholder="Имя клиента (например, Вася Опт)" style="flex:1; padding:8px; background:#2a2a2a; border:1px solid #3a3a3a; color:#fff; border-radius:4px;">
                            <button class="save-btn" onclick="createApiClient()">Создать</button>
                        </div>
                        <table class="simple-table">
                            <thead><tr><th>Имя</th><th>Токен</th><th>Режим работы</th><th>Действия</th></tr></thead>
                            <tbody id="api-clients-tbody"></tbody>
                        </table>
                    </div>
                    <div class="card" style="background:#1e1e1e; padding:16px; border-radius:8px; flex:1; min-width:350px; display:none;" id="api-markups-container">
                       <div class="card" style="background:#1e1e1e; padding:16px; border-radius:8px; flex:1; min-width:350px; display:none;" id="api-filters-container">
                        <h3 id="filters-client-title" style="color:#2ecc71; margin-bottom:10px;">Доступ: Дерево товаров</h3>
                        <p style="color:#aaa; font-size:12px; margin-bottom:15px;">Выберите конкретные товары и поставщиков. Если поставщики не выбраны, товар выдаваться не будет.</p>
                        
                        <div id="filter-access-tree" style="max-height:450px; overflow-y:auto; background:#2a2a2a; padding:10px; border-radius:4px; margin-bottom:15px; border:1px solid #3a3a3a;">
                            </div>

                        <button class="save-btn" style="background:#2ecc71; width: 100%; height: 35px;" onclick="saveClientFilters()">💾 Сохранить доступы</button>
                    </div>
                        <h3 id="markup-client-title" style="color:#4a90e2; margin-bottom:15px;">Настройка наценок</h3>
                        <form id="add-markup-form" style="display:flex; flex-wrap:wrap; gap:10px; margin-bottom: 20px; align-items:flex-end;">
                            <input type="hidden" id="markup-client-id">
                            <div class="form-group">
                                <label>Категория (Папка)</label>
                                <select id="markup-folder" style="padding:8px; background:#2a2a2a; border:1px solid #3a3a3a; color:#fff; border-radius:4px;">
                                    <option value="0">Базовая (Для остальных)</option>
                                    </select>
                            </div>
                            <div class="form-group">
                                <label>Тип</label>
                                <select id="markup-type" style="padding:8px; background:#2a2a2a; border:1px solid #3a3a3a; color:#fff; border-radius:4px;">
                                    <option value="percent">Процент (%)</option>
                                    <option value="fixed">Фикс (руб)</option>
                                </select>
                            </div>
                            <div class="form-group">
                                <label>Значение</label>
                                <input type="number" step="0.01" id="markup-value" value="0" style="width:70px; padding:8px; background:#2a2a2a; border:1px solid #3a3a3a; color:#fff; border-radius:4px;">
                            </div>
                            <div class="form-group">
                                <label>Округление до</label>
                                <select id="markup-rounding" style="padding:8px; background:#2a2a2a; border:1px solid #3a3a3a; color:#fff; border-radius:4px;">
                                    <option value="1">1 руб</option>
                                    <option value="10">10 руб</option>
                                    <option value="50">50 руб</option>
                                    <option value="100" selected>100 руб</option>
                                    <option value="500">500 руб</option>
                                    <option value="1000">1000 руб</option>
                                </select>
                            </div>
                            <button type="button" class="save-btn" onclick="saveMarkup()">💾 Сохранить</button>
                        </form>
                        <table class="simple-table">
                            <thead><tr><th>Папка</th><th>Наценка</th><th>Округление</th><th></th></tr></thead>
                            <tbody id="api-markups-tbody"></tbody>
                        </table>
                    </div>
                </div>
              
                <div class="card" style="background:#1e1e1e; padding:16px; border-radius:8px; margin-top:20px; border-left: 4px solid #4a90e2;">
                    <h3>🔗 Как использовать API?</h3>
                    <p style="color:#aaa; font-size:14px; margin-top:10px;">
                        Чтобы получить готовый каталог с рассчитанными ценами, отправьте GET-запрос: <br><br>
                        <code style="background:#2a2a2a; padding:8px 12px; border-radius:4px; color:#4a90e2; font-size:15px; display:inline-block;">https://engine.astoredirect.ru/api/v1/catalog?token=ТОКЕН_КЛИЕНТА</code>
                    </p>
                </div>
            </div>


            <div id="telegram-section" class="section" style="display: none;">
    <h2 style="margin-top: 0; padding-left: 15px;">✈️ Управление Telegram</h2>
    
    <div style="display:flex; gap:20px; align-items:flex-start; margin-top:16px; flex-wrap:wrap; padding-left: 15px;">
        
        <div class="card" style="background:#1e1e1e; padding:16px; border-radius:8px; margin-top: 20px; width: 100%;">
            <h3 style="color:#f39c12; margin-bottom:10px;">📢 Автопубликации прайс-листов</h3>
            
            <div style="display:flex; gap:10px; margin-bottom: 15px;">
                <button class="save-btn" onclick="openPublishEditor()">+ Создать публикацию</button>
            </div>

            <table class="simple-table">
                <thead>
                    <tr><th>Название</th><th>Чат</th><th>Интервал</th><th>Статус</th><th>Действия</th></tr>
                </thead>
                <tbody id="publications-tbody">
                </tbody>
            </table>

            <div id="publish-editor-container" style="display:none; margin-top: 20px; border-top: 1px solid #3a3a3a; padding-top: 20px;">
                <div style="display:flex; gap:20px; flex-wrap:wrap;">
                    
                    <div style="flex:1; min-width:300px;">
                        <h4 style="color:#4a90e2; margin-bottom:15px;">Настройки</h4>
                        <input type="hidden" id="pub_id">
                        <div class="form-group">
                            <label>Название (для себя)</label>
                            <input type="text" id="pub_name" placeholder="Например: Основной канал" style="width:100%; padding:8px; background:#2a2a2a; border:1px solid #3a3a3a; color:#fff; border-radius:4px;">
                        </div>
                        <div class="form-group" style="margin-top:10px;">
                            <label><input type="checkbox" id="pub_enabled"> Включить публикацию</label>
                        </div>
                        <div class="form-group" style="margin-top:10px;">
                            <label>Интервал (минут)</label>
                            <input type="number" id="pub_interval" value="60" style="width:100%; padding:8px; background:#2a2a2a; border:1px solid #3a3a3a; color:#fff; border-radius:4px;">
                        </div>
                        <div class="form-group" style="margin-top:10px;">
                            <label>ID чата (куда публиковать)</label>
                            <input type="text" id="pub_chat_id" style="width:100%; padding:8px; background:#2a2a2a; border:1px solid #3a3a3a; color:#fff; border-radius:4px;">
                        </div>
                        <div class="form-group" style="margin-top:10px;">
                            <label>ID сообщения (если нужно обновлять одно)</label>
                            <input type="text" id="pub_message_id" style="width:100%; padding:8px; background:#2a2a2a; border:1px solid #3a3a3a; color:#fff; border-radius:4px;">
                        </div>
                        <div class="form-group" style="margin-top:10px;">
                            <label>Юзербот для отправки</label>
                            <select id="pub_userbot_id" style="width:100%; padding:8px; background:#2a2a2a; border:1px solid #3a3a3a; color:#fff; border-radius:4px;"></select>
                        </div>
                        <div class="form-group" style="margin-top:10px;">
                            <label>Шаблон сообщения</label>
                            <textarea id="pub_template" style="width:100%; padding:8px; background:#2a2a2a; border:1px solid #3a3a3a; color:#fff; border-radius:4px; height: 100px;"></textarea>
                        </div>
                    </div>

                    <div style="flex:1; min-width:300px;">
                        <h4 style="color:#2ecc71; margin-bottom:10px;">Содержимое прайса</h4>
                        <p style="color:#aaa; font-size:12px; margin-bottom:15px;">Выберите категории и товары, которые войдут в эту публикацию.</p>
                        
                        <div id="pub-filter-tree" style="max-height:500px; overflow-y:auto; background:#2a2a2a; padding:10px; border-radius:4px; border:1px solid #3a3a3a;">
                        </div>
                    </div>
                    
                    <div style="flex:1; min-width:350px;">
                        <h4 style="color:#f39c12; margin-bottom:10px;">Настройки наценок</h4>
                        
                        <div id="pub-markups-warning" style="color:#aaa; font-size:13px; margin-bottom:15px; display:none; background: #2a2a2a; padding: 15px; border-radius: 6px; border: 1px dashed #f39c12;">
                            ⚠️ Сначала сохраните публикацию (кнопка в самом низу), чтобы появилась возможность настраивать наценки.
                        </div>
                        
                        <div id="pub-markups-content">
                            <form id="pub-markup-form" style="display:flex; flex-wrap:wrap; gap:10px; margin-bottom: 20px; align-items:flex-end;">
                                <div class="form-group">
                                    <label>Категория (Папка)</label>
                                    <select id="pub-markup-folder" style="padding:8px; background:#2a2a2a; border:1px solid #3a3a3a; color:#fff; border-radius:4px; max-width: 150px;">
                                        <option value="0">Базовая</option>
                                    </select>
                                </div>
                                <div class="form-group">
                                    <label>Тип</label>
                                    <select id="pub-markup-type" style="padding:8px; background:#2a2a2a; border:1px solid #3a3a3a; color:#fff; border-radius:4px;">
                                        <option value="percent">Процент (%)</option>
                                        <option value="fixed">Фикс (руб)</option>
                                    </select>
                                </div>
                                <div class="form-group">
                                    <label>Значение</label>
                                    <input type="number" step="0.01" id="pub-markup-value" value="0" style="width:70px; padding:8px; background:#2a2a2a; border:1px solid #3a3a3a; color:#fff; border-radius:4px;">
                                </div>
                                <div class="form-group">
                                    <label>Округление</label>
                                    <select id="pub-markup-rounding" style="padding:8px; background:#2a2a2a; border:1px solid #3a3a3a; color:#fff; border-radius:4px;">
                                        <option value="1">1 руб</option>
                                        <option value="10">10 руб</option>
                                        <option value="50">50 руб</option>
                                        <option value="100" selected>100 руб</option>
                                        <option value="500">500 руб</option>
                                        <option value="1000">1000 руб</option>
                                    </select>
                                </div>
                                <button type="button" class="save-btn" onclick="savePubMarkup()">➕</button>
                            </form>
                            <table class="simple-table" style="font-size: 13px;">
                                <thead><tr><th>Папка</th><th>Наценка</th><th>Округление</th><th></th></tr></thead>
                                <tbody id="pub-markups-tbody"></tbody>
                            </table>
                        </div>
                    </div>
                    
                </div>
                <button class="save-btn" style="background:#2ecc71; width: 100%; margin-top:20px; height: 40px; font-size:16px;" onclick="savePublication()">💾 Сохранить публикацию</button>
            </div>
        </div>
    </div>
</div>
            <div id="reports-section" style="display: none;">  
                <h2>📊 Сводный отчёт: Подтверждённые товары</h2>
                <p style="color: #aaa; margin-bottom: 20px;">Здесь отображаются все найденные товары, к которым вы привязали сообщения и подтвердили цену.</p>
                
                <div class="filters-panel" style="display: flex; gap: 15px; margin-bottom: 20px; flex-wrap: wrap;">
                    <div class="form-group" style="flex: 1; min-width: 200px;">
                        <label style="color: #aaa; font-size: 12px; margin-bottom: 4px; display: block;">Фильтр по товару</label>
                        <select id="report-filter-product" style="width: 100%; padding: 8px; background: #1e1e1e; border: 1px solid #3a3a3a; color: #fff; border-radius: 4px;" onchange="renderReportsTable()">
                            <option value="all">Все товары</option>
                        </select>
                    </div>
                    <div class="form-group" style="flex: 1; min-width: 200px;">
                        <label style="color: #aaa; font-size: 12px; margin-bottom: 4px; display: block;">Фильтр по продавцу</label>
                        <select id="report-filter-seller" style="width: 100%; padding: 8px; background: #1e1e1e; border: 1px solid #3a3a3a; color: #fff; border-radius: 4px;" onchange="renderReportsTable()">
                            <option value="all">Все продавцы</option>
                        </select>
                    </div>
                    <div class="form-group" style="flex: 1; min-width: 200px;">
                        <label style="color: #aaa; font-size: 12px; margin-bottom: 4px; display: block;">Сортировка</label>
                        <select id="report-sort" style="width: 100%; padding: 8px; background: #1e1e1e; border: 1px solid #3a3a3a; color: #fff; border-radius: 4px;" onchange="renderReportsTable()">
                            <option value="date_desc">Сначала новые (по дате)</option>
                            <option value="price_asc">Сначала дешевле (по цене)</option>
                            <option value="price_desc">Сначала дороже (по цене)</option>
                            <option value="seller_asc">По продавцу (А-Я)</option>
                        </select>
                    </div>
                </div>
                
                <table class="messages-table">
                    <thead>
                        <tr>
                            <th>ID</th> 
                            <th>Товар</th>
                            <th>Цена</th>
                            <th>Продавец (От кого)</th>
                            <th>Источник (Откуда)</th>
                            <th>Сообщение</th>
                            <th>Дата</th>
                        </tr>
                    </thead>
                    <tbody id="all-confirmed-reports-tbody"></tbody>
                </table>

            </div>



            {% if session.role == 'admin' %}
            <div id="admin-section" style="display: none;">
                <h2>Администрирование</h2>
                <div class="admin-grid" style="display: flex; gap: 20px; margin-top: 16px;">
                    <div class="card" style="background: #1e1e1e; padding: 16px; border-radius: 8px; flex: 1;">
                        <h4>Создать пользователя</h4>
                        <form id="add-user-form" style="margin-top: 10px;">
                            <input type="text" id="admin-new-login" placeholder="Логин" autocomplete="off" style="width:100%; margin-bottom:8px;">
                            <input type="password" id="admin-new-password" placeholder="Пароль" autocomplete="new-password" style="width:100%; margin-bottom:8px;">
                            <button type="submit" class="save-btn">Создать</button>
                        </form>
                    </div>
                    <div class="card" style="background: #1e1e1e; padding: 16px; border-radius: 8px; flex: 1;">
                        <h4>Список пользователей</h4>
                        <ul id="user-list" style="margin-top: 10px; list-style: inside;"></ul>
                    </div>
                </div>
                <div class="card" style="background: #1e1e1e; padding: 16px; border-radius: 8px; margin-top: 20px; border-left: 4px solid #e74c3c;">
                    <h4 style="color: #e74c3c; margin-bottom: 5px;">⚠️ Очистка базы сообщений</h4>
                    <p style="color: #aaa; font-size: 13px; margin-bottom: 15px;">
                        Прямое удаление сообщений из таблицы "messages". 
                        <strong>Внимание:</strong> при удалении сообщений сбросятся подтвержденные цены у товаров, привязанных к этим сообщениям!
                    </p>
                    <div style="display: flex; gap: 10px; align-items: flex-end; flex-wrap: wrap;">
                        <div class="form-group" style="min-width: 250px;">
                            <label style="color: #aaa; font-size: 12px; margin-bottom: 4px; display: block;">За какой период удалить?</label>
                            <select id="clear-messages-period" style="width: 100%; padding: 8px; background: #2a2a2a; border: 1px solid #3a3a3a; color: #fff; border-radius: 4px;">
                                <option value="all">За всё время (Очистить всё полностью)</option>
                                <option value="1">Оставить за сегодня (Удалить старше 1 дня)</option>
                                <option value="3">Удалить старше 3 дней</option>
                                <option value="7">Удалить старше 7 дней</option>
                                <option value="30">Удалить старше 30 дней</option>
                            </select>
                        </div>
                        <button onclick="adminClearMessages()" class="save-btn" style="background: #e74c3c; border: none; height: 35px; padding: 0 20px;">🗑️ Выполнить очистку</button>
                    </div>
                </div>
            </div>
            {% endif %}
        </div>
    </div>
    </div>
    <script>
    let currentPage = 1;
    let currentLimit = 20; // Дефолт 20
    const applyBtn = document.getElementById('apply-filters');
    const messagesTbody = document.getElementById('messages-tbody');
    const noMessagesDiv = document.getElementById('no-messages');
    const tableHeader = document.querySelector('.messages-table thead');
        let currentView = 'inbox';
        let searchWords = [];
        let excludeWords = [];
        let pollInterval = null;
         // Добавьте эту строку
         
let expandedFolders = new Set(JSON.parse(localStorage.getItem('expandedFolders') || '[]'));

        let currentProductFolderView = 'all'; // Текущий фильтр папки для товаров
let currentBindingProductId = null;   // ID товара, который сейчас привязываем

async function loadProductsFolderTree() {
    const folders = await cachedApiFetch('/api/folders/tree');
    const container = document.getElementById('products-folder-tree');
    if (!container) return;

    // Кнопка "Все товары"
    container.innerHTML = `
        <li data-id="all" class="${currentProductFolderView === 'all' ? 'active' : ''}" style="list-style: none; margin-bottom: 8px;">
            <div onclick="filterProductsByFolder('all', 'Все товары')" style="cursor:pointer; padding: 5px; border-radius: 4px;">
                <span class="folder-name" style="font-weight: bold;">📦 Все товары</span>
            </div>
        </li>
    `;

    function buildItem(folder, parent) {
        const li = document.createElement('li');
        li.style.listStyle = 'none';
        li.dataset.id = folder.id;
        if (currentProductFolderView == folder.id) li.classList.add('active');

        const hasChildren = folder.children && folder.children.length > 0;

        const div = document.createElement('div');
        div.style.display = 'flex';
        div.style.alignItems = 'center';
        div.style.padding = '3px 0';

        // Создаем стрелочку-переключатель
        const toggle = document.createElement('span');
        toggle.className = 'folder-toggle';
        toggle.innerHTML = hasChildren ? '▶' : ''; 
        toggle.onclick = (e) => {
            e.stopPropagation();
            li.classList.toggle('expanded'); // Переключаем видимость детей
        };

        const nameSpan = document.createElement('span');
        nameSpan.className = 'folder-name';
        nameSpan.style.cursor = 'pointer';
        nameSpan.style.padding = '4px 8px';
        nameSpan.style.borderRadius = '4px';
        nameSpan.style.flex = '1';
        nameSpan.innerHTML = `📂 ${folder.name}`;
        nameSpan.onclick = () => filterProductsByFolder(folder.id, folder.name);

        div.appendChild(toggle);
        div.appendChild(nameSpan);
        li.appendChild(div);

        if (hasChildren) {
            const ul = document.createElement('ul');
            folder.children.forEach(child => buildItem(child, ul));
            li.appendChild(ul);
            
            // Если внутри этой папки находится текущий выбранный товар, раскрываем её сразу
            if (isChildSelected(folder)) {
                li.classList.add('expanded');
            }
        }
        parent.appendChild(li);
    }

    // Вспомогательная функция, чтобы дерево было раскрыто до активной папки при загрузке
    function isChildSelected(folder) {
        if (!folder.children) return false;
        return folder.children.some(c => c.id == currentProductFolderView || isChildSelected(c));
    }

    folders.forEach(f => {
        if (f.name !== 'По умолчанию') buildItem(f, container);
    });
}

function filterProductsByFolder(folderId, folderName) {
    currentProductFolderView = folderId;
    document.getElementById('selected-product-folder-title').innerText = `📍 Папка: ${folderName}`;
    
    // Сбрасываем активные классы в дереве
    document.querySelectorAll('#products-folder-tree li').forEach(li => li.classList.remove('active'));
    const activeLi = document.querySelector(`#products-folder-tree li[data-id="${folderId}"]`);
    if (activeLi) activeLi.classList.add('active');
    
    loadProducts(); // Перерисовываем таблицу с учетом фильтра
}


        async function apiFetch(url, options = {}) {
            // --- ИСПРАВЛЕНИЕ: Автоматически сбрасываем кэш при любом изменении (POST, PUT, DELETE) ---
            if (options.method && options.method.toUpperCase() !== 'GET') {
                window.apiCache = {}; 
            }
            // ---------------------------------------------------------------------------------------

            const resp = await fetch(url, options);
            if (!resp.ok) {
                const text = await resp.text();
                let errorText = 'Неизвестная ошибка сервера';
                try {
                    errorText = JSON.parse(text).error || text;
                } catch (e) {
                    errorText = text;
                }
                alert('Внимание: ' + errorText);
                throw new Error(errorText);
            }
            return resp.json();
        }

        // --- ДОБАВЛЯЕМ УМНЫЙ КЭШ ПРЯМО СЮДА ---
        window.apiCache = {};
        
        async function cachedApiFetch(url, forceRefresh = false) {
            if (!forceRefresh && window.apiCache[url]) {
               
                return window.apiCache[url];
            }
           
            const data = await apiFetch(url);
            window.apiCache[url] = data; 
            return data;
        }
        // --------------------------------------

        function updateDateTime() {
            const now = new Date();
            document.getElementById('current-datetime').innerText = now.toLocaleString('ru-RU');
        }
        setInterval(updateDateTime, 1000);
        updateDateTime();

        const menuItems = document.querySelectorAll('.menu-items li');
        const sections = {
            parser: document.getElementById('parser-section'),
            interaction_bots: document.getElementById('interaction_bots-section'),
            products: document.getElementById('products-section'),
            userbots: document.getElementById('userbots-section'),
            api: document.getElementById('api-section'),
            telegram: document.getElementById('telegram-section'),
            excel: document.getElementById('excel-section'),
            office: document.getElementById('office-section'),
            orders: document.getElementById('orders-section'),
            contractors: document.getElementById('contractors-section'),
            charts: document.getElementById('charts-section'),
            cash: document.getElementById('cash-section'),
            reports: document.getElementById('reports-section'),
            profile: document.getElementById('profile-section'),
            admin: document.getElementById('admin-section')
        };
        const sidebar = document.getElementById('parser-sidebar');


        applyBtn.addEventListener('click', () => {
    const sInput = document.getElementById('filter-search-input');
    if (sInput && sInput.value.trim()) {
        const val = sInput.value.trim();
        if (!searchWords.includes(val)) {
            searchWords.push(val);
            updateSearchTags();
        }
        sInput.value = '';
    }
    const eInput = document.getElementById('filter-exclude-input');
    if (eInput && eInput.value.trim()) {
        const val = eInput.value.trim();
        if (!excludeWords.includes(val)) {
            excludeWords.push(val);
            updateExcludeTags();
        }
        eInput.value = '';
    }
    currentView = 'inbox';
    localStorage.setItem('activeFolder', currentView);
    syncActiveFolder();
    loadMessages();
});

       function showSection(sectionId) {

       localStorage.setItem('activeSection', sectionId);
    // 1. Скрываем вообще все стандартные разделы
    Object.values(sections).forEach(s => { if(s) s.style.display = 'none'; });
    
    // 2. Скрываем раздел каталога папок (он у нас отдельный)
    const fcSection = document.getElementById('folder-catalog-section');
    if (fcSection) fcSection.style.display = 'none';

    // 3. ЛОГИКА ОТОБРАЖЕНИЯ
  if (sectionId === 'parser') {
      // Если мы в парсере, проверяем: показать общую ленту или папку
      if (currentView === 'inbox') {
          if (sections['parser']) sections['parser'].style.display = 'block';
          if (typeof initInboxFilters === 'function') setTimeout(initInboxFilters, 50);
      } else {
          if (fcSection) fcSection.style.display = 'block';
      }
  } else if (sectionId === 'api') {
      sections['api'].style.display = 'block';
      loadApiClients();
      // Отсюда мы удалили loadPublications(), так как она теперь в Telegram
  } else if (sectionId === 'telegram') {                             
      if (sections['telegram']) sections['telegram'].style.display = 'block'; 
      if (typeof loadPublications === 'function') loadPublications();
  } else if (sectionId === 'excel') {
      sections['excel'].style.display = 'block'; 
      if (typeof loadExcelChats === 'function') {
          loadExcelChats();
          loadExcelConfigs();
      }
  }
    else {
        // ДЛЯ ВСЕХ ОСТАЛЬНЫХ ВКЛАДОК: показываем выбранный раздел
        if (sections[sectionId]) {
            sections[sectionId].style.display = 'block';
        }
    }

    // Боковую панель с папками показываем только в Парсере
    if (sidebar) {
        sidebar.style.display = (sectionId === 'parser') ? 'block' : 'none';
    }

    const botForm = document.getElementById('bot-form-container');
    if (botForm) botForm.style.display = 'none';

    // Подгрузка данных для каждой вкладки
    if (sectionId === 'userbots') {
        loadUserbots();
        loadTrackedChats();
        loadApiSources();
    } else if (sectionId === 'interaction_bots') {
        loadInteractionBots();
    } else if (sectionId === 'products') {
        loadProductsFolderTree(); // Сначала строим дерево
    loadProducts();
    } else if (sectionId === 'reports') {
        if (typeof loadAllConfirmedReports === 'function') loadAllConfirmedReports();
    } else if (sectionId === 'admin') {
        if (typeof loadUsers === 'function') loadUsers();
    }
}

            // Боковую панель с папками показываем только в Парсере
         
        

        menuItems.forEach(item => {
            item.addEventListener('click', () => {
                const section = item.dataset.section;
                if (section) {
                    menuItems.forEach(i => i.classList.remove('active'));
                    item.classList.add('active');
                    showSection(section);
                }
            });
        });

        const toggleSidebar = document.getElementById('toggle-sidebar');
        const folderTreeContainer = document.getElementById('folder-tree-container');
        let sidebarHidden = localStorage.getItem('sidebarHidden') === 'true';

        if (sidebarHidden) {
            folderTreeContainer.style.display = 'none';
            toggleSidebar.innerText = '▶';
        } else {
            folderTreeContainer.style.display = 'block';
            toggleSidebar.innerText = '◀';
        }

        toggleSidebar.addEventListener('click', () => {
            sidebarHidden = !sidebarHidden;
            folderTreeContainer.style.display = sidebarHidden ? 'none' : 'block';
            toggleSidebar.innerText = sidebarHidden ? '▶' : '◀';
            localStorage.setItem('sidebarHidden', sidebarHidden);
        });

        async function checkSession() {
            const data = await apiFetch('/api/check_session');
            loadFolders();
            
            // НОВОЕ: Запускаем загрузку сообщений и их обновление В ЛЮБОМ СЛУЧАЕ, 
            // даже если юзербот выключен (чтобы вы могли читать старые сообщения)
            startPolling();
            
            if (data.authorized) {
                document.getElementById('add-bot-btn').style.display = 'none';
                document.getElementById('bot-form-container').style.display = 'none';
            } else {
                document.getElementById('add-bot-btn').style.display = 'block';
            }
        }

        const addBotBtn = document.getElementById('add-bot-btn');
        addBotBtn.addEventListener('click', () => {
            const form = document.getElementById('bot-form-container');
            form.style.display = form.style.display === 'none' || form.style.display === '' ? 'block' : 'none';
        });
        // Оживляем кнопку "Добавить юзербота" на вкладке Юзерботов
        const addUserbotBtn = document.getElementById('add-userbot-btn');
        if (addUserbotBtn) {
            addUserbotBtn.addEventListener('click', () => {
                const form = document.getElementById('bot-form-container');
                form.style.display = form.style.display === 'none' || form.style.display === '' ? 'block' : 'none';
                form.scrollIntoView({ behavior: 'smooth' }); // Плавно прокручиваем к форме
            });
        }

        const qrBtn = document.getElementById('qr-btn');
        const qrContainer = document.getElementById('qr-container');
        const qrImage = document.getElementById('qr-image');
        const qrStatus = document.getElementById('qr-status');
        let checkInterval = null;

        // --- 1. ВЫНОСИМ ЛОГИКУ В ОТДЕЛЬНУЮ ФУНКЦИЮ ---
// Теперь она умеет принимать sessionId и отправлять его на сервер (для обновления старой сессии)
window.generateQR = async function(apiId, apiHash, sessionId = null) {
    qrStatus.innerText = 'Генерация QR...';
    qrContainer.style.display = 'block';
    qrImage.innerHTML = '';
    
    try {
        const bodyData = { api_id: apiId, api_hash: apiHash };
        
        // Если передан ID старой сессии, обязательно добавляем его в запрос!
        if (sessionId) {
            bodyData.session_id = sessionId;
        }

        const response = await fetch('/api/get_qr', {
            method: 'POST',
            headers: { 'Content-Type': 'application/json' },
            body: JSON.stringify(bodyData)
        });
        const data = await response.json();

        if (data.success) {
            if (data.already_authorized) {
                qrStatus.innerText = '✅ Вы уже авторизованы. Запускаем слушатель...';
                startPolling();
                loadUserbots();
                setTimeout(() => { document.getElementById('bot-form-container').style.display = 'none'; }, 2000);
            } else if (data.qr) {
                qrImage.innerHTML = `<img src="data:image/png;base64,${data.qr}" alt="QR код">`;
                qrStatus.innerText = 'Отсканируйте QR в Telegram. Ожидание входа...';
                if (checkInterval) clearInterval(checkInterval);
                checkInterval = setInterval(async () => {
                    const statusResp = await fetch('/api/check_login?session_id=' + data.session_id);
                    const statusData = await statusResp.json();
                    if (statusData.success) {
                        qrStatus.innerText = '✅ Авторизация успешна!';
                        clearInterval(checkInterval);
                        checkInterval = null;
                        startPolling();
                        loadUserbots();
                        setTimeout(() => { document.getElementById('bot-form-container').style.display = 'none'; }, 2000);
                    }
                }, 2000);
            }
        } else {
            qrStatus.innerText = 'Ошибка: ' + data.error;
        }
    } catch (e) {
        qrStatus.innerText = 'Ошибка запроса.';
    }
};

// --- 2. ОБНОВЛЯЕМ КНОПКУ В ФОРМЕ ---
qrBtn.addEventListener('click', async () => {
    const apiId = document.getElementById('api_id').value.trim();
    const apiHash = document.getElementById('api_hash').value.trim();
    
    // Пытаемся достать ID старой сессии из скрытого поля
    const sessionInput = document.getElementById('auth_session_id');
    const sessionId = sessionInput && sessionInput.value ? sessionInput.value : null;
    
    if (!apiId || !apiHash) {
        alert('Заполните API ID и API Hash');
        return;
    }
    
    // Передаем данные в нашу новую функцию (включая sessionId, если он есть)
    await window.generateQR(apiId, apiHash, sessionId);
});
        
        // --- Логика Сводного отчета ---
        // --- Логика Сводного отчета ---
        let allReportsData = []; // Переменная для хранения всех отчетов в памяти браузера

        async function loadAllConfirmedReports() {
            const tbody = document.getElementById('all-confirmed-reports-tbody');
            if (!tbody) return;
            
            try {
                tbody.innerHTML = '<tr><td colspan="7" style="text-align:center;">Загрузка...</td></tr>';
                // Загружаем данные с сервера один раз
                allReportsData = await apiFetch('/api/reports/confirmed');
                
                // Автоматически собираем все существующие Товары и Продавцов для выпадающих списков
                const productFilter = document.getElementById('report-filter-product');
                const sellerFilter = document.getElementById('report-filter-seller');
                
                if (productFilter && sellerFilter) {
                    const uniqueProducts = [...new Set(allReportsData.map(item => item.product_name))].sort();
                    // Объединяем имя отправителя и чат, чтобы было понятнее, кто это
                    const uniqueSellers = [...new Set(allReportsData.map(item => `${item.sender_name || 'Неизвестно'} (${item.chat_title})`))].sort();
                    
                    productFilter.innerHTML = '<option value="all">Все товары</option>' + 
                        uniqueProducts.map(p => `<option value="${p.replace(/"/g, '&quot;')}">${p}</option>`).join('');
                        
                    sellerFilter.innerHTML = '<option value="all">Все продавцы</option>' + 
                        uniqueSellers.map(s => `<option value="${s.replace(/"/g, '&quot;')}">${s}</option>`).join('');
                }
                
                // Отрисовываем таблицу
                renderReportsTable();
            } catch (e) {
                console.error('Ошибка загрузки отчета', e);
                tbody.innerHTML = '<tr><td colspan="7" style="text-align:center; color:#ff5555;">Ошибка загрузки</td></tr>';
            }
        }

        // Новая функция: фильтрует, сортирует и рисует таблицу мгновенно!
       function renderReportsTable() {
    const tbody = document.getElementById('all-confirmed-reports-tbody');
    if (!tbody) return;

    const productFilterVal = document.getElementById('report-filter-product')?.value || 'all';
    const sellerFilterVal = document.getElementById('report-filter-seller')?.value || 'all';
    const sortVal = document.getElementById('report-sort')?.value || 'date_desc';

    let filteredData = allReportsData.filter(item => {
        const matchProduct = productFilterVal === 'all' || item.product_name === productFilterVal;
        let senderName = item.sender_name || 'Неизвестно';
        if (senderName === 'None') senderName = 'Неизвестно';
        const itemSellerKey = `${senderName} (${item.chat_title})`;
        const matchSeller = sellerFilterVal === 'all' || itemSellerKey === sellerFilterVal;
        return matchProduct && matchSeller;
    });

    filteredData.sort((a, b) => {
        if (sortVal === 'price_asc') return (a.extracted_price || 0) - (b.extracted_price || 0);
        if (sortVal === 'price_desc') return (b.extracted_price || 0) - (a.extracted_price || 0);
        if (sortVal === 'seller_asc') return (`${a.sender_name || ''} ${a.chat_title || ''}`).toLowerCase().localeCompare(`${b.sender_name || ''} ${b.chat_title || ''}`.toLowerCase());
        return new Date(b.date) - new Date(a.date);
    });

    tbody.innerHTML = filteredData.map(item => {
        const isActual = item.is_actual === 1;
        const statusBadge = isActual ? '<span style="background: rgba(39, 174, 96, 0.2); color: #2ecc71; padding: 2px 6px; border-radius: 4px; font-size: 11px;">Актуален</span>' : '<span style="background: rgba(231, 76, 60, 0.2); color: #e74c3c; padding: 2px 6px; border-radius: 4px; font-size: 11px;">Не актуален</span>';
        
        let senderName = item.sender_name || 'Неизвестно';
        if (senderName === 'None') senderName = 'Неизвестно';

        // --- УМНЫЙ РЕНДЕР ИСТОЧНИКОВ ---
        let originalTitle = escapeHtml(item.chat_title || 'Неизвестный чат');
        let chatDisplayHtml = `<span style="color: #888;">${originalTitle}</span>`;
        
        if (item.sheet_url) {
            chatDisplayHtml = `<a href="${item.sheet_url}" target="_blank" style="color: #2ecc71; text-decoration: none; border-bottom: 1px dashed #2ecc71; font-weight: bold;" title="Открыть Google Таблицу">🔗 ${originalTitle} ↗</a>`;
        }
        else if (item.chat_id && item.telegram_message_id && (item.type === 'channel' || item.type === 'group')) {
            let cleanChatId = String(item.chat_id).replace(/^-100|^ -/, '');
            chatDisplayHtml = `<a href="https://t.me/c/${cleanChatId}/${item.telegram_message_id}" target="_blank" style="color: #4a90e2; text-decoration: none; border-bottom: 1px dashed #4a90e2; font-weight: bold;" title="Открыть сообщение в Telegram">🔗 ${originalTitle} ↗</a>`;
        } 
        else if (item.chat_id && String(item.chat_id).startsWith('api_src_')) {
            chatDisplayHtml = `<span style="color: #f39c12; font-weight: bold;" title="Обновлено по API">🌐 ${originalTitle}</span>`;
        
        } else if (item.type === 'private' || !item.type || (item.type !== 'excel' && !item.type.startsWith('excel'))) {
            let botUsername = null;
            let possibleNames = [
                item.custom_name, 
                item.chat_title, 
                item.tc_chat_title, 
                item.sender_name
            ];
            
            // 1. Агрессивно ищем любую строку с собачкой @
            for (let name of possibleNames) {
                if (name && typeof name === 'string' && name.includes('@')) {
                    let match = name.match(/@([a-zA-Z0-9_]+)/);
                    if (match) {
                        botUsername = match[1];
                        break;
                    }
                }
            }

            // 2. Если собачки нигде нет, но имя состоит из одного слова и заканчивается на "bot"
            if (!botUsername) {
                for (let name of possibleNames) {
                    if (name && typeof name === 'string') {
                        let cleanName = name.trim();
                        if (cleanName.toLowerCase().endsWith('bot') && !cleanName.includes(' ')) {
                            botUsername = cleanName;
                            break;
                        }
                    }
                }
            }

            if (botUsername) {
                // Делаем красивую веб-ссылку t.me
                let cleanUsername = botUsername.replace('@', '');
                chatDisplayHtml = `<a href="https://t.me/${cleanUsername}" target="_blank" title="Открыть бота" style="color: #4a90e2; text-decoration: none; border-bottom: 1px dashed #4a90e2; font-weight: bold;">🔗 ${originalTitle} ↗</a>`;
            } else {
                // Если юзернейма нет, открываем по ID через resolve
                let cleanUserId = item.chat_id ? String(item.chat_id).replace(/^-/, '') : '';
                chatDisplayHtml = `<a href="tg://resolve?domain=${cleanUserId}" title="Открыть чат" style="color: #4a90e2; text-decoration: none; border-bottom: 1px dashed #4a90e2; font-weight: bold;">🔗 ${originalTitle} ↗</a>`;
            }
        }        else if (item.type && item.type.startsWith('excel')) {
            chatDisplayHtml = `<span style="color: #2ecc71; font-weight: bold;" title="Локальный файл">🔗 ${originalTitle}</span>`;
        }

        return `
        <tr style="opacity: ${isActual ? '1' : '0.5'};">
            <td style="color: #888; font-weight: bold;">#${item.binding_id}</td>
            <td><strong>${item.product_name}</strong> <br>${statusBadge}</td>
            <td style="color: #27ae60; font-weight: bold; font-size: 16px;">
                ${item.extracted_price ? item.extracted_price.toLocaleString('ru-RU') : '—'}

            </td>
            <td>${senderName === 'Неизвестно' ? '<span style="color:#888;">—</span>' : escapeHtml(senderName)}</td>
            <td>${chatDisplayHtml}</td>
            <td style="white-space: pre-wrap; font-size: 12px; max-width: 400px;">${item.text}</td>
            <td style="font-size: 12px; color: #aaa;">${item.date}</td>
        </tr>
        `;
    }).join('') || '<tr><td colspan="7" class="no-messages" style="text-align:center; padding: 20px;">По заданным фильтрам ничего не найдено</td></tr>';
}

        // Обновляем таблицу при клике на вкладку "Отчёты по товарам"
        document.querySelector('[data-section="reports"]')?.addEventListener('click', loadAllConfirmedReports);

        // Обновляем таблицу при клике на вкладку "Отчёты по товарам"
        document.querySelector('[data-section="reports"]')?.addEventListener('click', loadAllConfirmedReports);


// --- Логика Товаров ---
        let currentMergeSourceId = null;

        // ПОИСК ПО ТОВАРАМ
        document.getElementById('product-search')?.addEventListener('input', (e) => {
            const term = e.target.value.toLowerCase();
            document.querySelectorAll('#products-tbody tr').forEach(tr => {
                const text = tr.innerText.toLowerCase();
                tr.style.display = text.includes(term) ? '' : 'none';
            });
        });

        // ДОБАВЛЕНИЕ ТОВАРА
        document.getElementById('add-product-form')?.addEventListener('submit', async (e) => {
            e.preventDefault();
            const folderIdVal = document.getElementById('prod_folder_id').value;
            await apiFetch('/api/products', {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify({
                    name: document.getElementById('prod_name').value,
                    synonyms: document.getElementById('prod_synonyms').value,
                    folder_id: folderIdVal ? parseInt(folderIdVal) : null
                })
            });
            e.target.reset();
            loadProducts();
            loadFolders();
        });

        // ЗАГРУЗКА И ОТРИСОВКА ТОВАРОВ
        async function loadProducts() {
    const prods = await cachedApiFetch('/api/products');
    const tbody = document.getElementById('products-tbody');
    if (!tbody) return;

    // Фильтруем данные по выбранной папке
    const filtered = currentProductFolderView === 'all' 
        ? prods 
        : prods.filter(p => p.folder_id == currentProductFolderView);

    tbody.innerHTML = filtered.map(p => {
        const safeName = String(p.name || '').replace(/['"]/g, '`');
        const safeSyns = String(p.synonyms || '').replace(/['"]/g, '`');
        
        return `
            <tr style="cursor: pointer;" 
                onclick="openProductBinding(${p.id}, '${safeName}')"
                onmouseover="this.style.background='#3a3a3a'" 
                onmouseout="this.style.background='transparent'">
                
                <td><strong>${p.name}</strong></td>
                <td style="color: #aaa;">${p.synonyms || ''}</td>
                <td style="white-space: nowrap;">
                    <button class="save-btn" onclick="event.stopPropagation(); editProduct(${p.id}, '${safeName}', '${safeSyns}')" style="background:#f39c12; margin-right:4px;">✏️</button>
                    <button class="save-btn" onclick="event.stopPropagation(); openMoveFolderModal(${p.id}, '${safeName}', ${p.folder_id || 'null'})" style="background:#3498db; margin-right:4px;">📁</button>
                    <button class="save-btn" onclick="event.stopPropagation(); deleteProduct(${p.id})" style="background:#e74c3c;">❌</button>
                </td>
            </tr>
        `}).join('');
}

        // РЕДАКТИРОВАНИЕ ТОВАРА
        let currentEditProductId = null;

        function editProduct(id, currentName, currentSynonyms) {
            currentEditProductId = id;
            document.getElementById('edit-prod-name').value = currentName;
            document.getElementById('edit-prod-synonyms').value = currentSynonyms;
            document.getElementById('edit-modal').style.display = 'block';
        }

 async function saveEditProduct() {
            if (!currentEditProductId) return;
            const newName = document.getElementById('edit-prod-name').value.trim();
            const newSynonyms = document.getElementById('edit-prod-synonyms').value.trim();

            if (!newName) {
                alert('Название не может быть пустым!');
                return;
            }

            await apiFetch(`/api/products/${currentEditProductId}`, {
                method: 'PUT',
                headers: {'Content-Type': 'application/json'},
                body: JSON.stringify({name: newName, synonyms: newSynonyms})
            });
            
            document.getElementById('edit-modal').style.display = 'none';
            
            // 1. Обновляем общий список товаров
            loadProducts();
            
            // 2. ИСПРАВЛЕНИЕ: Скачиваем свежие данные и перерисовываем каталог папки!
            await loadFolders(); 
            if (currentView !== 'inbox' && typeof loadFolderCatalog === 'function') {
                loadFolderCatalog(currentView);
            }
            
            // 3. Обновляем заголовок привязки, если он открыт
            if (currentBindingProductId === currentEditProductId) {
                const titleEl = document.getElementById('binding-product-title');
                if (titleEl) titleEl.innerText = `Привязка сообщений: ${newName}`;
            }
        }

        // НАСТРОЙКА ПАПКИ ДЛЯ ТОВАРА
        let currentMoveProductId = null;

        function openMoveFolderModal(id, name, currentFolderId) {
            currentMoveProductId = id;
            
            // ИСПРАВЛЕНИЕ: Вытаскиваем окно из ловушки в корень страницы
            const modal = document.getElementById('move-folder-modal');
            if (modal && modal.parentNode !== document.body) {
                document.body.appendChild(modal); 
            }

            // Заполняем данные
            const nameEl = document.getElementById('move-prod-name');
            if (nameEl) nameEl.innerText = name;

            const folderSelect = document.getElementById('move-folder-select');
            const sourceSelect = document.getElementById('prod_folder_id'); 
            if (sourceSelect && folderSelect) {
                // Копируем список папок
                folderSelect.innerHTML = sourceSelect.innerHTML;
                
                // --- НОВОЕ: Разблокируем текущую папку, так как для этого товара она "своя" ---
                if (currentFolderId) {
                    const currentOpt = Array.from(folderSelect.options).find(opt => opt.value == currentFolderId);
                    if (currentOpt) {
                        currentOpt.disabled = false; // Снимаем блокировку
                        currentOpt.textContent = currentOpt.textContent.replace(' (Занята)', ' (Текущая)');
                    }
                }
                
                folderSelect.value = currentFolderId || ""; 
            }

            // Показываем окно
            if (modal) modal.style.display = 'block';
        }

        async function saveProductFolder() {
            if (!currentMoveProductId) return;
            const newFolderId = document.getElementById('move-folder-select').value;

            await apiFetch(`/api/products/${currentMoveProductId}/link_folder`, {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify({ folder_id: newFolderId ? parseInt(newFolderId) : null })
            });

            document.getElementById('move-folder-modal').style.display = 'none';
            loadProducts(); 
            loadFolders();  
            
            // Если мы находимся в каталоге папки, обновим и его
            if (currentView !== 'inbox' && typeof loadFolderCatalog === 'function') {
                loadFolderCatalog(currentView);
            }
        }

        // УДАЛЕНИЕ ТОВАРА
        async function deleteProduct(id) {
            if(confirm('Удалить товар?')) {
                await apiFetch(`/api/products/${id}`, { method: 'DELETE' });
                if (currentBindingProductId === id) closeProductBinding();
                loadProducts();
                loadFolders(); // Обновляем папки, вдруг товар был привязан
            }
        }

        // ОБЪЕДИНЕНИЕ ТОВАРОВ (Вызывается из каталога)
        async function showMergeModal(id, name) {
            currentMergeSourceId = id;
            document.getElementById('merge-source-name').innerText = name;
            const prods = await cachedApiFetch('/api/products');
            const list = document.getElementById('merge-target-list');
            
            list.innerHTML = prods
                .filter(p => p.id !== id)
                .map(p => `<button onclick="mergeProducts(${id}, ${p.id})" style="padding:10px; background:#1e1e1e; color:#fff; border:1px solid #3a3a3a; cursor:pointer; text-align:left; border-radius:4px;">👉 ${p.name}</button>`)
                .join('');
            
            document.getElementById('merge-modal').style.display = 'block';
        }

        async function mergeProducts(sourceId, targetId) {
            if(confirm('Точно объединить? Товар исчезнет, а его синонимы и сообщения перейдут к выбранному.')) {
                await apiFetch(`/api/products/${sourceId}/merge`, {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({target_id: targetId})
                });
                document.getElementById('merge-modal').style.display = 'none';
                loadProducts();
                if (currentBindingProductId === sourceId) closeProductBinding();
            }
        }
        // --- Логика привязки сообщений (ПОСТРОЧНО) ---

     

        function closeProductBinding() {
            const container = document.getElementById('product-binding-container');
            if (container) container.style.display = 'none';
            currentBindingProductId = null;
        }

        async function openProductBinding(prodId, prodName) {
            currentBindingProductId = prodId;
            const titleEl = document.getElementById('binding-product-title');
            const container = document.getElementById('product-binding-container');
            
            if (titleEl) titleEl.innerText = `Привязка сообщений: ${prodName}`;
            if (container) {
                container.style.display = 'block';
                await refreshProductBinding();
                container.scrollIntoView({ behavior: 'smooth' });
            }
        }

        async function refreshProductBinding() {
            if (!currentBindingProductId) return;
            const prodId = currentBindingProductId;
            const data = await apiFetch(`/api/products/${prodId}/messages`);
            
            // Рендер предложенных СТРОК
            const propTbody = document.getElementById('proposed-messages-tbody');
            if (propTbody) {
                propTbody.innerHTML = data.proposed.map(m => `
                    <tr>
                        <td>👤 ${getSenderDisplayName(m)}</td>
                        <td>${m.custom_name || m.chat_title || 'Личный чат'}</td>
                        <td style="white-space: pre-wrap; font-size: 13px;">${m.match_line || m.text || ''}</td>
                        <td><input type="number" step="0.01" id="price_prop_${m.message_id}_${m.line_index}" value="${m.suggested_price || ''}" style="width:80px; padding:4px; background:#1e1e1e; color:#fff; border:1px solid #3a3a3a;"></td>
                        <td>
                             <button class="save-btn" onclick="confirmMessage(${prodId}, ${m.message_id}, ${m.line_index})">✅ Подтвердить</button>
                        </td>
                    </tr>
                `).join('') || '<tr><td colspan="5" class="no-messages">Нет новых совпадений</td></tr>';
            }

            // Рендер подтвержденных СТРОК
            const confTbody = document.getElementById('confirmed-messages-tbody');
            if (confTbody) {
                confTbody.innerHTML = data.confirmed.map(m => `
                    <tr>
                        <td><input type="checkbox" class="conf-cb" value="${m.pm_id}" data-sender="${getSenderDisplayName(m).replace(/"/g, '&quot;')}" data-chat="${(m.custom_name || m.chat_title || '').replace(/"/g, '&quot;')}" onchange="toggleMergeBtn()"></td>
                        <td>👤 ${getSenderDisplayName(m)}</td>
                        <td>${m.custom_name || m.chat_title || 'Личный чат'}</td>
                        <td style="white-space: pre-wrap; font-size: 13px;">${m.text}</td>
                        <td>
                            <input type="number" step="0.01" id="price_conf_${m.pm_id}" value="${m.extracted_price || ''}" style="width:80px; padding:4px; background:#1e1e1e; color:#fff; border:1px solid #3a3a3a;">
                            <button onclick="updatePrice(${m.pm_id})" style="cursor:pointer; background:none; border:none; font-size:16px;" title="Сохранить цену">💾</button>
                        </td>
                        <td>
                            <button class="save-btn" style="background:#e74c3c;" onclick="detachMessage(${prodId}, ${m.message_id})">Открепить</button>
                        </td>
                    </tr>
                `).join('') || '<tr><td colspan="6" class="no-messages">Нет прикрепленных сообщений</td></tr>';
                
                // Скрываем кнопку после перерисовки
                document.getElementById('btn-merge-conf').style.display = 'none';
                const selectAll = document.getElementById('select-all-conf');
                if(selectAll) selectAll.checked = false;
            }
        }


        async function confirmMessage(prodId, msgId, lineIndex) {
            const priceInput = document.getElementById(`price_prop_${msgId}_${lineIndex}`);
            const price = (priceInput && priceInput.value !== '') ? parseFloat(priceInput.value) : null;
            
            await apiFetch('/api/product_messages/confirm', {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify({ product_id: prodId, message_id: msgId, line_index: lineIndex, price: price })
            });
            refreshProductBinding();
        }

        async function detachMessage(prodId, msgId) {
            await apiFetch('/api/product_messages/detach', {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify({ product_id: prodId, message_id: msgId })
            });
            refreshProductBinding();
        }

        async function updatePrice(pmId) {
            const priceInput = document.getElementById(`price_conf_${pmId}`);
            const price = priceInput ? priceInput.value : null;
            await apiFetch('/api/product_messages/update_price', {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify({ pm_id: pmId, price: price })
            });
            alert('Цена обновлена');
        }

        // --- Логика Взаимодействия с ботами ---
       async function loadInteractionBots() {
            // Загружаем список доступных юзерботов
            const bots = await cachedApiFetch('/api/userbots');
            const select = document.getElementById('ib_userbot_id');
            if(select) {
                // Выводим @username в выпадающем меню
                select.innerHTML = bots.filter(b => b.status === 'active').map(b => `<option value="${b.id}">${b.account_name || 'API ID: ' + b.api_id}</option>`).join('');
            }

            const ibots = await apiFetch('/api/interaction_bots');
            const tbody = document.getElementById('interaction-bots-tbody');
            if(!tbody) return;
            
            tbody.innerHTML = ibots.map(ib => {
                // Ищем юзербота, чтобы показать его имя в таблице
                const ubot = bots.find(b => b.id === ib.userbot_id);
                const ubotName = ubot ? (ubot.account_name || ('API ID: ' + ubot.api_id)) : ib.userbot_id;

                return `
                <tr>
                    <td><strong style="color:#4a90e2;">${ubotName}</strong></td>
                    <td>${ib.bot_username}</td>
                    <td>${JSON.parse(ib.commands).join(', ')}</td>
                    <td>${ib.interval_minutes}</td>
                    <td><button class="save-btn" onclick="deleteInteractionBot(${ib.id})" style="background:#e74c3c;">Удалить</button></td>
                </tr>
            `}).join('');
        }

        document.getElementById('add-interaction-bot-form')?.addEventListener('submit', async (e) => {
            e.preventDefault();
            const rawCommands = document.getElementById('ib_commands').value.split(',').map(c => c.trim()).filter(c => c);
            await apiFetch('/api/interaction_bots', {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify({
                    userbot_id: parseInt(document.getElementById('ib_userbot_id').value),
                    bot_username: document.getElementById('ib_bot_username').value,
                    commands: rawCommands,
                    interval_minutes: parseInt(document.getElementById('ib_interval').value)
                })
            });
            e.target.reset();
            loadInteractionBots();
        });

        async function deleteInteractionBot(id) {
            if(confirm('Удалить бота из автоматизации?')) {
                await apiFetch(`/api/interaction_bots/${id}`, { method: 'DELETE' });
                loadInteractionBots();
            }
        }
      
        const searchTags = document.getElementById('search-tags');
       

        function updateSearchTags() {
        localStorage.setItem('searchWords', JSON.stringify(searchWords)); // Сохраняем массив
    searchTags.innerHTML = '';
           
            searchWords.forEach(word => {
                const tag = document.createElement('span');
                tag.className = 'tag';
                tag.innerHTML = `${word} <span class="remove-tag" data-word="${word}">&times;</span>`;
                searchTags.appendChild(tag);
            });
            document.querySelectorAll('#search-tags .remove-tag').forEach(btn => {
                btn.addEventListener('click', (e) => {
                    const word = e.target.dataset.word;
                    searchWords = searchWords.filter(w => w !== word);
                    updateSearchTags();
                });
            });
        }

        const excludeInput = document.getElementById('filter-exclude-input');
        const excludeTags = document.getElementById('exclude-tags');
        excludeInput.addEventListener('keydown', (e) => {
            if (e.key === 'Enter' || e.key === ',') {
                e.preventDefault();
                const word = excludeInput.value.trim();
                if (word && !excludeWords.includes(word)) {
                    excludeWords.push(word);
                    updateExcludeTags();
                }
                excludeInput.value = '';
            }
        });

        function updateExcludeTags() {
        localStorage.setItem('excludeWords', JSON.stringify(excludeWords)); // Сохраняем массив
    excludeTags.innerHTML = '';
            excludeWords.forEach(word => {
                const tag = document.createElement('span');
                tag.className = 'tag';
                tag.innerHTML = `${word} <span class="remove-tag" data-word="${word}">&times;</span>`;
                excludeTags.appendChild(tag);
            });
            document.querySelectorAll('#exclude-tags .remove-tag').forEach(btn => {
                btn.addEventListener('click', (e) => {
                    const word = e.target.dataset.word;
                    excludeWords = excludeWords.filter(w => w !== word);
                    updateExcludeTags();
                });
            });
        }

        

        function renderFolderTree(folders, container = document.getElementById('folder-tree')) {
            container.innerHTML = `
                <li data-id="inbox" style="padding: 6px 0 6px 16px; cursor: pointer; color: #ddd;">
                    <div style="display: flex; align-items: center;">
                        <span style="width: 16px; margin-right: 4px; display: inline-block;"></span>
                        <span class="folder-name" style="font-weight: 600;">📥 Все сообщения</span>
                    </div>
                </li>
            `;
            container.querySelector('li[data-id="inbox"]').addEventListener('click', (e) => {
                currentView = 'inbox';
                localStorage.setItem('activeFolder', currentView);
                loadMessages();
                syncActiveFolder();
            });
            folders.forEach(f => {
                if (f.name !== 'По умолчанию') {
                    buildFolderItem(f, container);
                }
            });
        }

        function syncActiveFolder() {
            document.querySelectorAll('.folder-tree li.active').forEach(el => el.classList.remove('active'));
            const activeLi = document.querySelector(`.folder-tree li[data-id="${currentView}"]`);
            if (activeLi) {
                activeLi.classList.add('active');
            }
        }

        function buildFolderItem(folder, parentElement) {
            const li = document.createElement('li');
            li.dataset.id = folder.id;

            const div = document.createElement('div');
            div.style.display = 'flex';
            div.style.alignItems = 'center';

            const toggle = document.createElement('span');
            toggle.className = 'folder-toggle';
            toggle.style.cursor = 'pointer';
            toggle.style.marginRight = '4px';
            toggle.style.width = '16px';
            toggle.style.display = 'inline-block';
            toggle.style.textAlign = 'center';

            const nameSpan = document.createElement('span');
            nameSpan.className = 'folder-name';
            nameSpan.textContent = folder.name;

            nameSpan.addEventListener('click', (e) => {
                e.stopPropagation();
                currentView = folder.id;
                localStorage.setItem('activeFolder', currentView);
                loadFolderMessages(folder.id);
                syncActiveFolder();
            });

            const actions = document.createElement('span');
            actions.className = 'folder-actions';
            actions.innerHTML = `
                <span class="add-subfolder" title="Создать подпапку">➕</span>
                <span class="delete-folder" title="Удалить папку">🗑️</span>
            `;
            actions.querySelector('.add-subfolder').addEventListener('click', (e) => {
                e.stopPropagation();
                const name = prompt('Введите название подпапки:');
                if (name) createSubfolder(folder.id, name);
            });
            actions.querySelector('.delete-folder').addEventListener('click', (e) => {
                e.stopPropagation();
                if (confirm('Удалить папку? Все сохранённые сообщения будут отвязаны.')) {
                    deleteFolder(folder.id);
                }
            });

            div.appendChild(toggle);
            div.appendChild(nameSpan);
            div.appendChild(actions);
            li.appendChild(div);

            if (folder.children && folder.children.length) {
            const ul = document.createElement('ul');
            ul.className = 'children';
            folder.children.forEach(child => buildFolderItem(child, ul));
            li.appendChild(ul);
            toggle.textContent = '▶';
            
            toggle.addEventListener('click', (e) => {
                e.stopPropagation();
                // ИДЕАЛЬНОЕ РЕШЕНИЕ: Просто вешаем класс, а CSS сам всё плавно откроет и повернет!
                li.classList.toggle('expanded');
            });
        } else {
            toggle.textContent = '▶';
            toggle.style.visibility = 'hidden';
        }
        parentElement.appendChild(li);
        }

        async function createSubfolder(parentId, name) {
            await apiFetch('/api/folders', {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify({ name, parent_id: parentId })
            });
            loadFolders();
        }
        async function renameFolder(folderId, newName) {
            try {
                const response = await apiFetch(`/api/folders/${folderId}/rename`, {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({ name: newName })
                });
                if (response && response.success) {
                    loadFolders(); // Обновляем дерево после успешного переименования
                } else {
                    alert('Ошибка переименования: ' + (response ? response.error : 'Неизвестная ошибка'));
                }
            } catch (e) {
                console.error('Ошибка переименования:', e);
            }
        }
        async function deleteFolder(folderId) {
            await apiFetch(`/api/folders/${folderId}`, { method: 'DELETE' });
            loadFolders();
            if (currentView === folderId) {
                currentView = 'inbox';
                localStorage.setItem('activeFolder', currentView);
                loadMessages();
                syncActiveFolder();
            }
        }

        document.getElementById('new-folder-btn').addEventListener('click', async () => {
            const name = prompt('Введите название папки:');
            if (!name) return;
            await apiFetch('/api/folders', {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify({ name })
            });
            loadFolders();
        });

        function loadFolderMessages(folderId) {
    // 1. Указываем системе, что мы переключились на папку
    currentView = 'folder_' + folderId;
    
    // 2. Очищаем поля поиска, чтобы старый поиск не мешал смотреть новую папку
    let sInp = document.getElementById('filter-search-input');
    let eInp = document.getElementById('filter-exclude-input');
    if (sInp) sInp.value = '';
    if (eInp) eInp.value = '';

    // 3. Вызываем нашу новую универсальную функцию!
    // Она сама скачает сообщения папки, уберет дубли, сделает пагинацию и правильные ссылки.
    loadMessages(1);
}
        


       

        
        const searchInputEl = document.getElementById('filter-search-input');
        
        let liveSearchTimeout = null;

        if (searchInputEl) {
            // 1. ЖИВОЙ ПОИСК: Реагируем на каждую напечатанную букву
            searchInputEl.addEventListener('input', function() {
            localStorage.setItem('filterSearchText', this.value);
                clearTimeout(liveSearchTimeout);
                liveSearchTimeout = setTimeout(() => {
                    if (typeof loadMessages === 'function') {
                        loadMessages(); // Обновляем таблицу спустя 0.3 сек после ввода
                    }
                }, 300); 
            });

            // 2. СОЗДАНИЕ ТЕГОВ: Оставляем возможность закрепить слово кнопкой Enter
            searchInputEl.addEventListener('keydown', (e) => {
                if (e.key === 'Enter' || e.key === ',') {
                    e.preventDefault();
                    const word = searchInputEl.value.trim();
                    if (word && !searchWords.includes(word)) {
                        searchWords.push(word);
                        updateSearchTags();
                    }
                    searchInputEl.value = ''; // Очищаем поле (слово стало тегом)
                    loadMessages(); // Окончательно обновляем таблицу
                }
            });
        }
function loadMessages(page = 1) { 
    currentPage = page; 
    let searchInput = document.getElementById('filter-search-input');
    let search = searchInput ? searchInput.value : ''; 
    let excludeInput = document.getElementById('filter-exclude-input'); 
    let exclude = excludeInput ? excludeInput.value : '';
    
    // --- НОВОЕ: Читаем выпадающие списки ---
    let senderSelect = document.getElementById('filter-sender');
    let sender = (senderSelect && senderSelect.value !== 'all') ? senderSelect.value : '';

    let typeSelect = document.getElementById('filter-type');
    let msgType = (typeSelect && typeSelect.value !== 'all') ? typeSelect.value : '';

    // Формируем URL и приклеиваем все параметры
    let url = `/api/messages?type=${currentView}&search=${encodeURIComponent(search)}&exclude=${encodeURIComponent(exclude)}&sender=${encodeURIComponent(sender)}&msg_type=${encodeURIComponent(msgType)}&page=${currentPage}&limit=${currentLimit}`;

    if (window.apiCache && window.apiCache[url]) { 
        renderMessagesData(window.apiCache[url]); 
        return;
    } 

    fetch(url) 
        .then(response => response.json()) 
        .then(data => { 
            if (window.apiCache) window.apiCache[url] = data; 
            renderMessagesData(data); 
        }) 
        .catch(err => { 
            console.error("Ошибка загрузки сообщений:", err); 
            let container = document.getElementById('messages-container');
            if(container) container.innerHTML = '<p style="color:red; text-align:center;">Ошибка загрузки сообщений.</p>'; 
        });
}

// --- ЖИВОЙ ПОИСК ---


function handleLiveSearch() {
    clearTimeout(liveSearchTimeout);
    // Ждем 500 миллисекунд после последнего нажатия клавиши
    liveSearchTimeout = setTimeout(() => {
        // Загружаем первую страницу с новыми параметрами поиска
        if (typeof loadMessages === 'function') {
            loadMessages(1); 
        }
    }, 500);
}

// Привязываем живой поиск к полям "Поиск" и "Исключить"
document.addEventListener("DOMContentLoaded", () => {
    let searchInput = document.getElementById('filter-search-input');
    let excludeInput = document.getElementById('filter-exclude-input');
    
    if (searchInput) searchInput.addEventListener('input', handleLiveSearch);
    if (excludeInput) excludeInput.addEventListener('input', handleLiveSearch);
});
// --- ФУНКЦИЯ ДЛЯ ПОДСВЕТКИ ТЕКСТА (ЗАЩИЩЕНА ОТ PYTHON/FLASK) ---
function highlightText(text) {
    if (!text) return '';
    
    // Встроенная защита от вредоносного кода
    const safeHtml = (str) => {
        let div = document.createElement('div');
        div.textContent = str;
        return div.innerHTML;
    };

    let searchInput = document.getElementById('filter-search-input');
    let searchStr = searchInput ? searchInput.value.trim() : '';
    
    let terms = [...(typeof searchWords !== 'undefined' ? searchWords : [])];
    
    // Разбиваем строку поиска на отдельные слова
    if (searchStr) {
        let words = searchStr.split(' ').filter(w => w.trim().length > 0);
        terms.push(...words);
    }
    
    terms = terms.filter(t => t.length > 0).sort((a, b) => b.length - a.length);
    let safeText = safeHtml(text);
    
    if (terms.length === 0) return safeText;
    
    // === 100% ЗАЩИТА ОТ FLASK ===
    // Собираем спецсимволы ВООБЩЕ БЕЗ СЛЕШЕЙ в коде (92 — это код слеша "\")
    const specialChars = '.*+?^${}()|[]/' + String.fromCharCode(92);
    const escapeRegExp = (s) => s.split('').map(c => specialChars.includes(c) ? String.fromCharCode(92) + c : c).join('');
    // =============================

    const escapedTerms = terms.map(escapeRegExp);
    const regex = new RegExp('(' + escapedTerms.join('|') + ')', 'gi');
    
    return safeText.replace(regex, '<span style="background-color: rgba(74, 144, 226, 0.2); color: #4a90e2; font-weight: bold; padding: 0 3px; border-radius: 3px;">$1</span>');
}

function renderPagination(total, page, totalPages) {
    let container = document.getElementById('pagination-container');
    
    // Защита от ошибок
    if (!container) return;

    // Если сообщений нет вообще, прячем пагинацию
    if (!total || total === 0) {
        container.innerHTML = '';
        return;
    }

    // Собираем HTML для панели управления страницами
    let html = `
        <div style="display: flex; justify-content: space-between; align-items: center; background: #1e1e1e; padding: 10px 15px; border-radius: 8px; margin-bottom: 20px; box-shadow: 0 4px 6px rgba(0,0,0,0.1);">
            
            <div style="display: flex; align-items: center; gap: 10px;">
                <span style="color: #bbb; font-size: 14px;">Показывать по:</span>
                <select onchange="changeLimit(this.value)" style="background: #2b2b2b; color: #fff; border: 1px solid #444; border-radius: 4px; padding: 5px; outline: none; cursor: pointer;">
                    <option value="20" ${currentLimit === 20 ? 'selected' : ''}>20</option>
                    <option value="40" ${currentLimit === 40 ? 'selected' : ''}>40</option>
                    <option value="60" ${currentLimit === 60 ? 'selected' : ''}>60</option>
                    <option value="80" ${currentLimit === 80 ? 'selected' : ''}>80</option>
                    <option value="100" ${currentLimit === 100 ? 'selected' : ''}>100</option>
                </select>
            </div>
            
            <div style="display: flex; align-items: center; gap: 15px;">
                <button 
                    onclick="loadMessages(${page - 1})" 
                    ${page <= 1 ? 'disabled style="opacity:0.4; cursor:not-allowed;"' : 'style="cursor:pointer;"'} 
                    class="action-btn">
                    &laquo; Назад
                </button>
                
                <span style="color: #fff; font-size: 14px; background: #2b2b2b; padding: 5px 10px; border-radius: 4px;">
                    Стр. <b style="color: #4a90e2;">${page}</b> из <b>${totalPages}</b> 
                    <span style="color: #888; font-size: 12px; margin-left: 5px;">(Всего: ${total})</span>
                </span>
                
                <button 
                    onclick="loadMessages(${page + 1})" 
                    ${page >= totalPages ? 'disabled style="opacity:0.4; cursor:not-allowed;"' : 'style="cursor:pointer;"'} 
                    class="action-btn">
                    Вперед &raquo;
                </button>
            </div>
        </div>
    `;
    
    container.innerHTML = html;
}

// 5. Функция смены лимита сообщений
function changeLimit(newLimit) {
    currentLimit = parseInt(newLimit);
    // При изменении лимита (например, с 20 на 100) обязательно сбрасываем на 1-ю страницу
    loadMessages(1); 
}

function renderMessagesData(data) {
    let container = document.getElementById('messages-tbody'); 
    let noMessages = document.getElementById('no-messages');
    let tableHeader = document.querySelector('#parser-section .messages-table thead');
    
    if (!container) return;

    // --- 1. ОБНОВЛЕНИЕ СПИСКА ОТПРАВИТЕЛЕЙ (ФИЛЬТР) ---
    let senderSelect = document.getElementById('filter-sender') || document.querySelector('select[id*="sender"]');
    if (senderSelect && data.messages) {
        let currentSelected = senderSelect.value;
        let uniqueSenders = new Set();
        
        Array.from(senderSelect.options).forEach(opt => {
            if (opt.value && opt.value !== 'all' && opt.value !== '') uniqueSenders.add(opt.value);
        });

        data.messages.forEach(msg => {
            let sender = typeof getSenderDisplayName === 'function' ? getSenderDisplayName(msg) : (msg.sender_name || 'Неизвестно');
            if (msg.type === 'private' || sender !== 'Неизвестно') {
                uniqueSenders.add(sender);
            }
        });

        senderSelect.innerHTML = '<option value="all">Все отправители</option>';
        Array.from(uniqueSenders).sort().forEach(sender => {
            let option = document.createElement('option');
            option.value = sender;
            option.textContent = sender;
            if (sender === currentSelected) option.selected = true;
            senderSelect.appendChild(option);
        });
    }

    // --- 2. ГЕНЕРАЦИЯ ТАБЛИЦЫ СООБЩЕНИЙ ---
    let html = '';

    if (data.messages && data.messages.length > 0) {
        if (noMessages) noMessages.style.display = 'none';
        if (tableHeader) tableHeader.style.display = '';

        const safeHtml = (str) => {
            if (!str) return '';
            let div = document.createElement('div');
            div.textContent = str;
            return div.innerHTML;
        };

        data.messages.forEach(msg => {
            const isExcel = msg.type && msg.type.startsWith('excel');
            const typeIcon = msg.type === 'group' ? '👥' : msg.type === 'channel' ? '📢' : (isExcel ? '📊' : '💬');
            const displaySender = typeof getSenderDisplayName === 'function' ? getSenderDisplayName(msg) : (msg.sender_name || 'Неизвестно');
            const rawText = msg.message_text || msg.text || ''; 
            const highlightedText = typeof highlightText === 'function' ? highlightText(rawText) : safeHtml(rawText);
            let dateStr = msg.date || (msg.timestamp ? new Date(msg.timestamp).toLocaleString() : 'Нет даты');

            let originalTitle = safeHtml(msg.chat_title || 'Личный чат / Бот');
            let customName = safeHtml(msg.custom_name || msg.chat_title || 'Личный чат / Бот');
            
            let nameBlockHtml = '';
            if (msg.custom_name && msg.custom_name !== msg.chat_title) {
                nameBlockHtml = `<div style="color:#888; font-size:11px; margin-bottom:2px;">${originalTitle}</div>`;
            }
            nameBlockHtml += `<strong style="font-size:14px;">${customName}</strong>`;

            let chatDisplayHtml = '';

            if (isExcel) {
                // СЦЕНАРИЙ 1: Excel
                chatDisplayHtml = `
                    <div style="color: #2ecc71; display: inline-block;" title="Файл на сервере / Таблица">
                        ${nameBlockHtml}
                    </div>`;
            } else if (msg.type === 'channel' || msg.type === 'group') {
                // СЦЕНАРИЙ 2: Канал или Группа
                let cleanChatId = msg.chat_id ? String(msg.chat_id).replace(/^-100|^ -|-/g, '') : '';
                chatDisplayHtml = `
                    <a href="https://t.me/c/${cleanChatId}/${msg.telegram_message_id || ''}" target="_blank" title="Открыть сообщение в Telegram" 
                       style="text-decoration: none; color: #4a90e2; display: inline-block; transition: opacity 0.2s;" 
                       onmouseover="this.style.opacity='0.7'; this.style.textDecoration='underline';" 
                       onmouseout="this.style.opacity='1'; this.style.textDecoration='none';">
                       ${nameBlockHtml} <span style="font-size:12px; font-weight:bold;">↗</span>
                    </a>`;
           } else {
                // СЦЕНАРИЙ 3: Боты и Личные сообщения
                let botUsername = null;
                
                // Собираем ВСЕ поля, где мог прийти юзернейм от сервера
                let possibleNames = [
                    msg.custom_name, 
                    msg.chat_title, 
                    msg.tc_chat_title, 
                    msg.sender_username,
                    msg.sender_name
                ];
                
                // 1. Агрессивно ищем любую строку с собачкой @
                for (let name of possibleNames) {
                    if (name && typeof name === 'string' && name.includes('@')) {
                        let match = name.match(/@([a-zA-Z0-9_]+)/);
                        if (match) {
                            botUsername = match[1];
                            break;
                        }
                    }
                }

                // 2. Если собачки нигде нет, но имя состоит из одного слова и заканчивается на "bot"
                if (!botUsername) {
                    for (let name of possibleNames) {
                        if (name && typeof name === 'string') {
                            let cleanName = name.trim();
                            if (cleanName.toLowerCase().endsWith('bot') && !cleanName.includes(' ')) {
                                botUsername = cleanName;
                                break;
                            }
                        }
                    }
                }

                if (botUsername) {
                    // Если нашли, делаем красивую t.me ссылку
                    let cleanUsername = botUsername.replace('@', '');
                    chatDisplayHtml = `
                        <a href="https://t.me/${cleanUsername}" target="_blank" title="Открыть бота в Telegram" 
                           style="text-decoration: none; color: #4a90e2; display: inline-block; transition: opacity 0.2s;" 
                           onmouseover="this.style.opacity='0.7'; this.style.textDecoration='underline';" 
                           onmouseout="this.style.opacity='1'; this.style.textDecoration='none';">
                           ${nameBlockHtml} <span style="font-size:12px; font-weight:bold;">↗</span>
                        </a>`;
                } else {
                    // Если совсем ничего нет, открываем по ID
                    let cleanUserId = msg.chat_id ? String(msg.chat_id).replace(/^-/, '') : '';
                    chatDisplayHtml = `
                        <a href="tg://openmessage?user_id=${cleanUserId}" title="Открыть чат" 
                           style="text-decoration: none; color: #4a90e2; display: inline-block; transition: opacity 0.2s;" 
                           onmouseover="this.style.opacity='0.7'; this.style.textDecoration='underline';" 
                           onmouseout="this.style.opacity='1'; this.style.textDecoration='none';">
                           ${nameBlockHtml} <span style="font-size:12px; font-weight:bold;">↗</span>
                        </a>`;
                }
            }

            let senderHtml = '';
            if (displaySender !== 'Неизвестно' || msg.type === 'private' || isExcel) {
                let icon = isExcel ? '📑' : '👤';
                senderHtml = `<br><span style="color:#888; font-size:12px; display:inline-block; margin-top:4px;">${icon} ${safeHtml(displaySender)}</span>`;
            }

            // Защита: если ID пустой, передаем 'null' чтобы не сломать JS синтаксис
            let safeMessageId = msg.id ? msg.id : 'null';

            html += `
            <tr style="transition: background 0.2s; border-bottom: 1px solid #333;" onmouseover="this.style.background='#2a2a2a'" onmouseout="this.style.background='transparent'">
                <td style="text-align:center; font-size:18px;" title="Тип: ${msg.type}">${typeIcon}</td>
                <td>
                    ${chatDisplayHtml}
                    ${senderHtml}
                </td>
                <td style="white-space: pre-wrap; max-width: 500px; word-break: break-word; line-height: 1.4;">${highlightedText}</td>
                <td style="white-space: nowrap; font-size:12px; color:#aaa;">${dateStr}</td>
                <td style="white-space: nowrap; vertical-align: middle;">
                    <div style="display:flex; flex-direction:column; gap:6px;">
                        <button class="save-btn" onclick="window.blockMessage(${safeMessageId}, this);" 
                                style="background:transparent; border:1px solid #e74c3c; color:#e74c3c; padding:4px 8px; font-size:12px; transition:0.2s; cursor:pointer;" 
                                onmouseover="this.style.background='#e74c3c'; this.style.color='#fff';" 
                                onmouseout="this.style.background='transparent'; this.style.color='#e74c3c';">
                            🚫 Удалить
                        </button>
                    </div>
                </td>
            </tr>`;
        });
    } else {
        if (noMessages) noMessages.style.display = 'block';
        if (tableHeader) tableHeader.style.display = 'none';
        html = '<tr><td colspan="5" style="color:#888; text-align:center; padding: 20px;">По данному запросу сообщений не найдено.</td></tr>';
    }
    
    container.innerHTML = html;
    
    // --- 3. ПАГИНАЦИЯ ---
    if (typeof renderPagination === 'function') {
        renderPagination(data.total, data.page, data.total_pages);
    }
}

   // НОВАЯ ФУНКЦИЯ: Умное определение имени (убирает баг "None" и подставляет кастомные имена)
   // Умное определение имени (теперь игнорирует авторов и в каналах, и в группах)
    function getSenderDisplayName(msg) {
        if (msg.custom_name) return msg.custom_name;
        // Если это канал ИЛИ группа - берем только название чата, игнорируем автора
        if (msg.type === 'channel' || msg.type === 'group') return msg.chat_title || 'Чат';
        
        let name = msg.sender_name || '';
        name = name.replace(/ None/g, '').trim();
        if (!name || name === 'Unknown') return msg.chat_title || 'Неизвестно';
        return name;
    }

    // ИСПРАВЛЕННАЯ ФУНКЦИЯ ФИЛЬТРА: Теперь она использует красивые имена
 
   // ИСПРАВЛЕННОЕ ФОРМИРОВАНИЕ СПИСКА (теперь берет чистые имена)
function updateSenderFilterOptions(messages) {
    const select = document.getElementById('filter-sender');
    if (!select) return;
    const currentVal = select.value;

    // ЗДЕСЬ ИЗМЕНЕНИЕ: используем getSenderDisplayName вместо сырого m.sender_name
    const senders = [...new Set(messages.map(m => getSenderDisplayName(m)).filter(Boolean))].sort();

    let options = '<option value="all">Все отправители</option>';
    senders.forEach(s => {
        options += `<option value="${s.replace(/"/g, '&quot;')}">${s}</option>`;
    });
    select.innerHTML = options;

    if (currentVal !== 'all' && senders.includes(currentVal)) {
        select.value = currentVal;
    } else {
        select.value = 'all';
    }
}
    // ИСПРАВЛЕННЫЙ ПАРСЕР: Рисует таблицу и применяет фильтр "От кого"
    // ИСПРАВЛЕННЫЙ ПАРСЕР: Рисует таблицу, применяет фильтр "От кого" и ПОДСВЕЧИВАЕТ ПОИСК
   // ИСПРАВЛЕННЫЙ ПАРСЕР: Рисует таблицу для папок и делает ИСТОЧНИКИ КЛИКАБЕЛЬНЫМИ
function renderMessages(messages) {
    const tbody = document.getElementById('messages-tbody');
    const noMessages = document.getElementById('no-messages');
    const tableHeader = document.querySelector('#parser-section .messages-table thead');

    if (typeof updateSenderFilterOptions === 'function') updateSenderFilterOptions(messages);

    const senderFilterVal = document.getElementById('filter-sender')?.value || 'all';
    let filteredMessages = messages;
    if (senderFilterVal !== 'all') {
        filteredMessages = messages.filter(m => getSenderDisplayName(m) === senderFilterVal);
    }

    if (!filteredMessages || filteredMessages.length === 0) {
        if (tbody) tbody.innerHTML = '';
        if (noMessages) noMessages.style.display = 'block';
        if (tableHeader) tableHeader.style.display = 'none';
        return;
    }

    if (noMessages) noMessages.style.display = 'none';
    if (tableHeader) tableHeader.style.display = '';

    // Вспомогательная функция для безопасности
    const safeHtml = (str) => {
        if (!str) return '';
        let div = document.createElement('div');
        div.textContent = str;
        return div.innerHTML;
    };

 
function highlightText(text) {
    if (!text) return '';
    
    // Встроенная защита от вредоносного кода
    const safeHtml = (str) => {
        let div = document.createElement('div');
        div.textContent = str;
        return div.innerHTML;
    };

    let searchInput = document.getElementById('filter-search-input');
    let searchStr = searchInput ? searchInput.value.trim() : '';
    
    let terms = [...(typeof searchWords !== 'undefined' ? searchWords : [])];
    
    // Разбиваем строку поиска на отдельные слова
    if (searchStr) {
        let words = searchStr.split(' ').filter(w => w.trim().length > 0);
        terms.push(...words);
    }
    
    terms = terms.filter(t => t.length > 0).sort((a, b) => b.length - a.length);
    let safeText = safeHtml(text);
    
    if (terms.length === 0) return safeText;
    
    // === 100% ЗАЩИТА ОТ FLASK ===
    // Собираем спецсимволы ВООБЩЕ БЕЗ СЛЕШЕЙ в коде (92 — это код слеша "\")
    const specialChars = '.*+?^${}()|[]/' + String.fromCharCode(92);
    const escapeRegExp = (s) => s.split('').map(c => specialChars.includes(c) ? String.fromCharCode(92) + c : c).join('');
    // =============================

    const escapedTerms = terms.map(escapeRegExp);
    const regex = new RegExp('(' + escapedTerms.join('|') + ')', 'gi');
    
    return safeText.replace(regex, '<span style="background-color: rgba(74, 144, 226, 0.2); color: #4a90e2; font-weight: bold; padding: 0 3px; border-radius: 3px;">$1</span>');
}

    let html = '';
    filteredMessages.forEach(msg => {
        const isExcel = msg.type && msg.type.startsWith('excel');
        const isTGGroupChannel = msg.chat_id && msg.telegram_message_id && (msg.type === 'channel' || msg.type === 'group');
        
        const typeIcon = msg.type === 'group' ? '👥' : msg.type === 'channel' ? '📢' : (isExcel ? '📊' : '💬');
        const displaySender = getSenderDisplayName(msg);
        const rawText = msg.message_text || msg.text || ''; 
        const highlightedText = highlightText(rawText);
        let dateStr = msg.date || (msg.timestamp ? new Date(msg.timestamp).toLocaleString() : 'Нет даты');

        // --- БЛОК ОТОБРАЖЕНИЯ ИСТОЧНИКА ---
        let originalTitle = safeHtml(msg.chat_title || 'Личный чат / Бот');
        let customName = safeHtml(msg.custom_name || msg.chat_title || 'Личный чат / Бот');
        
        let nameBlockHtml = '';
        if (msg.custom_name && msg.custom_name !== msg.chat_title) {
            nameBlockHtml = `<div style="color:#888; font-size:11px; margin-bottom:2px;">${originalTitle}</div>`;
        }
        nameBlockHtml += `<strong style="font-size:14px;">${customName}</strong>`;

        // Ищем юзернейм (он начинается с @ в базе)
        let botUsername = msg.tc_chat_title && msg.tc_chat_title.startsWith('@') ? msg.tc_chat_title : null;

        let chatDisplayHtml = '';

        if (isTGGroupChannel) {
            let cleanChatId = String(msg.chat_id).replace(/^-100|^ -/, '');
            chatDisplayHtml = `
                <a href="https://t.me/c/${cleanChatId}/${msg.telegram_message_id}" target="_blank" title="Открыть сообщение в Telegram" 
                   style="text-decoration: none; color: #4a90e2; display: inline-block; transition: opacity 0.2s;" 
                   onmouseover="this.style.opacity='0.7'; this.style.textDecoration='underline';" 
                   onmouseout="this.style.opacity='1'; this.style.textDecoration='none';">
                   ${nameBlockHtml} <span style="font-size:12px; font-weight:bold;">↗</span>
                </a>`;
        } 
        else if (!isExcel) {
            if (botUsername) {
                // Идеальная ситуация: у нас есть @username бота!
                let cleanUsername = botUsername.replace('@', '');
                chatDisplayHtml = `
                    <a href="https://t.me/${cleanUsername}" target="_blank" title="Открыть чат с ботом" 
                       style="text-decoration: none; color: #4a90e2; display: inline-block; transition: opacity 0.2s;" 
                       onmouseover="this.style.opacity='0.7'; this.style.textDecoration='underline';" 
                       onmouseout="this.style.opacity='1'; this.style.textDecoration='none';">
                       ${nameBlockHtml} <span style="font-size:12px; font-weight:bold;">↗</span>
                    </a>`;
            } else {
                // Если юзернейма нет, пытаемся открыть по ID
                let cleanUserId = msg.chat_id ? String(msg.chat_id).replace(/^-/, '') : '';
                chatDisplayHtml = `
                    <a href="tg://openmessage?user_id=${cleanUserId}" title="Открыть чат (Telegram)" 
                       style="text-decoration: none; color: #4a90e2; display: inline-block; transition: opacity 0.2s;" 
                       onmouseover="this.style.opacity='0.7'; this.style.textDecoration='underline';" 
                       onmouseout="this.style.opacity='1'; this.style.textDecoration='none';">
                       ${nameBlockHtml} <span style="font-size:12px; font-weight:bold;">↗</span>
                    </a>`;
            }
        } 
        else if (isExcel) {
            chatDisplayHtml = `
                <div style="color: #2ecc71; display: inline-block;" title="Файл на сервере / Таблица">
                    ${nameBlockHtml}
                </div>`;
        }

        let senderHtml = '';
        if (displaySender !== 'Неизвестно' || msg.type === 'private' || isExcel) {
            let icon = isExcel ? '📑' : '👤';
            senderHtml = `<br><span style="color:#888; font-size:12px; display:inline-block; margin-top:4px;">${icon} ${safeHtml(displaySender)}</span>`;
        }

        html += `
        <tr style="transition: background 0.2s; border-bottom: 1px solid #333;" onmouseover="this.style.background='#2a2a2a'" onmouseout="this.style.background='transparent'">
            <td style="text-align:center; font-size:18px;" title="Тип: ${msg.type}">${typeIcon}</td>
            <td>
                ${chatDisplayHtml}
                ${senderHtml}
            </td>
            <td style="white-space: pre-wrap; max-width: 500px; word-break: break-word; line-height: 1.4;">${highlightedText}</td>
            <td style="white-space: nowrap; font-size:12px; color:#aaa;">${dateStr}</td>
            <td style="white-space: nowrap; vertical-align: middle;">
                <div style="display:flex; flex-direction:column; gap:6px;">
                    <button class="save-btn" onclick="window.blockMessage(${msg.id}, this);" 
                            style="background:transparent; border:1px solid #e74c3c; color:#e74c3c; padding:4px 8px; font-size:12px; transition:0.2s; cursor:pointer;" 
                            onmouseover="this.style.background='#e74c3c'; this.style.color='#fff';" 
                            onmouseout="this.style.background='transparent'; this.style.color='#e74c3c';">
                        🚫 Удалить
                    </button>
                </div>
            </td>
        </tr>`;
    });

    if (tbody) tbody.innerHTML = html;
}


// Перенесите инициализацию всех фильтров в одну функцию
function initInboxFilters() {
    const chatInput = document.getElementById('filter-chat-input');
    const searchInput = document.getElementById('filter-search-input');
    const excludeInput = document.getElementById('filter-exclude-input');

    if (chatInput && !chatInput.dataset.bound) {
        chatInput.addEventListener('keypress', function(e) {
            if (e.key === 'Enter' && this.value.trim()) {
                const val = this.value.trim();
                if (!activeChatFilters.includes(val)) {
                    activeChatFilters.push(val);
                    renderFilterTags('chat-tags', activeChatFilters);
                }
                this.value = '';
            }
        });
        chatInput.dataset.bound = "true";
    }

    // Слушатели для Поиска и Исключений (уже есть в вашем коде[cite: 480, 487], но лучше перенести сюда для безопасности)
    if (searchInput && !searchInput.dataset.bound) {
        searchInput.dataset.bound = "true";
        // Ваш существующий код для searchInput...
    }
}





        
   window.blockMessage = async function(id, btnElement) {
    if (confirm('Точно удалить это сообщение? Оно навсегда исчезнет из парсера.')) {
        
        // Визуально отключаем кнопку, чтобы не нажать дважды
        if (btnElement) {
            btnElement.innerText = 'Удаление...';
            btnElement.style.pointerEvents = 'none';
        }
        
        try {
            const response = await fetch(`/api/messages/${id}/block`, { method: 'POST' });
            const data = await response.json();
            
            if (data.success) {
                window.apiCache = {}; // Очищаем кэш
                if (typeof currentView !== 'undefined' && currentView === 'inbox') {
                    loadMessages();
                } else {
                    loadFolderMessages(currentView);
                }
            } else {
                alert('Ошибка сервера: ' + data.error);
                if (btnElement) { btnElement.innerText = '🚫 Удалить'; btnElement.style.pointerEvents = 'auto'; }
            }
        } catch(e) {
            alert('Ошибка сети: ' + e.message);
            if (btnElement) { btnElement.innerText = '🚫 Удалить'; btnElement.style.pointerEvents = 'auto'; }
        }
    }
};
       
  // 2. ФУНКЦИИ УПРАВЛЕНИЯ ЮЗЕРБОТАМИ
 async function loadUserbots() {
    const bots = await cachedApiFetch('/api/userbots');
    const container = document.getElementById('userbot-list');
    if (!container) return;
    
    if (!bots.length) {
        container.innerHTML = '<p>Нет добавленных аккаунтов.</p>';
        return;
    }
    
    let html = '<table class="simple-table"><tr><th>Аккаунт</th><th>Статус</th><th>Время</th><th>Действия</th></tr>';
    
    bots.forEach(bot => {
        const ubotName = bot.account_name || 'Неизвестно';
        const isChecked = bot.schedule_enabled ? 'checked' : '';
        
        let statusHtml = '';
        let toggleBtnHtml = '';
        
        if (bot.status === 'active') {
            statusHtml = '🟢 Активен';
            toggleBtnHtml = `<button class="toggle-bot save-btn" style="padding:4px 8px; font-size:12px;" data-id="${bot.id}">Остановить</button>`;
        } else if (bot.status === 'unauthorized') {
            statusHtml = '<span style="color:#e74c3c; font-weight:bold;">🔴 Слетела авторизация</span>';
            // Наша синяя кнопка для повторной авторизации
            toggleBtnHtml = `<button class="reauth-bot save-btn" style="background:#3498db; padding:4px 8px; font-size:12px;" data-id="${bot.id}" data-api="${bot.api_id}" data-hash="${bot.api_hash}">Авторизоваться</button>`; 
        } else {
            statusHtml = '⚪ Выключен';
            toggleBtnHtml = `<button class="toggle-bot save-btn" style="padding:4px 8px; font-size:12px;" data-id="${bot.id}">Запустить</button>`;
        }
        
        html += `
            <tr>
                <td>
                    <strong style="color:#4a90e2;">${ubotName}</strong><br>
                    <span style="font-size:11px; color:#888;">API ID: ${bot.api_id}</span>
                </td>
                <td>${statusHtml}</td>
                <td>
                    <div style="display:flex; gap:5px; align-items:center; font-size: 13px;">
                        <input type="checkbox" id="sch_en_${bot.id}" ${isChecked} title="Включить расписание по времени">
                        <input type="time" id="sch_start_${bot.id}" value="${bot.time_start || '00:00'}" style="background:#1e1e1e; color:#fff; border:1px solid #3a3a3a; padding:4px; border-radius:4px;">
                        - 
                        <input type="time" id="sch_end_${bot.id}" value="${bot.time_end || '23:59'}" style="background:#1e1e1e; color:#fff; border:1px solid #3a3a3a; padding:4px; border-radius:4px;">
                        <button class="save-btn" onclick="saveSchedule(${bot.id})" style="padding:4px 8px; font-size:14px; margin-left:5px;" title="Сохранить тайминг">💾</button>
                    </div>
                </td>
                <td>
                    ${toggleBtnHtml}
                    <button class="delete-bot save-btn" style="background:#e74c3c; padding:4px 8px; font-size:12px;" data-id="${bot.id}">Удалить</button>
                </td>
            </tr>`;
    });
    
    html += '</table>';
    container.innerHTML = html;

    // --- ОБРАБОТЧИКИ КНОПОК ---

    // 1. Кнопка Запустить / Остановить
    container.querySelectorAll('.toggle-bot').forEach(btn => {
        btn.addEventListener('click', async () => {
            await apiFetch(`/api/userbots/${btn.dataset.id}/toggle`, { method: 'POST' });
            loadUserbots();
        });
    });

    // 2. Кнопка Авторизоваться (при слетевшей сессии)
    // 2. Кнопка Авторизоваться (при слетевшей сессии)
    container.querySelectorAll('.reauth-bot').forEach(btn => {
        btn.addEventListener('click', async () => {
            const form = document.getElementById('bot-form-container');
            form.style.display = 'block';
            form.scrollIntoView({ behavior: 'smooth' });
            
            // Подставляем старые данные
            document.getElementById('api_id').value = btn.dataset.api;
            document.getElementById('api_hash').value = btn.dataset.hash;
            
            // Записываем ID сессии в скрытое поле
            const sessionInput = document.getElementById('auth_session_id');
            if (sessionInput) sessionInput.value = btn.dataset.id;
            
            // Автоматически запускаем генерацию QR для этой сессии
            await window.generateQR(btn.dataset.api, btn.dataset.hash, btn.dataset.id);
        });
    });

    // 3. Кнопка Удалить
    container.querySelectorAll('.delete-bot').forEach(btn => {
        btn.addEventListener('click', async () => {
            if (confirm('Удалить юзербота?')) {
                await apiFetch(`/api/userbots/${btn.dataset.id}`, { method: 'DELETE' });
                loadUserbots();
                loadTrackedChats();
            }
        });
    });
} // <--- Конец функции loadUserbots
    // НОВОЕ: Обработчик для кнопки переавторизации
        
    // НОВАЯ ФУНКЦИЯ ДЛЯ ОТПРАВКИ ТАЙМИНГА НА СЕРВЕР
    window.saveSchedule = async function(botId) {
        const enabled = document.getElementById(`sch_en_${botId}`).checked ? 1 : 0;
        const start = document.getElementById(`sch_start_${botId}`).value;
        const end = document.getElementById(`sch_end_${botId}`).value;
        
        await apiFetch(`/api/userbots/${botId}/schedule`, {
            method: 'POST',
            headers: { 'Content-Type': 'application/json' },
            body: JSON.stringify({ schedule_enabled: enabled, time_start: start, time_end: end })
        });
        
        // Маленькое уведомление, чтобы понимать, что сохранилось
        alert('✅ Расписание для юзербота успешно сохранено!');
        loadUserbots();
    };

        
// Обработчики для кнопок редактирования


// Обработчики для кнопок удаления (уже были)
document.querySelectorAll('.delete-chat').forEach(btn => {
        btn.addEventListener('click', async () => {
            const id = btn.dataset.id;
            if (confirm('Удалить чат из отслеживаемых?')) {
                await apiFetch(`/api/tracked_chats/${id}`, { method: 'DELETE' });
                loadTrackedChats();
            }
        });
    });


        

        document.getElementById('add-chat-btn')?.addEventListener('click', async () => {
            const chatId = document.getElementById('new-chat-id').value.trim();
            const customName = document.getElementById('new-chat-name').value.trim(); // Получаем кастомное имя
            if (!chatId) return;
            
            try {
                await apiFetch('/api/tracked_chats', {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({ chat_id: chatId, custom_name: customName }) // Отправляем кастомное имя
                });
                
                // Очищаем оба поля после успешного добавления
                document.getElementById('new-chat-id').value = '';
                document.getElementById('new-chat-name').value = '';
                
                loadTrackedChats();
            } catch (error) {
                console.log("Добавление отменено:", error.message);
            }
        });

        

        const addUserForm = document.getElementById('add-user-form');
        if (addUserForm) {
            addUserForm.addEventListener('submit', async (e) => {
                e.preventDefault();
                // Берем данные по новым ID
                const login = document.getElementById('admin-new-login').value;
                const password = document.getElementById('admin-new-password').value;
        
                await apiFetch('/api/admin/add_user', {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({ login, password })
                });
                alert('Пользователь создан');
                // Очищаем форму после успеха
                document.getElementById('admin-new-login').value = '';
                document.getElementById('admin-new-password').value = '';
        
                loadUsers();
            });
        }

        async function loadUsers() {
            const users = await apiFetch('/api/admin/users');
            const list = document.getElementById('user-list');
            if (list) {
                list.innerHTML = users.map(u => `<li>${u.login} (${u.role})</li>`).join('');
            }
        }

        if (document.getElementById('admin-section')) {
            loadUsers();
        }

        checkSession();

        function toggleAllConf(el) {
            document.querySelectorAll('.conf-cb').forEach(cb => cb.checked = el.checked);
            toggleMergeBtn();
        }

        function toggleMergeBtn() {
            const checkedBoxes = document.querySelectorAll('.conf-cb:checked');
            const btn = document.getElementById('btn-merge-conf');
            
            if (!btn) return;

            // Если выбрано меньше 2 товаров — скрываем кнопку
            if (checkedBoxes.length < 2) {
                btn.style.display = 'none';
                return;
            }

            // Берем данные "От кого" и "Откуда" у ПЕРВОЙ выбранной галочки
            const firstSender = checkedBoxes[0].dataset.sender;
            const firstChat = checkedBoxes[0].dataset.chat;
            
            // Проверяем, совпадают ли они у всех остальных выбранных
            let allMatch = true;
            for (let i = 1; i < checkedBoxes.length; i++) {
                if (checkedBoxes[i].dataset.sender !== firstSender || checkedBoxes[i].dataset.chat !== firstChat) {
                    allMatch = false;
                    break;
                }
            }

            // Показываем кнопку только если все совпадает
            btn.style.display = allMatch ? 'block' : 'none';
        }

        async function mergeSelectedConfirmed() {
            const checked = Array.from(document.querySelectorAll('.conf-cb:checked')).map(cb => parseInt(cb.value));
            if (checked.length < 2) return;
            
            await apiFetch('/api/product_messages/merge', {
                method: 'POST',
                headers: {'Content-Type': 'application/json'},
                body: JSON.stringify({ pm_ids: checked })
            });
            
            // После объединения сразу обновляем таблицу
            refreshProductBinding();
        }


        let allProducts = []; // Добавляем глобальную переменную

        async function loadFolders() {
            allProducts = await cachedApiFetch('/api/products');
            const folders = await cachedApiFetch('/api/folders/tree');
    
            renderFolderTree(folders);
            syncActiveFolder();
            
            const folderSelect = document.getElementById('folder-select');
            const prodFolderSelect = document.getElementById('prod_folder_id');
            
            if(folderSelect) folderSelect.innerHTML = '';
            if(prodFolderSelect) prodFolderSelect.innerHTML = '<option value="">Без папки (Общие)</option>';
            
            function addOptions(list, level=0) {
                list.forEach(f => {
                    if (f.name === 'По умолчанию') return; 
                    
                    // Убрали проверку isOccupied. Теперь папки всегда доступны.
                    const labelText = '—'.repeat(level) + ' ' + f.name;

                    if (folderSelect) {
                        const opt = document.createElement('option');
                        opt.value = f.id;
                        opt.textContent = labelText;
                        folderSelect.appendChild(opt);
                    }
                    if (prodFolderSelect) {
                        const opt2 = document.createElement('option');
                        opt2.value = f.id;
                        opt2.textContent = labelText;
                        prodFolderSelect.appendChild(opt2);
                    }
                    if (f.children) addOptions(f.children, level+1);
                });
            }
            
            // ВОТ ЭТИХ ДВУХ СТРОК НЕ ХВАТАЛО:
            addOptions(folders); 
        } // <--- Эта скобка закрывает loadFolders!

        function renderFolderTree(folders, container = document.getElementById('folder-tree')) {
            container.innerHTML = `
                <li data-id="inbox" style="padding: 6px 0 6px 16px; cursor: pointer; color: #ddd;">
                    <div style="display: flex; align-items: center;">
                        <span style="width: 16px; margin-right: 4px; display: inline-block;"></span>
                        <span class="folder-name" style="font-weight: 600;">📥 Все сообщения</span>
                    </div>
                </li>
            `;
            container.querySelector('li[data-id="inbox"]').addEventListener('click', (e) => {
                currentView = 'inbox';
                localStorage.setItem('activeFolder', currentView);
                document.getElementById('folder-catalog-section').style.display = 'none';
                document.getElementById('parser-section').style.display = 'block';
                loadMessages();
                syncActiveFolder();
            });
            folders.forEach(f => {
                if (f.name !== 'По умолчанию') buildFolderItem(f, container);
            });
        }

function buildFolderItem(folder, parentElement) {
            const li = document.createElement('li');
            li.dataset.id = folder.id;

            const div = document.createElement('div');
            div.style.display = 'flex';
            div.style.alignItems = 'center';

            const toggle = document.createElement('span');
            toggle.className = 'folder-toggle';
            toggle.style.cursor = 'pointer';
            toggle.style.marginRight = '4px';
            toggle.style.width = '16px';
            toggle.style.display = 'inline-block';
            toggle.style.textAlign = 'center';

            const nameSpan = document.createElement('span');
            nameSpan.className = 'folder-name';
            nameSpan.textContent = folder.name;

            // КЛИК ПО ПАПКЕ ОТКРЫВАЕТ КАТАЛОГ ТОВАРОВ
            nameSpan.addEventListener('click', (e) => {
                e.stopPropagation();
                currentView = folder.id;
                localStorage.setItem('activeFolder', currentView);
                syncActiveFolder();
                
                document.getElementById('parser-section').style.display = 'none';
                document.getElementById('folder-catalog-section').style.display = 'block';
                loadFolderCatalog(folder.id);
            });

            // ПРОВЕРЯЕМ ЕСТЬ ЛИ В ПАПКЕ ТОВАР. Если есть - скрываем "создать подпапку"
            const hasProduct = allProducts.some(p => p.folder_id === folder.id);
            const actions = document.createElement('span');
            actions.className = 'folder-actions';
            actions.innerHTML = `
            <span class="rename-folder" title="Переименовать">✏️</span>
                <span class="add-subfolder" title="Создать подпапку">➕</span>
                <span class="delete-folder" title="Удалить папку">🗑️</span>
            `;
            actions.querySelector('.rename-folder').addEventListener('click', (e) => {
        e.stopPropagation();
        const newName = prompt('Введите новое название папки:', folder.name);
        if (newName && newName.trim() !== '' && newName.trim() !== folder.name) {
            renameFolder(folder.id, newName.trim());
        }
    });
            actions.querySelector('.add-subfolder')?.addEventListener('click', (e) => {
                e.stopPropagation();
                const name = prompt('Введите название подпапки:');
                if (name) createSubfolder(folder.id, name);
            });

            actions.querySelector('.delete-folder').addEventListener('click', (e) => {
                e.stopPropagation();
                if (confirm('Удалить папку?')) deleteFolder(folder.id);
            });

            div.appendChild(toggle);
            div.appendChild(nameSpan);
            div.appendChild(actions);
            li.appendChild(div);

           if (folder.children && folder.children.length) {
            const ul = document.createElement('ul');
            ul.className = 'children';
            folder.children.forEach(child => buildFolderItem(child, ul));
            li.appendChild(ul);
            toggle.textContent = '▶';
            if (expandedFolders.has(folder.id.toString())) {
            li.classList.add('expanded');
        }
            toggle.addEventListener('click', (e) => {
                e.stopPropagation();
                // ИДЕАЛЬНОЕ РЕШЕНИЕ: Просто вешаем класс, а CSS сам всё плавно откроет и повернет!
                li.classList.toggle('expanded');
                // СОХРАНЕНИЕ СОСТОЯНИЯ В ПАМЯТЬ
            if (li.classList.contains('expanded')) {
                expandedFolders.add(folder.id.toString());
            } else {
                expandedFolders.delete(folder.id.toString());
            }
            localStorage.setItem('expandedFolders', JSON.stringify([...expandedFolders]));
            });
        } else {
            toggle.textContent = '▶';
            toggle.style.visibility = 'hidden';
        }
        parentElement.appendChild(li);
        }

        // --- ЛОГИКА КАТАЛОГА В ПАПКАХ ---
     

        let fcCurrentProducts = [];

        async function loadFolderCatalog(folderId) {
            const activeLi = document.querySelector(`.folder-tree li[data-id="${folderId}"] .folder-name`);
            const titleEl = document.getElementById('fc-title');
            if (titleEl) titleEl.innerText = activeLi ? `📂 Папка: ${activeLi.innerText}` : '📂 Каталог';

            fcCurrentProducts = allProducts.filter(p => p.folder_id === folderId);

            const linkBlock = document.getElementById('fc-link-block');
            const productsList = document.getElementById('fc-products-list');
            const reportsBlock = document.getElementById('fc-reports-block');

            // Показываем только товары без папки
            const unlinked = allProducts.filter(p => !p.folder_id);
            const select = document.getElementById('fc-unlinked-products');
            if (select) {
                select.innerHTML = unlinked.map(p => `<option value="${p.id}">${p.name}</option>`).join('') || '<option value="">Нет свободных товаров</option>';
            }

            if (fcCurrentProducts.length === 0) {
                // 1. Прячем блок привязки навсегда
                if (linkBlock) linkBlock.style.display = 'none'; 
                
                // 2. Выводим понятную инструкцию
                if (productsList) {
                    productsList.innerHTML = `
                        <div style="color:#aaa; padding:30px 20px; text-align:center; background:#2a2a2a; border-radius:8px; border:1px dashed #3a3a3a;">
                            <div style="font-size: 40px; margin-bottom: 10px;">📭</div>
                            <div style="font-size: 16px; color: #fff; margin-bottom: 5px;">В этой папке пока нет товаров</div>
                            <div style="font-size: 13px;">Чтобы переместить сюда товар, перейдите в главную вкладку <b>«Товары»</b>,<br>нажмите <b>«📁 Настроить папку»</b> рядом с нужным товаром и выберите эту папку.</div>
                        </div>
                    `;
                }
                if (reportsBlock) reportsBlock.style.display = 'none';
            } else {
                if (linkBlock) linkBlock.style.display = 'none';
                
                if (productsList) {
                    productsList.innerHTML = fcCurrentProducts.map(p => {
                        
                        // АБСОЛЮТНО БЕЗОПАСНАЯ ОЧИСТКА СТРОК (БЕЗ РЕГУЛЯРНЫХ ВЫРАЖЕНИЙ И СЛЕШЕЙ)
                        const n1 = p.name ? String(p.name) : '';
                        const safeName = n1.split(String.fromCharCode(10)).join(' ').split(String.fromCharCode(13)).join('').split("'").join("&#39;").split('"').join("&quot;");
                        
                        const s1 = p.synonyms ? String(p.synonyms) : '';
                        const safeSyns = s1.split(String.fromCharCode(10)).join(' ').split(String.fromCharCode(13)).join('').split("'").join("&#39;").split('"').join("&quot;");
                        
                        return `
                        <div class="card" style="background:#2a2a2a; border:1px solid #3a3a3a; padding:16px; border-radius:8px; margin-bottom:10px; display:flex; justify-content:space-between; align-items:center; flex-wrap:wrap; gap:10px;">
                            <div>
                                <strong style="font-size:18px; color:#4a90e2;">${p.name}</strong>
                                <div style="color:#aaa; font-size:13px; margin-top:4px;">Синонимы: ${p.synonyms || 'нет'}</div>
                            </div>
                            
                        </div>`;
                    }).join('');
                }

                if (reportsBlock) reportsBlock.style.display = 'block';
                allReportsData = await apiFetch('/api/reports/confirmed');
                if (typeof renderFCReports === 'function') renderFCReports();
            }
        }

        async function linkProductToFolder() {
            const select = document.getElementById('fc-unlinked-products');
            const prodId = select.value;
            if (!prodId) return;
            await apiFetch(`/api/products/${prodId}/link_folder`, {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify({ folder_id: currentView })
            });
            await loadFolders();
            loadFolderCatalog(currentView);
        }

        async function unlinkProduct(prodId) {
            if (!confirm('Отвязать этот товар от текущей папки?')) return;
            await apiFetch(`/api/products/${prodId}/link_folder`, {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify({ folder_id: null })
            });
            await loadFolders();
            loadFolderCatalog(currentView);
        }

      function renderFCReports() {
    const tbody = document.getElementById('fc-reports-tbody');
    const sortVal = document.getElementById('fc-report-sort').value;
    const prodIds = fcCurrentProducts.map(p => p.id);
    
    let data = allReportsData.filter(r => prodIds.includes(r.product_id));

    data.sort((a, b) => {
        if (sortVal === 'price_asc') return (a.extracted_price || 0) - (b.extracted_price || 0);
        if (sortVal === 'price_desc') return (b.extracted_price || 0) - (a.extracted_price || 0);
        if (sortVal === 'seller_asc') return (`${a.sender_name} ${a.chat_title}`).localeCompare(`${b.sender_name} ${b.chat_title}`);
        if (sortVal === 'name_asc') return (a.product_name || '').localeCompare(b.product_name || '');
        return new Date(b.date) - new Date(a.date);
    });

    tbody.innerHTML = data.map(item => {
        const isActual = item.is_actual === 1;
        const statusBadge = isActual ? '<span style="background: rgba(39, 174, 96, 0.2); color: #2ecc71; padding: 2px 6px; border-radius: 4px; font-size: 11px;">Актуален</span>' : '<span style="background: rgba(231, 76, 60, 0.2); color: #e74c3c; padding: 2px 6px; border-radius: 4px; font-size: 11px;">Не актуален</span>';
        
        let senderName = item.sender_name || 'Неизвестно';
        if (senderName === 'None') senderName = 'Неизвестно';

        // --- УМНЫЙ РЕНДЕР ИСТОЧНИКОВ ---
        let originalTitle = escapeHtml(item.chat_title || 'Неизвестный чат');
        let chatDisplayHtml = `<span style="color: #888;">${originalTitle}</span>`;
        
        if (item.sheet_url) {
            chatDisplayHtml = `<a href="${item.sheet_url}" target="_blank" style="color: #2ecc71; text-decoration: none; border-bottom: 1px dashed #2ecc71; font-weight: bold;" title="Открыть Google Таблицу">🔗 ${originalTitle} ↗</a>`;
        }
        else if (item.chat_id && item.telegram_message_id && (item.type === 'channel' || item.type === 'group')) {
            let cleanChatId = String(item.chat_id).replace(/^-100|^ -/, '');
            chatDisplayHtml = `<a href="https://t.me/c/${cleanChatId}/${item.telegram_message_id}" target="_blank" style="color: #4a90e2; text-decoration: none; border-bottom: 1px dashed #4a90e2; font-weight: bold;" title="Открыть сообщение в Telegram">🔗 ${originalTitle} ↗</a>`;
        } 
        else if (item.chat_id && String(item.chat_id).startsWith('api_src_')) {
            chatDisplayHtml = `<span style="color: #f39c12; font-weight: bold;" title="Обновлено по API">🌐 ${originalTitle}</span>`;
        
        } else if (item.type === 'private' || !item.type || (item.type !== 'excel' && !item.type.startsWith('excel'))) {
            let botUsername = null;
            let possibleNames = [
                item.custom_name, 
                item.chat_title, 
                item.tc_chat_title, 
                item.sender_name
            ];
            
            // 1. Агрессивно ищем любую строку с собачкой @
            for (let name of possibleNames) {
                if (name && typeof name === 'string' && name.includes('@')) {
                    let match = name.match(/@([a-zA-Z0-9_]+)/);
                    if (match) {
                        botUsername = match[1];
                        break;
                    }
                }
            }

            // 2. Если собачки нигде нет, но имя состоит из одного слова и заканчивается на "bot"
            if (!botUsername) {
                for (let name of possibleNames) {
                    if (name && typeof name === 'string') {
                        let cleanName = name.trim();
                        if (cleanName.toLowerCase().endsWith('bot') && !cleanName.includes(' ')) {
                            botUsername = cleanName;
                            break;
                        }
                    }
                }
            }

            if (botUsername) {
                // Делаем красивую веб-ссылку t.me
                let cleanUsername = botUsername.replace('@', '');
                chatDisplayHtml = `<a href="https://t.me/${cleanUsername}" target="_blank" title="Открыть бота" style="color: #4a90e2; text-decoration: none; border-bottom: 1px dashed #4a90e2; font-weight: bold;">🔗 ${originalTitle} ↗</a>`;
            } else {
                // Если юзернейма нет, открываем по ID через resolve
                let cleanUserId = item.chat_id ? String(item.chat_id).replace(/^-/, '') : '';
                chatDisplayHtml = `<a href="tg://resolve?domain=${cleanUserId}" title="Открыть чат" style="color: #4a90e2; text-decoration: none; border-bottom: 1px dashed #4a90e2; font-weight: bold;">🔗 ${originalTitle} ↗</a>`;
            }
        }
        else if (item.type && item.type.startsWith('excel')) {
            chatDisplayHtml = `<span style="color: #2ecc71; font-weight: bold;" title="Локальный файл">🔗 ${originalTitle}</span>`;
        }

        return `
        <tr style="opacity: ${isActual ? '1' : '0.5'};">
            <td style="color: #888; font-weight: bold;">#${item.binding_id}</td>
            <td><strong>${item.product_name}</strong> <br>${statusBadge}</td>
            <td style="color: #27ae60; font-weight: bold; font-size:15px;">
                ${item.extracted_price ? item.extracted_price.toLocaleString('ru-RU') : '—'}
            </td>
            <td>${senderName === 'Неизвестно' ? '<span style="color:#888;">—</span>' : escapeHtml(senderName)}</td>
            <td>${chatDisplayHtml}</td>
            <td style="white-space: pre-wrap; font-size: 12px; max-width: 300px;">${item.text}</td>
            <td style="font-size: 12px; color: #aaa;">${item.date}</td>
        </tr>
        `;
    }).join('') || '<tr><td colspan="7" class="no-messages">Пока нет подтвержденных цен для этого товара</td></tr>';
}

        // Массивы для хранения активных фильтров
let activeChatFilters = [];

// Обработка ввода для тегов чатов
document.getElementById('filter-chat-input')?.addEventListener('keypress', function(e) {
    if (e.key === 'Enter' && this.value.trim()) {
        const val = this.value.trim();
        if (!activeChatFilters.includes(val)) {
            activeChatFilters.push(val);
            renderFilterTags('chat-tags', activeChatFilters);
        }
        this.value = '';
    }
});

async function editTrackedChat(id, currentName) {
    const newName = prompt('Введите новое название для чата:', currentName);
    if (newName === null) return; // пользователь нажал "Отмена"
    await apiFetch(`/api/tracked_chats/${id}`, {
        method: 'PUT',
        headers: { 'Content-Type': 'application/json' },
        body: JSON.stringify({ custom_name: newName })
    });
    loadTrackedChats(); // обновляем список
}
// Универсальная функция отрисовки тегов (если у вас её еще нет)
function renderFilterTags(containerId, array) {
    const container = document.getElementById(containerId);
    container.innerHTML = array.map((tag, index) => `
        <span class="tag">${tag} <span onclick="removeFilterTag('${containerId}', ${index})" style="cursor:pointer; margin-left:5px;">×</span></span>
    `).join('');
}

// Удаление тега
window.removeFilterTag = function(containerId, index) {
    if (containerId === 'chat-tags') activeChatFilters.splice(index, 1);
    // Добавьте сюда логику для ваших старых массивов (search и exclude)
    renderFilterTags(containerId, containerId === 'chat-tags' ? activeChatFilters : []);
};



// Функция для безопасного вывода текста
function escapeHtml(unsafe) {
    if (!unsafe) return '';
    return String(unsafe)
         .replace(/&/g, "&amp;")
         .replace(/</g, "&lt;")
         .replace(/>/g, "&gt;")
         .replace(/"/g, "&quot;")
         .replace(/'/g, "&#039;");
}
const socket = io();
// --- ВОССТАНАВЛИВАЕМ ФУНКЦИЮ ---
function startPolling() { 
    if (typeof currentView !== 'undefined') { 
        if (currentView === 'inbox' && typeof loadMessages === 'function') {
            loadMessages(typeof currentPage !== 'undefined' ? currentPage : 1);
        } else if (typeof loadFolderMessages === 'function') {
            loadFolderMessages(currentView); 
        }
    } 
}
        // -------------------------------
    // Слушаем событие 'db_updated' от сервера


    // ОБЯЗАТЕЛЬНО добавляем эту переменную ПЕРЕД сокетом
let dbUpdateTimeout = null; 
socket.on('db_updated', function(data) { 
    clearTimeout(dbUpdateTimeout); 
    dbUpdateTimeout = setTimeout(() => { 
        window.apiCache = {}; 
        
        if (typeof currentView !== 'undefined') { 
            if (document.getElementById('parser-section').style.display === 'block') { 
                // --- ИСПРАВЛЕНО: Передаем currentPage, чтобы оставаться на текущей странице ---
                if (currentView === 'inbox' && typeof loadMessages === 'function') {
                    loadMessages(typeof currentPage !== 'undefined' ? currentPage : 1);
                } else if (typeof loadFolderMessages === 'function') {
                    loadFolderMessages(currentView);
                }
            } 
        }
        
        if (document.getElementById('excel-section').style.display === 'block' && typeof loadExcelConfigs === "function") { 
            loadExcelConfigs(); 
        } 
        
        // ИСПРАВЛЕНА ПРОВЕРКА: проверяем, что открыта именно вкладка с юзерботами (а не просто существует элемент)
        if (document.getElementById('userbots-section').style.display === 'block' && typeof loadTrackedChats === "function") { 
            loadTrackedChats(); 
        } 
        
        if (document.getElementById('products-section').style.display === 'block' && typeof loadProducts === "function") { 
            loadProducts(); 
            if (typeof refreshProductBinding === 'function' && window.currentBindingProductId) { refreshProductBinding(); } 
        } 
        
        if (document.getElementById('folder-catalog-section').style.display === 'block' && typeof loadFolderCatalog === "function") { 
            loadFolderCatalog(currentView); 
        } 
    }, 500); // 500 мс задержки
});

async function loadTrackedChats() {
    const chats = await cachedApiFetch('/api/tracked_chats');
    const list = document.getElementById('tracked-chats-list');
    if (!list) return;

    if (!chats.length) {
        list.innerHTML = '<li>Нет отслеживаемых чатов.</li>';
        return;
    }
    list.innerHTML = '';
    chats.forEach(chat => {
        const li = document.createElement('li');
        li.style.margin = '8px 0';
        
        // Формируем отображаемое имя
        let displayHtml = '';
        const original = chat.chat_title || ('ID: ' + chat.chat_id);
        const custom = chat.custom_name;
        if (custom && custom !== original) {
            displayHtml = `<strong style="color:#4a90e2;">${escapeHtml(custom)}</strong> <span style="color:#888; font-size:12px;">(${escapeHtml(original)})</span>`;
        } else {
            displayHtml = `<span>${escapeHtml(original)}</span>`;
        }
        
        // Значение для редактирования: текущий custom_name (может быть null/undefined)
        const currentEditName = custom || '';
        
        li.innerHTML = `
            ${displayHtml}
            <button class="edit-chat" data-id="${chat.id}" data-name="${escapeHtml(currentEditName)}" style="margin-left:10px; background:none; border:none; cursor:pointer;" title="Редактировать название">✏️</button>
            <button class="delete-chat" data-id="${chat.id}" style="margin-left:5px; background:none; border:none; cursor:pointer;" title="Удалить">❌</button>
        `;
        list.appendChild(li);
    });

    // Обработчики для кнопок редактирования
    document.querySelectorAll('.edit-chat').forEach(btn => {
        btn.addEventListener('click', async (e) => {
            e.stopPropagation();
            const id = btn.dataset.id;
            const currentName = btn.dataset.name;
            await editTrackedChat(id, currentName);
        });
    });

    // Обработчики для кнопок удаления
   // Обработчики для кнопок удаления
    document.querySelectorAll('.delete-chat').forEach(btn => {
        btn.addEventListener('click', async () => {
            const id = btn.dataset.id;
            if (confirm('Удалить чат из отслеживаемых?')) {
                await apiFetch(`/api/tracked_chats/${id}`, { method: 'DELETE' });
                loadTrackedChats();
            }
        });
    });
    
    }

    function restoreUIState() {
    // 1. Восстанавливаем текст в поле поиска
    const savedSearchText = localStorage.getItem('filterSearchText');
    const searchInput = document.getElementById('filter-search-input');
    if (savedSearchText && searchInput) {
        searchInput.value = savedSearchText;
    }

    // 2. Восстанавливаем теги поиска
    const savedSearchWords = localStorage.getItem('searchWords');
    if (savedSearchWords) {
        searchWords = JSON.parse(savedSearchWords);
        updateSearchTags();
    }

    // 3. Восстанавливаем слова-исключения
    const savedExcludeWords = localStorage.getItem('excludeWords');
    if (savedExcludeWords) {
        excludeWords = JSON.parse(savedExcludeWords);
        updateExcludeTags();
    }

    // 4. Восстанавливаем активную вкладку меню
    const savedSection = localStorage.getItem('activeSection') || 'parser';
    const menuItem = document.querySelector(`.menu-items li[data-section="${savedSection}"]`);
    if (menuItem) {
        document.querySelectorAll('.menu-items li').forEach(i => i.classList.remove('active'));
        menuItem.classList.add('active');
        showSection(savedSection);
    } else {
        showSection('parser'); // По умолчанию
    }

    // 5. Восстанавливаем открытую папку
    const savedFolder = localStorage.getItem('activeFolder');
    if (savedFolder) {
        // Если это число (ID папки), преобразуем обратно в число
        currentView = isNaN(savedFolder) || savedFolder === 'inbox' ? savedFolder : parseInt(savedFolder);
        // Интерфейс дерева папок подтянется сам после вызова loadFolders()
    }
}



// ================= ЛОГИКА API И КЛИЕНТОВ =================
async function loadApiClients() {
    const clients = await apiFetch('/api/api_clients');
    const pubEditor = document.getElementById('publish-editor-container');
if (pubEditor) pubEditor.style.display = 'none';
    const tbody = document.getElementById('api-clients-tbody');
    tbody.innerHTML = clients.map(c => `
        <tr style="cursor:pointer; transition: background 0.2s;" onmouseover="this.style.background='#3a3a3a'" onmouseout="this.style.background='transparent'" onclick="openMarkups(${c.id}, '${escapeHtml(c.name)}', '${c.allowed_folders || 'all'}', '${c.allowed_chats || 'all'}')">
            <td><strong>${escapeHtml(c.name)}</strong></td>
            <td><code style="background:#1e1e1e; padding:2px 6px; border-radius:4px; font-size:11px; color:#4a90e2;">${c.token}</code></td>
            <td onclick="event.stopPropagation()">
                <div style="display:flex; gap:5px; align-items:center; font-size: 13px;">
                    <input type="checkbox" id="api_sch_en_${c.id}" ${c.schedule_enabled ? 'checked' : ''} title="Включить расписание">
                    <input type="time" id="api_sch_start_${c.id}" value="${c.time_start || '00:00'}" style="background:#1e1e1e; color:#fff; border:1px solid #3a3a3a; padding:4px; border-radius:4px;">
                    - 
                    <input type="time" id="api_sch_end_${c.id}" value="${c.time_end || '23:59'}" style="background:#1e1e1e; color:#fff; border:1px solid #3a3a3a; padding:4px; border-radius:4px;">
                    <button class="save-btn" onclick="saveApiClientSchedule(${c.id})" style="padding:4px 8px; font-size:14px; margin-left:5px;" title="Сохранить тайминг">💾</button>
                </div>
            </td>
            <td>
                <button class="save-btn" style="background:#e74c3c; font-size:12px; padding:4px 8px;" onclick="event.stopPropagation(); deleteApiClient(${c.id})">❌</button>
            </td>
        </tr>
    `).join('') || '<tr><td colspan="4" style="text-align:center; color:#aaa;">Нет созданных клиентов</td></tr>';
}

async function createApiClient() {
    const name = document.getElementById('new-api-client-name').value.trim();
    if(!name) return;
    await apiFetch('/api/api_clients', { 
        method: 'POST', 
        headers: { 'Content-Type': 'application/json' }, 
        body: JSON.stringify({name: name, schedule_enabled: 0, time_start: '00:00', time_end: '23:59'}) 
    });
    document.getElementById('new-api-client-name').value = '';
    loadApiClients();
}
window.saveApiClientSchedule = async function(clientId) {
    const enabled = document.getElementById(`api_sch_en_${clientId}`).checked ? 1 : 0;
    const start = document.getElementById(`api_sch_start_${clientId}`).value;
    const end = document.getElementById(`api_sch_end_${clientId}`).value;
    
    await apiFetch(`/api/api_clients/${clientId}/schedule`, {
        method: 'POST',
        headers: { 'Content-Type': 'application/json' },
        body: JSON.stringify({ schedule_enabled: enabled, time_start: start, time_end: end })
    });
    alert('✅ Расписание для клиента успешно сохранено!');
    loadApiClients();
};
async function deleteApiClient(id) {
    if(!confirm('Удалить клиента и все его настройки наценок?')) return;
    
    await apiFetch(`/api/api_clients/${id}`, { method: 'DELETE' });
    document.getElementById('api-filters-container').style.display = 'none';
    document.getElementById('api-markups-container').style.display = 'none';
    loadApiClients();
}

// Восстанавливаем оригинальный вызов без лишних переменных
// В loadApiClients() должно быть: onclick="openMarkups(${c.id}, '${escapeHtml(c.name)}')"



// Функция загрузки настроек публикации для выбранного клиента
async function loadPublishSettings(pubId) {
    // 1. Пытаемся найти контейнер редактора. Если его нет — выходим, чтобы не было ошибок в консоли.
    const editorContainer = document.getElementById('publish-editor-container');
    if (!editorContainer) return;

    // 2. Получаем данные конкретной публикации из нашего нового API
    const publications = await cachedApiFetch('/api/publications');
    const pub = publications.find(p => p.id == pubId);
    if (!pub) return;

    // 3. Заполняем основные поля (используем новые ID с префиксом pub_)
    document.getElementById('pub_id').value = pub.id;
    document.getElementById('pub_name').value = pub.name || '';
    document.getElementById('pub_enabled').checked = pub.is_active === 1;
    document.getElementById('pub_interval').value = pub.interval_min || 60;
    document.getElementById('pub_chat_id').value = pub.chat_id || '';
    document.getElementById('pub_message_id').value = pub.message_id || '';
    document.getElementById('pub_template').value = pub.template || '';

    // 4. Заполня email список юзерботов
    const bots = await cachedApiFetch('/api/userbots');
    const botSelect = document.getElementById('pub_userbot_id');
    if (botSelect) {
        botSelect.innerHTML = '<option value="">-- Выберите бота --</option>' + 
            bots.filter(b => b.status === 'active')
                .map(b => `<option value="${b.id}" ${pub.userbot_id == b.id ? 'selected' : ''}>
                    ${escapeHtml(b.account_name || 'ID: ' + b.api_id)}
                </option>`).join('');
    }

    // 5. Рендерим дерево товаров с галочками поставщиков
    // Сначала парсим сохраненные настройки товаров
    let allowedItems = {};
    try { 
        allowedItems = typeof pub.allowed_items === 'string' ? JSON.parse(pub.allowed_items) : pub.allowed_items || {}; 
    } catch(e) { 
        console.error("Ошибка парсинга allowed_items:", e); 
    }

    // Загружаем актуальное дерево товаров и поставщиков
    const treeData = await cachedApiFetch('/api/publish_tree_data');
    
    // Вызываем функцию отрисовки дерева, которую мы написали ранее
    if (typeof renderPublishTree === 'function') {
        renderPublishTree(treeData.folders, treeData.products, treeData.suppliers, allowedItems);
    }

    // Показываем редактор
    editorContainer.style.display = 'block';
}

// Сохранение настроек публикации
async function savePublishSettings(clientId) {
    const enabled = document.getElementById('publish_enabled').checked ? 1 : 0;
    const chatId = document.getElementById('publish_chat_id').value.trim();
    const messageId = document.getElementById('publish_message_id').value.trim();
    const userbotId = document.getElementById('publish_userbot_id').value;
    const template = document.getElementById('publish_template').value;
    const interval = document.getElementById('publish_interval').value;
    // Собрать выбранные папки
    const folderSelect = document.getElementById('publish_folders');
    const selectedFolders = Array.from(folderSelect.selectedOptions).map(opt => opt.value);
    const foldersJson = JSON.stringify(selectedFolders);
    
    await apiFetch(`/api/api_clients/${clientId}/publish`, {
        method: 'POST',
        headers: { 'Content-Type': 'application/json' },
        body: JSON.stringify({
            publish_enabled: enabled,
            publish_interval: parseInt(interval) || 60,
            publish_chat_id: chatId || null,
            publish_message_id: messageId || null,
            userbot_id: userbotId || null,
            publish_template: template,
            folders: foldersJson
        })
    });
    alert('✅ Настройки публикации сохранены!');
}



async function openMarkups(clientId, clientName) {
    // 1. Показываем наценки (оригинальный код без изменений)
    document.getElementById('api-markups-container').style.display = 'block';
    document.getElementById('markup-client-title').innerHTML = `Настройка наценок: <span style="color:#fff;">${clientName}</span>`;
    document.getElementById('markup-client-id').value = clientId;
    

    const folders = await cachedApiFetch('/api/folders/tree');
    const select = document.getElementById('markup-folder');
    select.innerHTML = '<option value="0">Базовая (Для остальных)</option>';
    function addOpt(list, level=0) {
        list.forEach(f => {
            if(f.name === 'По умолчанию') return;
            select.innerHTML += `<option value="${f.id}">${'—'.repeat(level)} ${f.name}</option>`;
            if(f.children) addOpt(f.children, level+1);
        });
    }
    addOpt(folders);
    loadMarkups(clientId);

    // 2. ЗАГРУЖАЕМ И РИСУЕМ ДЕРЕВО ДОСТУПОВ
    document.getElementById('api-filters-container').style.display = 'block';
    document.getElementById('filters-client-title').innerHTML = `Доступ: <span style="color:#fff;">${clientName}</span>`;
    const treeContainer = document.getElementById('filter-access-tree');
    treeContainer.innerHTML = '<div style="text-align:center; padding:20px;">Загрузка структуры... ⏳</div>';
    
    // Получаем сырые настройки клиента и всё дерево
   // Получаем сырые настройки клиента и всё дерево
    const clients = await cachedApiFetch('/api/api_clients', true);
    const client = clients.find(c => c.id === clientId);
    let rules = {};
    try { rules = JSON.parse(client.access_rules || '{}'); } catch(e){}
    
    const treeData = await cachedApiFetch('/api/access_tree');

    // 1. Создаем карту всех папок (свойство children для вложенности)
    const foldersMap = { 0: {id: 0, name: '[Без папки / Общие]', parent_id: null, children: [], products: []} };
    treeData.folders.forEach(f => {
        foldersMap[f.id] = {...f, children: [], products: []};
    });

    // 2. Распределяем товары по папкам
    treeData.products.forEach(p => {
        const fid = p.folder_id || 0;
        if(foldersMap[fid]) foldersMap[fid].products.push(p);
    });

    // 3. Собираем реальное дерево (вкладываем дочерние папки в родительские)
    const rootFolders = [];
    Object.values(foldersMap).forEach(f => {
        if (f.id !== 0 && f.parent_id && foldersMap[f.parent_id]) {
            foldersMap[f.parent_id].children.push(f);
        } else {
            rootFolders.push(f); // Корневые папки
        }
    });

    // 4. Сортируем папки и товары по алфавиту
    function sortFolderTree(folder) {
        folder.children.sort((a, b) => a.name.localeCompare(b.name));
        folder.products.sort((a, b) => a.name.localeCompare(b.name));
        folder.children.forEach(sortFolderTree);
    }
    rootFolders.forEach(sortFolderTree);

    // 5. РЕКУРСИВНАЯ ФУНКЦИЯ: Отрисовывает папки внутри папок
    function buildFolderHtml(folder) {
        // Проверяем, есть ли вообще товары в этой ветке (если пусто — скрываем)
        function hasProducts(f) {
            if (f.products.length > 0) return true;
            return f.children.some(hasProducts);
        }
        if (!hasProducts(folder)) return '';

        const folderContentId = 'api-folder-content-' + folder.id;

        // Открываем папку по умолчанию, если внутри уже есть выданные доступы
        function isAnyProductSelected(f) {
            if (f.products.some(prod => rules[prod.id] !== undefined)) return true;
            return f.children.some(isAnyProductSelected);
        }

        const hasSelectedProducts = isAnyProductSelected(folder);
        const initialDisplay = hasSelectedProducts ? 'block' : 'none';
        const initialIcon = hasSelectedProducts ? '▼' : '▶';

        // HTML самой папки
        let html = `<li style="margin-bottom:8px; background:rgba(255,255,255,0.03); padding:8px; border-radius:6px; border:1px solid #333;">
            <div style="display:flex; align-items:center; gap:8px;">
                <span onclick="toggleApiVisibility('${folderContentId}', this)" style="cursor:pointer; color:#aaa; font-size:12px; width:20px; text-align:center; display:inline-block; user-select:none;">${initialIcon}</span>
                <label style="font-weight:bold; color:#f39c12; display:flex; align-items:center; gap:8px; cursor:pointer; margin:0; flex:1;">
                    <input type="checkbox" class="cb-folder" onchange="toggleAccessLevel(this, '.cb-prod', '.cb-chat')">
                    📁 ${escapeHtml(folder.name)}
                </label>
            </div>
            <ul id="${folderContentId}" style="display:${initialDisplay}; list-style:none; padding-left:25px; margin-top:8px; border-left:1px solid #444;">`;

        // СНАЧАЛА рисуем вложенные папки (рекурсивно)
        folder.children.forEach(child => {
            html += buildFolderHtml(child);
        });

        // ЗАТЕМ рисуем товары, которые лежат конкретно в этой папке
        folder.products.forEach(prod => {
            const isProdAllowed = rules[prod.id] !== undefined;
            const allowedChats = rules[prod.id] || [];

            html += `<li style="margin-bottom:8px; background:#1e1e1e; padding:8px; border-radius:6px; border:1px solid #2a2a2a;" class="prod-group" data-pid="${prod.id}">
                <label style="color:#4a90e2; display:flex; gap:8px; border-bottom:1px solid #333; padding-bottom:4px; cursor:pointer;">
                    <input type="checkbox" class="cb-prod" ${isProdAllowed ? 'checked' : ''} onchange="toggleAccessLevel(this, '.cb-chat', null)">
                    📦 ${escapeHtml(prod.name)}
                </label>
                <ul style="list-style:none; padding-left:25px; margin-top:6px;">`;
            
            if(prod.chats.length === 0) {
                html += `<li style="color:#777; font-size:12px; font-style:italic;">Цены только ручные (нет поставщиков)</li>`;
            } else {
                prod.chats.forEach(chat => {
                    const isChatAllowed = isProdAllowed && (allowedChats.includes('all') || allowedChats.includes(String(chat.chat_id)));
                    html += `<li>
                        <label style="color:#ccc; display:flex; gap:8px; font-size:13px; cursor:pointer; padding:3px 0;">
                            <input type="checkbox" class="cb-chat" value="${chat.chat_id}" ${isChatAllowed ? 'checked' : ''} onchange="updateParentChecks(this)">
                            💬 ${escapeHtml(chat.name)}
                        </label>
                    </li>`;
                });
            }
            html += `</ul></li>`;
        });

        html += `</ul></li>`;
        return html;
    }

    // 6. Запускаем сборку HTML с самых верхних (корневых) папок
    let finalHtml = '<ul style="list-style:none; padding:0; margin:0;">';
    rootFolders.forEach(root => {
        finalHtml += buildFolderHtml(root);
    });
    finalHtml += '</ul>';

    treeContainer.innerHTML = finalHtml;
    
    // Синхронизируем состояние родительских галочек
    treeContainer.querySelectorAll('.cb-prod').forEach(cb => updateParentChecks(cb));

    await loadPublishSettings(clientId);
    
    // Привязать событие сохранения
  
}


// Функции для каскадного выделения (нажатие на папку выделяет товары, на товар - чаты)
window.toggleAccessLevel = function(el, childClass1, childClass2) {
    const li = el.closest('li');
    li.querySelectorAll(childClass1).forEach(cb => { cb.checked = el.checked; });
    if(childClass2) li.querySelectorAll(childClass2).forEach(cb => { cb.checked = el.checked; });
    updateParentChecks(el);
};

// Функция для сворачивания/разворачивания папок в API
window.toggleApiVisibility = function(targetId, iconEl) {
    const target = document.getElementById(targetId);
    if (!target) return;
    if (target.style.display === 'none') {
        target.style.display = 'block';
        if (iconEl) iconEl.innerText = '▼';
    } else {
        target.style.display = 'none';
        if (iconEl) iconEl.innerText = '▶';
    }
};

window.updateParentChecks = function(el) {
    const ul = el.closest('ul');
    if(!ul) return;
    const parentLi = ul.closest('li');
    if(!parentLi) return;
    const parentCb = parentLi.querySelector('input[type="checkbox"]');
    if(!parentCb || parentCb === el) return;

    // Умный поиск чекбоксов: берем только те, которые лежат непосредственно в этой ветке, 
    // чтобы не путать папки и вложенные товары
    const siblings = Array.from(ul.querySelectorAll('input[type="checkbox"]')).filter(cb => cb.closest('ul') === ul);
    
    parentCb.checked = siblings.some(cb => cb.checked);
    updateParentChecks(parentCb); // Идем выше
};

// Функция сбора данных и сохранения
window.saveClientFilters = async function() {
    const clientId = document.getElementById('markup-client-id').value;
    const rules = {};
    
    document.querySelectorAll('.prod-group').forEach(pGroup => {
        const prodCb = pGroup.querySelector('.cb-prod');
        if (prodCb && prodCb.checked) {
            const pid = pGroup.dataset.pid;
            const chatCbs = Array.from(pGroup.querySelectorAll('.cb-chat'));
            
            if (chatCbs.length === 0) {
                rules[pid] = ['all']; // Нет чатов, но товар разрешен
            } else {
                const checkedChats = chatCbs.filter(cb => cb.checked).map(cb => cb.value);
                if (checkedChats.length === chatCbs.length) {
                    rules[pid] = ['all']; // Разрешены все
                } else if (checkedChats.length > 0) {
                    rules[pid] = checkedChats; // Разрешены конкретные
                } else {
                    rules[pid] = ['all']; // Товар отмечен, но чаты сняты. Отдаем всё.
                }
            }
        }
    });
    
    await apiFetch(`/api/api_clients/${clientId}/filters`, {
        method: 'POST',
        headers: {'Content-Type': 'application/json'},
        body: JSON.stringify({ access_rules: rules })
    });
    
    alert('✅ Дерево доступов успешно сохранено!');
    window.apiCache = {}; 
};




async function loadMarkups(clientId) {
    const markups = await apiFetch(`/api/api_clients/${clientId}/markups`);
    const tbody = document.getElementById('api-markups-tbody');
    tbody.innerHTML = markups.map(m => `
        <tr>
            <td><strong style="color:${m.folder_id === 0 ? '#f39c12' : '#fff'};">${m.folder_id === 0 ? 'Базовая (Для остальных)' : escapeHtml(m.folder_name)}</strong></td>
            <td>${m.markup_value > 0 ? '+' : ''}${m.markup_value}${m.markup_type === 'percent' ? '%' : ' руб'}</td>
            <td>До ${m.rounding}</td>
            <td>
                ${m.folder_id === 0 ? '<span style="color:#888; font-size:11px;">(Неудаляемая)</span>' : `<button class="save-btn" style="background:#e74c3c; font-size:12px; padding:4px 8px;" onclick="deleteMarkup(${m.id})">❌</button>`}
            </td>
        </tr>
    `).join('');
}

async function saveMarkup() {
    const clientId = document.getElementById('markup-client-id').value;
    const body = {
        folder_id: parseInt(document.getElementById('markup-folder').value),
        markup_type: document.getElementById('markup-type').value,
        markup_value: parseFloat(document.getElementById('markup-value').value),
        rounding: parseInt(document.getElementById('markup-rounding').value)
    };
    await apiFetch(`/api/api_clients/${clientId}/markups`, { method: 'POST', headers: { 'Content-Type': 'application/json' }, body: JSON.stringify(body) });
    loadMarkups(clientId);
}

async function deleteMarkup(id) {
    if(!confirm('Удалить эту наценку?')) return;
    const clientId = document.getElementById('markup-client-id').value;
    await apiFetch(`/api/api_markups/${id}`, { method: 'DELETE' });
    loadMarkups(clientId);
}

// ================= ЛОГИКА ПАРСИНГА EXCEL =================
let excelStep = 0;
let exCoords = { n1:null, p1:null, n2:null, sku:null };
let exChatId = null;

let excelSheetsData = {};
let exCurrentSheet = null;

function toggleExcelMode() {
    selectExcelSheet(exCurrentSheet); // Сбрасываем шаги при переключении режима
}

async function loadExcelChats() {
    const chats = await apiFetch('/api/tracked_chats');
    const select = document.getElementById('excel-chat-select');
    select.innerHTML = '<option value="">-- Выберите чат --</option>' + 
        chats.map(c => `<option value="${c.chat_id}">${c.custom_name || c.chat_title}</option>`).join('');
}
async function parseLatestExcel(chatId) {
    if (!confirm('Найти и спарсить последний Excel/PDF файл в этом чате?')) return;
    
    try {
        const response = await apiFetch(`/api/excel/parse_latest/${chatId}`, {
            method: 'POST'
        });
        
        if (response && response.success) {
            alert('Задача запущена! Парсер начал искать и обрабатывать последний файл. Результаты скоро появятся на вкладке "Склад".');
        } else if (response && response.error) {
            alert('Ошибка: ' + response.error);
        }
    } catch (e) {
        console.error('Ошибка ручного парсинга:', e);
    }
}
async function loadExcelConfigs() {
    // 1. Загружаем чаты и папки для выпадающих списков
    const chats = await cachedApiFetch('/api/tracked_chats');
    const folders = await cachedApiFetch('/api/folders/tree');

    const chatSelect = document.getElementById('excel-chat-select');
    const gsChatSelect = document.getElementById('gs-chat-select');
    const pdfChatSelect = document.getElementById('pdf-chat-select');

    // --- УМНОЕ ОБНОВЛЕНИЕ: ЗАПОМИНАЕМ ВЫБОР ---
    // Запоминаем, что именно выбрал пользователь до перерисовки
    const currentChat = chatSelect ? chatSelect.value : '';
    const currentGsChat = gsChatSelect ? gsChatSelect.value : '';
    const currentPdfChat = pdfChatSelect ? pdfChatSelect.value : '';

    const chatOpts = '<option value="">-- Выберите чат --</option>' + chats.map(c => `<option value="${c.chat_id}">${escapeHtml(c.custom_name || c.chat_title)}</option>`).join('');

    // Вставляем новые пункты и сразу возвращаем сохраненный выбор
    if (chatSelect) {
        chatSelect.innerHTML = chatOpts;
        if (currentChat) chatSelect.value = currentChat;
    }
    if (gsChatSelect) {
        gsChatSelect.innerHTML = chatOpts;
        if (currentGsChat) gsChatSelect.value = currentGsChat;
    }
    if (pdfChatSelect) {
        pdfChatSelect.innerHTML = chatOpts;
        if (currentPdfChat) pdfChatSelect.value = currentPdfChat;
    }
    // ------------------------------------------

    

    // 2. Рисуем саму таблицу конфигов
    const configs = await apiFetch('/api/excel/configs');
    const tbody = document.getElementById('excel-configs-tbody');
    if (!tbody) return;

    // Умная группировка конфигов по ссылке (URL)
    const groupedConfigs = [];
    const urlMap = {};

    configs.forEach(c => {
        if (c.source_type === 'google_sheet' && c.sheet_url) {
            if (!urlMap[c.sheet_url]) {
                urlMap[c.sheet_url] = {
                    ...c,
                    sheets_count: 1,
                    all_ids: [c.id],
                    all_names: [c.sheet_name]
                };
                groupedConfigs.push(urlMap[c.sheet_url]);
            } else {
                urlMap[c.sheet_url].sheets_count++;
                urlMap[c.sheet_url].all_ids.push(c.id);
                urlMap[c.sheet_url].all_names.push(c.sheet_name);
            }
        } else {
            // Обычные файлы (Excel/PDF) добавляем как есть без группировки
            groupedConfigs.push({
                ...c, 
                sheets_count: 1, 
                all_ids: [c.id], 
                all_names: [c.sheet_name || 'Без названия']
            });
        }
    });

    tbody.innerHTML = groupedConfigs.map(c => {
        // --- ОПРЕДЕЛЯЕМ ТИП ИСТОЧНИКА ---
        const isGoogle = c.source_type === 'google_sheet';
        const isPDF = c.source_type === 'pdf';
        
        let typeIcon = '';
        let titleText = '';
        let color = '';

        if (isGoogle) {
            typeIcon = '🔗 Google Таблица';
            titleText = `<a href="${c.sheet_url}" target="_blank" style="color:#4a90e2;">Открыть ссылку</a>`;
            color = '#2ecc71'; // Зеленый
        } else if (isPDF) {
            typeIcon = '📑 PDF Документ';
            titleText = '<span style="color:#888;">Файл на сервере</span>';
            color = '#e74c3c'; // Красный
        } else {
            typeIcon = '📊 Excel Файл';
            titleText = '<span style="color:#888;">Файл на сервере</span>';
            color = '#3498db'; // Темно-зеленый
        }
        // --------------------------------

        const intervalText = isGoogle ? `<br><small style="color:#aaa;">Парсинг раз в ${c.parse_interval} мин</small>` : '';
        
        // Красивое отображение количества листов (только для Google)
        const sheetsDisplay = isGoogle && c.sheets_count > 1 
            ? `<div style="margin-top: 4px; color: #f39c12; font-size: 12px; font-weight: bold;">Внутри ${c.sheets_count} лист(ов)</div>` 
            : `<div style="margin-top: 4px; color: #aaa; font-size: 12px;">Лист: ${c.all_names[0]}</div>`;

        // Упаковываем все ID, чтобы кнопка удаляла всё разом
        const idsJson = JSON.stringify(c.all_ids);

        return `<tr>
            <td>
                <strong style="color: ${color};">${typeIcon}</strong>
                ${intervalText}
            </td>
            <td>${escapeHtml(c.chat_title || c.chat_id)}</td>
            <td>Папка ID: ${c.folder_id || 'Нет'}</td>
            <td>Товар: <b>${c.name_col}</b><br>Цена: <b>${c.price_col}</b>${sheetsDisplay}</td>
            <td>${titleText}</td>
            <td>
            <button class="save-btn" onclick="parseLatestExcel('${c.chat_id}')" 
                style="background:#9b59b6; margin-bottom: 4px; margin-right: 4px;" title="Спарсить последний файл в этом чате">
            🔄 Спарсить последнее
        </button>
                <button class="save-btn" style="background:#e74c3c; padding:5px 10px;" onclick='deleteMultipleConfigs(${idsJson})'>Удалить</button>
            </td>
        </tr>`;
    }).join('');
}

// Глобальная функция для удаления сразу нескольких листов сгруппированной таблицы
window.deleteMultipleConfigs = async function(ids) {
    if(!confirm('Удалить эту таблицу и настройки всех её листов?')) return;
    for (let id of ids) {
        await apiFetch(`/api/excel/configs/${id}`, { method: 'DELETE' });
    }
    loadExcelConfigs();
};
async function ignoreMissingSheet(id) {
    await apiFetch(`/api/excel/missing_sheets/${id}`, { method: 'DELETE' });
    loadExcelConfigs();
}

// Новая функция для удаления всех настроек чата разом
async function deleteExcelChatConfigs(chatId) {
    if(!confirm('Удалить ВСЕ настройки листов для этого чата?')) return;
    await apiFetch(`/api/excel/configs/chat/${chatId}`, { method: 'DELETE' });
    loadExcelConfigs();
}

// НОВАЯ ФУНКЦИЯ ДЛЯ ОТПРАВКИ КОМАНДЫ НА ПАРСИНГ
async function parseLatestExcel(chatId) {
    if (!confirm('Найти и спарсить последний Excel/PDF файл в этом чате?')) return;
     
    try {
        const response = await apiFetch(`/api/excel/parse_latest/${chatId}`, {
            method: 'POST'
        });
        
        if (response && response.success) {
            alert('Задача запущена! Парсер начал искать и обрабатывать последний файл. Результаты скоро появятся на вкладке "Склад".');
        } else if (response && response.error) {
            alert('Ошибка: ' + response.error);
        }
    } catch (e) {
        console.error('Ошибка ручного парсинга:', e);
    }
}

async function deleteExcelConfig(id) {
    if(!confirm('Удалить настройку для этого листа?')) return;
    await apiFetch(`/api/excel/configs/${id}`, { method: 'DELETE' });
    loadExcelConfigs();
}

async function setupMissingSheet(chatId, sheetName) {
    alert('⏳ Скачиваю последний прайс-лист из чата напрямую через Telegram... Это займет пару секунд.');
    try {
        const resp = await fetch(`/api/excel/preview_latest/${chatId}`);
        const data = await resp.json();
        
        if (!data.success && data.error) {
            alert('❌ Ошибка: ' + data.error);
            return;
        }
        
        // Заполняем глобальные переменные для предпросмотра
        exChatId = chatId;
        excelSheetsData = data.sheets_data;
        const sheetNames = Object.keys(excelSheetsData);
        
        // Показываем карточку предпросмотра
        const previewCard = document.getElementById('excel-preview-card');
        previewCard.style.display = 'block';
        
        // Генерируем табы листов
        let tabsHtml = '<div style="margin-bottom: 15px; display:flex; gap:8px; flex-wrap:wrap; border-bottom: 1px solid #3a3a3a; padding-bottom: 10px;">';
        sheetNames.forEach(sheet => {
            tabsHtml += `<button class="save-btn sheet-tab-btn" data-sheet="${escapeHtml(sheet)}" style="background:#333; color:#ccc; border:none;" onclick="selectExcelSheet('${escapeHtml(sheet)}')">${escapeHtml(sheet)}</button>`;
        });
        tabsHtml += '</div><div id="excel-table-container" style="overflow-x:auto;"></div>';
        
        previewCard.innerHTML = `
            <h3 style="color:#fff; margin-bottom:15px;">Быстрая настройка листа</h3>
            ${tabsHtml}
            <label style="display: inline-flex; align-items: center; gap: 8px; margin-top: 10px; cursor: pointer; color: #2ecc71; font-weight: bold; background: rgba(46, 204, 113, 0.1); padding: 8px 12px; border-radius: 6px; border: 1px solid #2ecc71;">
                <input type="checkbox" id="excel-apply-all-sheets" style="width: 18px; height: 18px;"> 
                Применить это правило ко ВСЕМ листам (и текущим, и будущим)
            </label>
            <div id="excel-status" style="color:#f39c12; margin-top:15px; font-size:16px; font-weight:bold;"></div>
        `;
        
        // Плавно скроллим страницу к таблице
        previewCard.scrollIntoView({ behavior: 'smooth' });
        
        // Автоматически открываем именно тот лист, которого не хватает
        if (sheetNames.includes(sheetName)) {
            selectExcelSheet(sheetName);
        } else {
            alert(`⚠️ Лист "${sheetName}" не найден в самом последнем файле. Возможно, поставщик удалил его.`);
            // Открываем первый доступный
            if (sheetNames.length > 0) selectExcelSheet(sheetNames[0]);
        }
        
    } catch (e) {
        alert('❌ Ошибка сети при попытке скачать файл.');
    }
}

async function uploadExcelPreview(event) {
    // 1. ПРЕДОТВРАЩАЕМ ПЕРЕЗАГРУЗКУ СТРАНИЦЫ
    if (event) event.preventDefault();
    
    const fileInput = document.getElementById('excel-file');
    const chatSelect = document.getElementById('excel-chat-select');
    
    if (!fileInput.files[0] || !chatSelect.value) { 
        alert("Выберите чат и файл"); 
        return; 
    }
    
    exChatId = chatSelect.value;
    const formData = new FormData();
    formData.append('file', fileInput.files[0]);
    
    const previewCard = document.getElementById('excel-preview-card');
    previewCard.style.display = 'block';
    previewCard.innerHTML = '<h3 style="color:#f39c12;">⏳ Загрузка и анализ файла...</h3>';
    
    try {
        const resp = await fetch('/api/excel/preview', { method: 'POST', body: formData });
        const data = await resp.json();
        
        if(!data.success) { 
            alert(data.error); 
            previewCard.style.display = 'none';
            return; 
        }
        
        excelSheetsData = data.sheets_data;
        const sheetNames = Object.keys(excelSheetsData);
        
        // Рисуем кнопки-вкладки для каждого листа
        let tabsHtml = '<div style="margin-bottom: 15px; display:flex; gap:8px; flex-wrap:wrap; border-bottom: 1px solid #3a3a3a; padding-bottom: 10px;">';
        sheetNames.forEach(sheet => {
            // Важно: добавляем type="button", чтобы кнопка вклада не перегружала страницу
            tabsHtml += `<button type="button" class="save-btn sheet-tab-btn" data-sheet="${escapeHtml(sheet)}" style="background:#333; color:#ccc; border:none; cursor:pointer;" onclick="selectExcelSheet('${escapeHtml(sheet)}')">${escapeHtml(sheet)}</button>`;
        });
        tabsHtml += '</div><div id="excel-table-container" style="overflow-x:auto;"></div>';
        
        previewCard.innerHTML = `
            <h3 style="color:#fff; margin-bottom:15px;">Выберите лист для настройки</h3>
            <label style="display: flex; align-items: center; gap: 8px; margin-bottom: 15px; color: #f39c12; font-weight: bold; background: rgba(243, 156, 18, 0.1); padding: 8px 12px; border-radius: 6px; cursor: pointer; border: 1px solid #f39c12;">
                <input type="checkbox" id="excel-grouped-mode" style="width: 18px; height: 18px;" onchange="toggleExcelMode()"> 
                Сложный формат (Товар / Категория ➔ Артикул ➔ Цена)
            </label>
            ${tabsHtml}
            <label style="display: inline-flex; align-items: center; gap: 8px; margin-top: 10px; cursor: pointer; color: #2ecc71; font-weight: bold; background: rgba(46, 204, 113, 0.1); padding: 8px 12px; border-radius: 6px; border: 1px solid #2ecc71;">
                <input type="checkbox" id="excel-apply-all-sheets" style="width: 18px; height: 18px;"> 
                Применить это правило ко ВСЕМ листам (и текущим, и будущим)
            </label>
            <div id="excel-status" style="color:#f39c12; margin-top:15px; font-size:16px; font-weight:bold;"></div>
        `;
        
        if(sheetNames.length > 0) selectExcelSheet(sheetNames[0]);

    } catch (err) {
        console.error(err);
        alert("Ошибка при загрузке файла");
        previewCard.style.display = 'none';
    }
}

function selectExcelSheet(sheetName) {
    exCurrentSheet = sheetName;
    excelStep = 0;
    exCoords = { n1:null, p1:null, n2:null, sku:null };
    
    document.querySelectorAll('.sheet-tab-btn').forEach(btn => {
        if(btn.dataset.sheet === sheetName) {
            btn.style.background = '#4a90e2'; btn.style.color = '#fff';
        } else {
            btn.style.background = '#333'; btn.style.color = '#ccc';
        }
    });

    const isGroupedMode = document.getElementById('excel-grouped-mode')?.checked;

    const sheetData = excelSheetsData[sheetName];
    let html = '<table class="simple-table" style="background:#fff; color:#000; font-size:13px; cursor:crosshair; width: max-content; user-select: none;">';
    sheetData.forEach((row, rIdx) => {
        html += '<tr>';
        row.forEach((cell, cIdx) => {
            html += `<td style="border:1px solid #ccc; padding:6px; min-width: 80px;" onclick="onExcelCellClick(${rIdx}, ${cIdx}, '${escapeHtml(String(cell))}', this)">${escapeHtml(String(cell))}</td>`;
        });
        html += '</tr>';
    });
    html += '</table>';
    
    document.getElementById('excel-table-container').innerHTML = html;
    
    if (isGroupedMode) {
        document.getElementById('excel-status').innerText = `[${sheetName}] СЛОЖНЫЙ РЕЖИМ. Шаг 1: Кликните на ПЕРВУЮ КАТЕГОРИЮ (Например: iPhone 17 Pro Max 1Tb Silver)`;
    } else {
        document.getElementById('excel-status').innerText = `[${sheetName}] Шаг 1: Кликните на ячейку с ПЕРВЫМ НАЗВАНИЕМ товара`;
    }
}

async function onExcelCellClick(r, c, val, el) {
    el.style.background = '#f39c12';
    el.style.color = '#fff';
    el.style.fontWeight = 'bold';

    const isGroupedMode = document.getElementById('excel-grouped-mode')?.checked;

    if (isGroupedMode) {
        // ЛОГИКА ДЛЯ СЛОЖНОГО ФОРМАТА
        if (excelStep === 0) {
            exCoords.n1 = {r, c, val};
            excelStep++;
            document.getElementById('excel-status').innerText = `[${exCurrentSheet}] ✅ Группа: "${val}". Шаг 2: Кликните на АРТИКУЛ под ней.`;
        } else if (excelStep === 1) {
            exCoords.sku = {r, c, val};
            excelStep++;
            document.getElementById('excel-status').innerText = `[${exCurrentSheet}] ✅ Артикул: "${val}". Шаг 3: Кликните на ЦЕНУ для этого артикула.`;
        } else if (excelStep === 2) {
            exCoords.p1 = {r, c, val};
            await finishExcelSetup(exCoords.n1.c, exCoords.sku.c, exCoords.p1.c, exCoords.n1.r, 1);
        }
    } else {
        // ЛОГИКА ДЛЯ СТАНДАРТНОГО ФОРМАТА
        if (excelStep === 0) {
            exCoords.n1 = {r, c, val};
            excelStep++;
            document.getElementById('excel-status').innerText = `[${exCurrentSheet}] ✅ Название: "${val}". Шаг 2: Кликните на ячейку с ЦЕНОЙ.`;
        } else if (excelStep === 1) {
            exCoords.p1 = {r, c, val};
            excelStep++;
            document.getElementById('excel-status').innerText = `[${exCurrentSheet}] ✅ Цена: "${val}". Шаг 3: Кликните на ВТОРОЕ НАЗВАНИЕ (следующий товар).`;
        } else if (excelStep === 2) {
            exCoords.n2 = {r, c, val};
            const blockSize = exCoords.n2.r - exCoords.n1.r;
            if (blockSize <= 0 || exCoords.n1.c !== exCoords.n2.c) {
                alert("Ошибка! Второе название должно быть строго НИЖЕ первого.");
                selectExcelSheet(exCurrentSheet);
                return;
            }
            await finishExcelSetup(exCoords.n1.c, -1, exCoords.p1.c, exCoords.n1.r, 0, exCoords.p1.r - exCoords.n1.r, blockSize);
        }
    }
}

// Вынес сохранение в отдельную функцию, чтобы не дублировать код
async function finishExcelSetup(n_col, sku_col, p_col, start_row, is_grouped, p_off = 0, block_step = 1) {
    const applyAll = document.getElementById('excel-apply-all-sheets')?.checked;
    const targetSheetName = applyAll ? '*' : exCurrentSheet;

    const config = {
        chat_id: exChatId,
        sheet_name: targetSheetName,
        name_col: n_col,
        name_row_offset: 0,
        sku_col: sku_col,
        price_col: p_col,
        price_row_offset: p_off,
        block_step: block_step,
        start_row: start_row,
        is_grouped: is_grouped
    };

    await apiFetch('/api/excel/save_config', {
        method: 'POST', headers: {'Content-Type': 'application/json'}, body: JSON.stringify(config)
    });
    
    loadExcelConfigs(); 

    if (applyAll) {
        alert('🎉 Правило сохранено! Оно применено ко всем листам.');
        document.getElementById('excel-preview-card').style.display = 'none';
        return;
    }

    const sheetNames = Object.keys(excelSheetsData);
    const currentIndex = sheetNames.indexOf(exCurrentSheet);
    
    if (currentIndex < sheetNames.length - 1) {
        selectExcelSheet(sheetNames[currentIndex + 1]);
    } else {
        alert('🎉 Настройка для всех листов успешно сохранена!');
        document.getElementById('excel-preview-card').style.display = 'none';
    }
}



// --- ЛОГИКА ПРОФИЛЯ: Смена пароля ---
        function togglePasswordVisibility(inputId, iconElement) {
            const input = document.getElementById(inputId);
            if (input.type === 'password') {
                input.type = 'text';
                iconElement.innerText = '🙈'; // Иконка закрытого глаза
            } else {
                input.type = 'password';
                iconElement.innerText = '👁️'; // Иконка открытого глаза
            }
        }

        document.getElementById('change-password-form')?.addEventListener('submit', async (e) => {
            e.preventDefault();
            const oldPassword = document.getElementById('old-password').value;
            const newPassword = document.getElementById('new-password').value;
            const confirmPassword = document.getElementById('confirm-password').value;

            if (newPassword !== confirmPassword) {
                alert('Новые пароли не совпадают!');
                return;
            }

            try {
                const response = await fetch('/api/profile/change_password', {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({
                        old_password: oldPassword,
                        new_password: newPassword
                    })
                });

                const data = await response.json();
                if (data.success) {
                    alert('✅ Пароль успешно изменён!');
                    e.target.reset(); // Очищаем форму после успеха
                    
                    // Возвращаем все поля к скрытому виду
                    document.querySelectorAll('.toggle-password').forEach(icon => icon.innerText = '👁️');
                    document.getElementById('old-password').type = 'password';
                    document.getElementById('new-password').type = 'password';
                    document.getElementById('confirm-password').type = 'password';
                } else {
                    alert('❌ Ошибка: ' + data.error);
                }
            } catch (error) {
                alert('❌ Ошибка сети при смене пароля');
            }
        });



        async function logoutOtherDevices() {
            if (!confirm('Вы уверены, что хотите выйти со всех остальных устройств?')) return;
            
            try {
                const response = await fetch('/api/profile/logout_others', { method: 'POST' });
                const data = await response.json();
                
                if (data.success) {
                    alert('✅ Сеансы на других устройствах успешно завершены! Теперь доступ есть только с этого браузера.');
                } else {
                    alert('❌ Ошибка: ' + data.error);
                }
            } catch (error) {
                alert('❌ Ошибка сети при попытке завершить сеансы.');
            }
        }

        // --- НОВЫЙ БЛОК: Слушаем команду на принудительный выход ---
    socket.on('force_logout', async function() {
        try {
            // Делаем невидимый фоновый пинг на сервер
            const resp = await fetch('/api/check_session');
            
            // Если сервер ответил 401 (токен устарел) — мгновенно выкидываем на логин
            if (resp.status === 401) {
                window.location.href = '/login';
            }
        } catch (e) {
            console.error("Ошибка проверки сессии:", e);
        }
    });



    // Функция очистки сообщений из Админ-панели
        async function adminClearMessages() {
            const period = document.getElementById('clear-messages-period').value;
            let confirmMsg = 'Вы уверены, что хотите удалить старые сообщения? Это действие необратимо!';
            
            if (period === 'all') {
                confirmMsg = 'ВНИМАНИЕ! Вы собираетесь удалить ВООБЩЕ ВСЕ сообщения из базы. Все товары потеряют привязанные цены. Продолжить?';
            }
            
            if (!confirm(confirmMsg)) return;
            
            try {
                const response = await fetch('/api/admin/clear_messages', {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({ period: period })
                });
                const data = await response.json();
                
                if (data.success) {
                    alert('✅ База сообщений успешно очищена!');
                    // Сбрасываем кэш и перезагружаем таблицу, если открыт парсер
                    window.apiCache = {};
                    if (typeof currentView !== 'undefined' && currentView === 'inbox' && typeof loadMessages === 'function') {
                        loadMessages();
                    }
                } else {
                    alert('❌ Ошибка сервера: ' + data.error);
                }
            } catch (error) {
                alert('❌ Ошибка сети при отправке команды.');
            }
        }



// ================= ЛОГИКА ПАРСИНГА GOOGLE ТАБЛИЦ =================
let gsCurrentSheet = null;
let gsStep = 0;
let gsCoords = { n1:null, p1:null, n2:null, sku:null };
let gsSheetsData = {};

async function previewGoogleSheet() {
    const urlInput = document.getElementById('gs-url').value.trim();
    const chatSelect = document.getElementById('gs-chat-select').value;
    if (!urlInput || !chatSelect) { alert("Укажите ссылку на таблицу и выберите поставщика!"); return; }
    
    document.getElementById('gs-preview-card').style.display = 'block';
    document.getElementById('gs-preview-card').innerHTML = '<h3 style="color:#2ecc71;">⏳ Загрузка таблицы...</h3>';
    
    try {
        const resp = await fetch('/api/excel/google_sheet/preview', {
            method: 'POST',
            headers: {'Content-Type': 'application/json'},
            body: JSON.stringify({ sheet_url: urlInput })
        });
        const data = await resp.json();
        if(!data.success) { alert("❌ Ошибка: " + data.error); document.getElementById('gs-preview-card').style.display='none'; return; }
        
        gsSheetsData = data.sheets_data;
        const sheetNames = Object.keys(gsSheetsData);
        
        let tabsHtml = '<div style="margin-bottom: 15px; display:flex; gap:8px; flex-wrap:wrap; border-bottom: 1px solid #3a3a3a; padding-bottom: 10px;">';
        sheetNames.forEach(sheet => {
            tabsHtml += `<button class="save-btn gs-sheet-tab-btn" data-sheet="${escapeHtml(sheet)}" style="background:#333; color:#ccc; border:none;" onclick="selectGsSheet('${escapeHtml(sheet)}')">${escapeHtml(sheet)}</button>`;
        });
        tabsHtml += '</div><div id="gs-table-container" style="overflow-x:auto;"></div>';
        
        document.getElementById('gs-preview-card').innerHTML = `
            <h3 style="color:#fff; margin-bottom:15px;">Выберите лист Google Таблицы</h3>
            <label style="display: flex; align-items: center; gap: 8px; margin-bottom: 15px; color: #2ecc71; font-weight: bold; background: rgba(46, 204, 113, 0.1); padding: 8px 12px; border-radius: 6px; cursor: pointer; border: 1px solid #2ecc71;">
                <input type="checkbox" id="gs-grouped-mode" style="width: 18px; height: 18px;" onchange="toggleGsMode()"> 
                Сложный формат (Категория ➔ Артикул ➔ Цена)
            </label>
            ${tabsHtml}
            <label style="display: inline-flex; align-items: center; gap: 8px; margin-top: 10px; cursor: pointer; color: #f39c12; font-weight: bold; background: rgba(243, 156, 18, 0.1); padding: 8px 12px; border-radius: 6px; border: 1px solid #f39c12;">
                <input type="checkbox" id="gs-apply-all-sheets" style="width: 18px; height: 18px;"> 
                Применить правило ко ВСЕМ листам (текущим и будущим)
            </label>
            <div id="gs-status" style="color:#2ecc71; margin-top:15px; font-size:16px; font-weight:bold;"></div>
        `;
        if(sheetNames.length > 0) selectGsSheet(sheetNames[0]);
    } catch(e) {
        alert('Ошибка сети: ' + e);
        document.getElementById('gs-preview-card').style.display = 'none';
    }
}

function toggleGsMode() { selectGsSheet(gsCurrentSheet); }

function selectGsSheet(sheetName) {
    gsCurrentSheet = sheetName;
    gsStep = 0;
    gsCoords = { n1:null, p1:null, n2:null, sku:null };
    
    document.querySelectorAll('.gs-sheet-tab-btn').forEach(btn => {
        if(btn.dataset.sheet === sheetName) {
            btn.style.background = '#2ecc71'; btn.style.color = '#fff';
        } else {
            btn.style.background = '#333'; btn.style.color = '#ccc';
        }
    });

    const isGroupedMode = document.getElementById('gs-grouped-mode')?.checked;
    const sheetData = gsSheetsData[sheetName];
    
    let html = '<table class="simple-table" style="background:#fff; color:#000; font-size:13px; cursor:crosshair; width: max-content; user-select: none;">';
    sheetData.forEach((row, rIdx) => {
        html += '<tr>';
        row.forEach((cell, cIdx) => {
            html += `<td style="border:1px solid #ccc; padding:6px; min-width: 80px;" onclick="onGsCellClick(${rIdx}, ${cIdx}, '${escapeHtml(String(cell))}', this)">${escapeHtml(String(cell))}</td>`;
        });
        html += '</tr>';
    });
    html += '</table>';
    
    document.getElementById('gs-table-container').innerHTML = html;
    
    if (isGroupedMode) {
        document.getElementById('gs-status').innerText = `[${sheetName}] СЛОЖНЫЙ РЕЖИМ. Шаг 1: Кликните на ПЕРВУЮ КАТЕГОРИЮ.`;
    } else {
        document.getElementById('gs-status').innerText = `[${sheetName}] ПРОСТОЙ РЕЖИМ. Шаг 1: Кликните на ПЕРВОЕ НАЗВАНИЕ товара`;
    }
}

async function onGsCellClick(r, c, val, el) {
    el.style.background = '#2ecc71';
    el.style.color = '#fff';
    el.style.fontWeight = 'bold';
    
    const isGroupedMode = document.getElementById('gs-grouped-mode')?.checked;

    if (isGroupedMode) {
        if (gsStep === 0) {
            gsCoords.n1 = {r, c, val}; gsStep++;
            document.getElementById('gs-status').innerText = `[${gsCurrentSheet}] ✅ Группа: "${val}". Шаг 2: Кликните на АРТИКУЛ под ней.`;
        } else if (gsStep === 1) {
            gsCoords.sku = {r, c, val}; gsStep++;
            document.getElementById('gs-status').innerText = `[${gsCurrentSheet}] ✅ Артикул: "${val}". Шаг 3: Кликните на ЦЕНУ для этого артикула.`;
        } else if (gsStep === 2) {
            gsCoords.p1 = {r, c, val};
            await finishGsSetup(gsCoords.n1.c, gsCoords.sku.c, gsCoords.p1.c, gsCoords.n1.r, 1, 0, 1);
        }
    } else {
        if (gsStep === 0) {
            gsCoords.n1 = {r, c, val}; gsStep++;
            document.getElementById('gs-status').innerText = `[${gsCurrentSheet}] ✅ Название: "${val}". Шаг 2: Кликните на ячейку с ЦЕНОЙ.`;
        } else if (gsStep === 1) {
            gsCoords.p1 = {r, c, val}; gsStep++;
            document.getElementById('gs-status').innerText = `[${gsCurrentSheet}] ✅ Цена: "${val}". Шаг 3: Кликните на ВТОРОЕ НАЗВАНИЕ (следующий товар).`;
        } else if (gsStep === 2) {
            gsCoords.n2 = {r, c, val};
            const blockSize = gsCoords.n2.r - gsCoords.n1.r;
            if (blockSize <= 0 || gsCoords.n1.c !== gsCoords.n2.c) {
                alert("Ошибка! Второе название должно быть строго НИЖЕ первого в той же колонке.");
                selectGsSheet(gsCurrentSheet);
                return;
            }
            await finishGsSetup(gsCoords.n1.c, -1, gsCoords.p1.c, gsCoords.n1.r, 0, gsCoords.p1.r - gsCoords.n1.r, blockSize);
        }
    }
}

async function finishGsSetup(n_col, sku_col, p_col, start_row, is_grouped, p_off = 0, block_step = 1) {
    const chatId = document.getElementById('gs-chat-select').value;
    if (!chatId) {
        alert('Пожалуйста, обязательно выберите "Привязать к поставщику" перед сохранением!');
        return; // Блокируем сохранение
    }

    const applyAll = document.getElementById('gs-apply-all-sheets')?.checked;
    const targetSheetName = applyAll ? '*' : gsCurrentSheet;

    const config = {
        chat_id: chatId,
        sheet_name: targetSheetName,
        name_col: n_col,
        sku_col: sku_col,
        price_col: p_col,
        price_row_offset: p_off,
        block_step: block_step,
        start_row: start_row,
        is_grouped: is_grouped,
        sheet_url: document.getElementById('gs-url').value.trim(),
        parse_interval: document.getElementById('gs-interval').value || 60
    };

    await apiFetch('/api/excel/google_sheet/save_config', {
        method: 'POST', headers: {'Content-Type': 'application/json'}, body: JSON.stringify(config)
    });
    
    loadExcelConfigs(); // Обновляем список настроек
    
    if (applyAll) {
        alert('🎉 Правило для Google Таблицы сохранено! Оно применено КО ВСЕМ листам.');
        document.getElementById('gs-preview-card').style.display = 'none';
        return;
    }

    // Если "Применить ко всем" не стоит, автоматически переходим к следующему листу
    const sheetNames = Object.keys(gsSheetsData);
    const currentIndex = sheetNames.indexOf(gsCurrentSheet);
    if (currentIndex < sheetNames.length - 1) {
        selectGsSheet(sheetNames[currentIndex + 1]);
    } else {
        alert('🎉 Настройка для всех листов Google Таблицы успешно завершена!');
        document.getElementById('gs-preview-card').style.display = 'none';
    }
}

// ================= ЛОГИКА ПАРСИНГА PDF =================
let pdfStep = 0;
let pdfCoords = { n1:null, p1:null, n2:null, sku:null };
let pdfSheetsData = {};

async function uploadPdfPreview(event) {
    // 1. Предотвращаем перезагрузку страницы браузером
    if (event) event.preventDefault();

    const fileInput = document.getElementById('pdf-file');
    const chatSelect = document.getElementById('pdf-chat-select');
    
    // Проверка на заполнение полей
    if (!fileInput.files[0] || !chatSelect.value) { 
        alert("Выберите чат и загрузите PDF файл!"); 
        return; 
    }
    
    const previewCard = document.getElementById('pdf-preview-card');
    previewCard.style.display = 'block';
    // Показываем статус загрузки
    previewCard.innerHTML = '<h3 style="color:#e74c3c;">⏳ Разбор PDF документа... Пожалуйста, подождите.</h3>';
    
    const formData = new FormData();
    formData.append('file', fileInput.files[0]);
    formData.append('chat_id', chatSelect.value); // Передаем ID чата, если нужно на бэкенде
    
    try {
        const resp = await fetch('/api/pdf/preview', { 
            method: 'POST', 
            body: formData 
        });
        
        const data = await resp.json();
        
        if (!data.success) { 
            alert("❌ Ошибка: " + data.error); 
            previewCard.style.display = 'none';
            return; 
        }
        
        // Сохраняем данные во внешнюю переменную для функции отрисовки
        pdfSheetsData = data.sheets_data;
        const sheetName = Object.keys(pdfSheetsData)[0] || 'Страница 1';
        
        // Формируем интерфейс внутри карточки
        previewCard.innerHTML = `
            <h3 style="color:#fff; margin-bottom:15px;">Настройка ячеек PDF</h3>
            
            <label style="display: flex; align-items: center; gap: 8px; margin-bottom: 15px; color: #e74c3c; font-weight: bold; background: rgba(231, 76, 60, 0.1); padding: 8px 12px; border-radius: 6px; cursor: pointer; border: 1px solid #e74c3c; width: fit-content;">
                <input type="checkbox" id="pdf-grouped-mode" style="width: 18px; height: 18px;" onchange="renderPdfTable()"> 
                Сложный формат (Категория ➔ Артикул ➔ Цена)
            </label>
            
            <div id="pdf-status" style="color:#e74c3c; margin-bottom:15px; font-size:16px; font-weight:bold;">
                Шаг 1: Кликните на ПЕРВОЕ НАЗВАНИЕ товара
            </div>
            
            <div id="pdf-table-container" style="overflow:auto; max-height: 550px; border: 1px solid #333; border-radius: 4px;">
                </div>
        `;
        
        // Вызываем функцию отрисовки таблицы
        renderPdfTable(sheetName);

    } catch (e) {
        console.error("Ошибка при загрузке PDF:", e);
        alert('Ошибка сети при попытке обработать PDF.');
        previewCard.style.display = 'none';
    }
}

function renderPdfTable() {
    pdfStep = 0;
    pdfCoords = { n1:null, p1:null, n2:null, sku:null };
    const isGroupedMode = document.getElementById('pdf-grouped-mode')?.checked;
    
    document.getElementById('pdf-status').innerText = isGroupedMode 
        ? `СЛОЖНЫЙ РЕЖИМ. Шаг 1: Кликните на ПЕРВУЮ КАТЕГОРИЮ.`
        : `ПРОСТОЙ РЕЖИМ. Шаг 1: Кликните на ПЕРВОЕ НАЗВАНИЕ товара.`;

    let html = '<table class="simple-table" style="background:#fff; color:#000; font-size:13px; cursor:crosshair; width: max-content; user-select: none;">';
    pdfSheetsData['Страница 1'].forEach((row, rIdx) => {
        html += '<tr>';
        row.forEach((cell, cIdx) => {
            html += `<td style="border:1px solid #ccc; padding:6px; min-width: 80px;" onclick="onPdfCellClick(${rIdx}, ${cIdx}, '${escapeHtml(String(cell))}', this)">${escapeHtml(String(cell))}</td>`;
        });
        html += '</tr>';
    });
    html += '</table>';
    
    document.getElementById('pdf-table-container').innerHTML = html;
}

async function onPdfCellClick(r, c, val, el) {
    el.style.background = '#e74c3c';
    el.style.color = '#fff';
    el.style.fontWeight = 'bold';
    
    const isGroupedMode = document.getElementById('pdf-grouped-mode')?.checked;

    if (isGroupedMode) {
        if (pdfStep === 0) {
            pdfCoords.n1 = {r, c, val}; pdfStep++;
            document.getElementById('pdf-status').innerText = `✅ Группа: "${val}". Шаг 2: Кликните на АРТИКУЛ под ней.`;
        } else if (pdfStep === 1) {
            pdfCoords.sku = {r, c, val}; pdfStep++;
            document.getElementById('pdf-status').innerText = `✅ Артикул: "${val}". Шаг 3: Кликните на ЦЕНУ для этого артикула.`;
        } else if (pdfStep === 2) {
            pdfCoords.p1 = {r, c, val};
            await finishPdfSetup(pdfCoords.n1.c, pdfCoords.sku.c, pdfCoords.p1.c, pdfCoords.n1.r, 1, 0, 1);
        }
    } else {
        if (pdfStep === 0) {
            pdfCoords.n1 = {r, c, val}; pdfStep++;
            document.getElementById('pdf-status').innerText = `✅ Название: "${val}". Шаг 2: Кликните на ячейку с ЦЕНОЙ.`;
        } else if (pdfStep === 1) {
            pdfCoords.p1 = {r, c, val}; pdfStep++;
            document.getElementById('pdf-status').innerText = `✅ Цена: "${val}". Шаг 3: Кликните на ВТОРОЕ НАЗВАНИЕ (следующий товар).`;
        } else if (pdfStep === 2) {
            pdfCoords.n2 = {r, c, val};
            const blockSize = pdfCoords.n2.r - pdfCoords.n1.r;
            await finishPdfSetup(pdfCoords.n1.c, -1, pdfCoords.p1.c, pdfCoords.n1.r, 0, pdfCoords.p1.r - pdfCoords.n1.r, blockSize);
        }
    }
}

async function finishPdfSetup(n_col, sku_col, p_col, start_row, is_grouped, p_off = 0, block_step = 1) {
    const config = {
        chat_id: document.getElementById('pdf-chat-select').value,
        sheet_name: 'Страница 1', // Для PDF используем заглушку
        name_col: n_col,
        sku_col: sku_col,
        price_col: p_col,
        price_row_offset: p_off,
        block_step: block_step,
        start_row: start_row,
        is_grouped: is_grouped
    };

    await apiFetch('/api/pdf/save_config', {
        method: 'POST', headers: {'Content-Type': 'application/json'}, body: JSON.stringify(config)
    });
    
    alert('🎉 Настройка для PDF успешно завершена!');
    document.getElementById('pdf-preview-card').style.display = 'none';
    if(typeof loadExcelConfigs === 'function') loadExcelConfigs(); // Обновляем общую таблицу
}
// --- ЛОГИКА АВТОПУБЛИКАЦИЙ ---

async function loadPublications() {
    const tbody = document.getElementById('publications-tbody');
    if (!tbody) return;
    const pubs = await cachedApiFetch('/api/publications', true); // Принудительно обновляем кэш
    
    tbody.innerHTML = pubs.map(p => `
        <tr>
            <td><strong>${escapeHtml(p.name)}</strong></td>
            <td>${escapeHtml(p.chat_id || 'Не указан')}</td>
            <td>${p.interval_min} мин.</td>
            <td>
                <label style="cursor:pointer; display:flex; align-items:center; gap:5px;">
                    <input type="checkbox" onchange="togglePublication(${p.id}, this.checked)" ${p.is_active ? 'checked' : ''}> 
                    ${p.is_active ? '<span style="color:#2ecc71;">Вкл</span>' : '<span style="color:#e74c3c;">Выкл</span>'}
                </label>
            </td>
            <td>
                <button class="save-btn" style="background:#f39c12; padding:4px 8px; font-size:12px; margin-right:4px;" onclick="openPublishEditor(${p.id})">✏️ Ред.</button>
                <button class="save-btn" style="background:#e74c3c; padding:4px 8px; font-size:12px;" onclick="deletePublication(${p.id})">❌</button>
            </td>
        </tr>
    `).join('') || '<tr><td colspan="5" class="no-messages">Нет публикаций. Создайте первую!</td></tr>';
}

async function openPublishEditor(pubId = null) {
    document.getElementById('publish-editor-container').style.display = 'block';
    document.getElementById('pub-filter-tree').innerHTML = '<div style="padding: 10px;">Загрузка дерева товаров...</div>';
    
    // Загружаем список юзерботов в select
    const bots = await cachedApiFetch('/api/userbots');
    const botSelect = document.getElementById('pub_userbot_id');
    if(botSelect) {
        botSelect.innerHTML = '<option value="">-- Выберите бота --</option>' + 
            bots.filter(b => b.status === 'active').map(b => `<option value="${b.id}">${escapeHtml(b.account_name || 'API ID: ' + b.api_id)}</option>`).join('');
    }

    let pubData = null;
    let allowedItems = {};

    if (pubId) {
        const pubs = await cachedApiFetch('/api/publications');
        pubData = pubs.find(p => p.id === pubId);
        if (pubData) {
            document.getElementById('pub_id').value = pubData.id;
            document.getElementById('pub_name').value = pubData.name || '';
            document.getElementById('pub_enabled').checked = pubData.is_active === 1;
            document.getElementById('pub_interval').value = pubData.interval_min || 60;
            document.getElementById('pub_chat_id').value = pubData.chat_id || '';
            document.getElementById('pub_message_id').value = pubData.message_id || '';
            document.getElementById('pub_userbot_id').value = pubData.userbot_id || '';
            document.getElementById('pub_template').value = pubData.template || '';
            try { allowedItems = JSON.parse(pubData.allowed_items) || {}; } catch(e) {}
        }
    } else {
        document.getElementById('pub_id').value = '';
        document.getElementById('pub_name').value = '';
        document.getElementById('pub_enabled').checked = true;
        document.getElementById('pub_interval').value = 60;
        document.getElementById('pub_chat_id').value = '';
        document.getElementById('pub_message_id').value = '';
        document.getElementById('pub_userbot_id').value = '';
        document.getElementById('pub_template').value = 'Цены на товары:';
    }
    
    // Загрузка папок в выпадающий список
    const folders = await cachedApiFetch('/api/folders/tree');
    const folderSelect = document.getElementById('pub-markup-folder');
    if (folderSelect) {
        folderSelect.innerHTML = '<option value="0">Базовая (Для остальных)</option>';
        function addOpt(list, level=0) {
            list.forEach(f => {
                if(f.name === 'По умолчанию') return;
                folderSelect.innerHTML += `<option value="${f.id}">${'—'.repeat(level)} ${f.name}</option>`;
                if(f.children) addOpt(f.children, level+1);
            });
        }
        addOpt(folders);
    }

    // --- ИЗМЕНЕНИЕ: НАЦЕНКИ ВСЕГДА ДОСТУПНЫ ---
    document.getElementById('pub-markups-content').style.display = 'block';
    const warningText = document.getElementById('pub-markups-warning');
    if (warningText) warningText.style.display = 'none';

    if (pubId) {
        loadPubMarkups(pubId);
    } else {
        // Очищаем таблицу и создаем временную память для новых наценок
        window.tempPubMarkups = [];
        document.getElementById('pub-markups-tbody').innerHTML = '';
    }
    
    // Рендерим дерево
    const treeData = await cachedApiFetch('/api/publish_tree_data'); 
    renderPublishTree(treeData.folders, treeData.products, treeData.suppliers, allowedItems); 
}

function renderPublishTree(folders, products, suppliers, allowedItems) {
        const container = document.getElementById('pub-filter-tree');
        let html = '';

        // --- ВОТ ЭТУ ФУНКЦИЮ ПОЛНОСТЬЮ ЗАМЕНЯЕМ ---
        function buildNode(folder, level) {
            // Генерируем уникальные ID для контейнера и иконки-стрелочки
            const folderContentId = 'pub-folder-content-' + folder.id;
            const folderIconId = 'pub-folder-icon-' + folder.id;

            // Рисуем название папки как КЛИКАБЕЛЬНУЮ кнопку
            let nodeHtml = `<div style="margin-left: ${level * 20}px; margin-bottom: 5px;">
                <strong style="color: #f39c12; cursor: pointer; user-select: none;" 
                        onclick="toggleFolderNode('${folderContentId}', '${folderIconId}')">
                    <span id="${folderIconId}" style="display: inline-block; width: 15px;">▼</span> 📁 ${escapeHtml(folder.name)}
                </strong>
            </div>
            
            <div id="${folderContentId}" style="display: block;">`; // <--- ОТКРЫВАЕМ КОНТЕЙНЕР ДЛЯ СКРЫТИЯ

            // Рисуем товары внутри папки
            const folderProducts = products.filter(p => p.folder_id === folder.id);
            folderProducts.forEach(p => {
                const productSuppliers = suppliers.filter(s => s.product_id === p.id);
                if (productSuppliers.length === 0) return; // Скрываем товары без цен

                nodeHtml += `<div style="margin-left: ${(level + 1) * 20}px; margin-bottom: 3px;">
                    <div style="display:flex; align-items:center; margin-bottom: 2px;">
                        <span>📦 ${escapeHtml(p.name)}</span>
                    </div>`;

                // Рисуем чекбоксы с поставщиками
                productSuppliers.forEach(sup => {
                    const isChecked = (allowedItems[p.id] && allowedItems[p.id].includes(sup.supplier_id.toString())) ? 'checked' : '';
                    nodeHtml += `<div style="margin-left: 24px; display:flex; align-items:center; font-size:13px; color:#aaa;">
                        <input type="checkbox" class="pub-supplier-cb" data-product="${p.id}" value="${sup.supplier_id}" ${isChecked} style="margin-right: 8px; cursor:pointer;">
                        <span>👤 ${escapeHtml(sup.supplier_name || 'Неизвестный чат')} (${sup.extracted_price} руб.)</span>
                    </div>`;
                });
                nodeHtml += `</div>`;
            });

            // Рисуем подпапки
            const children = folders.filter(f => f.parent_id === folder.id);
            children.forEach(c => {
                nodeHtml += buildNode(c, level + 1);
            });

            nodeHtml += `</div>`; // <--- ЗАКРЫВАЕМ КОНТЕЙНЕР, КОТОРЫЙ БУДЕМ СВОРАЧИВАТЬ

            return nodeHtml;
        }
        // --- КОНЕЦ ЗАМЕНЫ ---

        const rootFolders = folders.filter(f => f.parent_id === null);
        rootFolders.forEach(f => {
            html += buildNode(f, 0);
        });
        container.innerHTML = html;
    }

async function savePublication() {
    const pubId = document.getElementById('pub_id').value;
    const allowedItems = {};
    
    document.querySelectorAll('.pub-supplier-cb:checked').forEach(cb => {
        const prodId = cb.getAttribute('data-product');
        const supId = cb.value;
        if (!allowedItems[prodId]) allowedItems[prodId] = [];
        allowedItems[prodId].push(supId);
    });

    const payload = {
        id: pubId ? parseInt(pubId) : null,
        name: document.getElementById('pub_name').value,
        is_active: document.getElementById('pub_enabled').checked ? 1 : 0,
        interval_min: parseInt(document.getElementById('pub_interval').value) || 60,
        chat_id: document.getElementById('pub_chat_id').value,
        message_id: document.getElementById('pub_message_id').value,
        userbot_id: document.getElementById('pub_userbot_id').value,
        template: document.getElementById('pub_template').value,
        allowed_items: allowedItems,
        temp_markups: window.tempPubMarkups || [] // <--- Передаем временные наценки
    };

    if (!payload.name || !payload.chat_id || !payload.userbot_id) {
        alert("Заполните Название, ID чата и выберите Юзербота!");
        return;
    }

    await apiFetch('/api/publications', {
        method: 'POST',
        headers: { 'Content-Type': 'application/json' },
        body: JSON.stringify(payload)
    });

    document.getElementById('publish-editor-container').style.display = 'none';
    loadPublications();
}

async function deletePublication(id) {
    if(confirm('Удалить эту настройку автопубликации?')) {
        await apiFetch(`/api/publications/${id}`, { method: 'DELETE' });
        loadPublications();
    }
}

async function togglePublication(id, isActive) {
    await apiFetch(`/api/publications/${id}/toggle`, {
        method: 'POST',
        headers: { 'Content-Type': 'application/json' },
        body: JSON.stringify({ is_active: isActive ? 1 : 0 })
    });
    loadPublications();
}
// --- ЛОГИКА ПАРСИНГА ЧУЖИХ API ---
async function loadPubMarkups(pubId) {
    const markups = await apiFetch(`/api/publications/${pubId}/markups`);
    const tbody = document.getElementById('pub-markups-tbody');
    tbody.innerHTML = markups.map(m => `
        <tr>
            <td><strong style="color:${m.folder_id === 0 ? '#f39c12' : '#fff'};">${m.folder_id === 0 ? 'Базовая (Для остальных)' : escapeHtml(m.folder_name)}</strong></td>
            <td>${m.markup_value > 0 ? '+' : ''}${m.markup_value}${m.markup_type === 'percent' ? '%' : ' руб'}</td>
            <td>До ${m.rounding}</td>
            <td>
                ${m.folder_id === 0 ? '<span style="color:#888; font-size:11px;">(Базовая)</span>' : `<button class="save-btn" style="background:#e74c3c; font-size:12px; padding:4px 8px;" onclick="deletePubMarkup(${m.id})">❌</button>`}
            </td>
        </tr>
    `).join('');
}

async function savePubMarkup() {
    const pubId = document.getElementById('pub_id').value;

    const folderSelect = document.getElementById('pub-markup-folder');
    const folderId = parseInt(folderSelect.value);
    const folderName = folderSelect.options[folderSelect.selectedIndex].text.replace(/—/g, '').trim();

    const mType = document.getElementById('pub-markup-type').value;
    const mValue = parseFloat(document.getElementById('pub-markup-value').value) || 0;
    const rounding = parseInt(document.getElementById('pub-markup-rounding').value) || 100;

    if (pubId) {
        // Если публикация уже сохранена, отправляем сразу в БД
        const body = { folder_id: folderId, markup_type: mType, markup_value: mValue, rounding: rounding };
        await apiFetch(`/api/publications/${pubId}/markups`, { method: 'POST', headers: { 'Content-Type': 'application/json' }, body: JSON.stringify(body) });
        loadPubMarkups(pubId);
    } else {
        // Если публикация новая, сохраняем в оперативную память браузера
        if (!window.tempPubMarkups) window.tempPubMarkups = [];
        
        const existingIdx = window.tempPubMarkups.findIndex(m => m.folder_id === folderId);
        if (existingIdx !== -1) {
            window.tempPubMarkups[existingIdx] = { folder_id: folderId, folder_name: folderName, markup_type: mType, markup_value: mValue, rounding: rounding };
        } else {
            window.tempPubMarkups.push({ folder_id: folderId, folder_name: folderName, markup_type: mType, markup_value: mValue, rounding: rounding });
        }
        renderTempPubMarkups();
    }
}

// Новая функция для отрисовки временных наценок
function renderTempPubMarkups() {
    const tbody = document.getElementById('pub-markups-tbody');
    tbody.innerHTML = (window.tempPubMarkups || []).map(m => `
        <tr>
            <td><strong style="color:${m.folder_id === 0 ? '#f39c12' : '#fff'};">${m.folder_id === 0 ? 'Базовая (Для остальных)' : escapeHtml(m.folder_name)}</strong></td>
            <td>${m.markup_value > 0 ? '+' : ''}${m.markup_value}${m.markup_type === 'percent' ? '%' : ' руб'}</td>
            <td>До ${m.rounding}</td>
            <td>
                ${m.folder_id === 0 ? '<span style="color:#888; font-size:11px;">(Базовая)</span>' : `<button class="save-btn" style="background:#e74c3c; font-size:12px; padding:4px 8px;" onclick="deleteTempPubMarkup(${m.folder_id})">❌</button>`}
            </td>
        </tr>
    `).join('');
}

// Новая функция для удаления временных наценок
window.deleteTempPubMarkup = function(folderId) {
    if(!confirm('Удалить эту наценку?')) return;
    window.tempPubMarkups = window.tempPubMarkups.filter(m => m.folder_id !== folderId);
    renderTempPubMarkups();
};

async function deletePubMarkup(id) {
    if(!confirm('Удалить эту наценку?')) return;
    const pubId = document.getElementById('pub_id').value;
    await apiFetch(`/api/pub_markups/${id}`, { method: 'DELETE' });
    loadPubMarkups(pubId);
}
// --- ЛОГИКА ПАРСИНГА ЧУЖИХ API ---

window.loadApiSources = async function() {
    const tbody = document.getElementById('api-sources-tbody');
    if (!tbody) return; 
    
    // Получаем свежие данные с сервера
    const sources = await cachedApiFetch('/api/api_sources', true);
    
    tbody.innerHTML = sources.map(s => `
        <tr>
            <td><strong>${escapeHtml(s.name)}</strong></td>
            <td><span style="color:#aaa; font-size:12px;">${escapeHtml(s.url)}</span></td>
            <td>${s.interval_min} мин.</td>
            <td>
                <button type="button" class="save-btn" style="background:#e74c3c; padding:4px 8px; font-size:12px;" onclick="deleteApiSource(${s.id})">❌ Удалить</button>
            </td>
        </tr>
    `).join('') || '<tr><td colspan="4" class="no-messages">Нет подключенных API</td></tr>';
};

window.loadApiSources = async function() {
    const tbody = document.getElementById('api-sources-tbody');
    if (!tbody) return; 
    
    // Получаем свежие данные с сервера
    const sources = await cachedApiFetch('/api/api_sources', true);
    
    tbody.innerHTML = sources.map(s => `
        <tr>
            <td><strong>${escapeHtml(s.name)}</strong></td>
            <td><span style="color:#aaa; font-size:12px;">${escapeHtml(s.url)}</span></td>
            <td>${s.interval_min} мин.</td>
            <td>
                <button type="button" class="save-btn" style="background:#e74c3c; padding:4px 8px; font-size:12px;" onclick="window.deleteApiSource(${s.id})">❌ Удалить</button>
            </td>
        </tr>
    `).join('') || '<tr><td colspan="4" class="no-messages">Нет подключенных API</td></tr>';
};

window.deleteApiSource = async function(id) {
    console.log("Вызвана функция удаления для ID:", id); // Вывод в консоль для проверки
    if (confirm('Удалить этот источник API?')) {
        try {
            await apiFetch(`/api/api_sources/${id}`, { method: 'DELETE' });
            loadApiSources(); // Обновляем таблицу после удаления
        } catch (e) {
            console.error("Ошибка при удалении:", e);
            alert("Произошла ошибка при удалении. Посмотрите консоль.");
        }
    }
};


window.saveApiSource = async function() {
    const payload = {
        name: document.getElementById('api_source_name').value.trim(),
        url: document.getElementById('api_source_url').value.trim(),
        token: document.getElementById('api_source_token').value.trim(),
        interval_min: parseInt(document.getElementById('api_source_interval').value) || 60
    };
    
    if (!payload.name || !payload.url || !payload.token) {
        return alert('Заполните Название, Ссылку и Токен!');
    }
    
    try {
        await apiFetch('/api/api_sources', { 
            method: 'POST', 
            headers: {'Content-Type': 'application/json'}, 
            body: JSON.stringify(payload) 
        });
        
        // Очищаем форму
        document.getElementById('api_source_name').value = '';
        document.getElementById('api_source_url').value = '';
        document.getElementById('api_source_token').value = '';
        document.getElementById('api_source_interval').value = 60;
        
        // Обновляем таблицу
        window.loadApiSources();
    } catch (e) {
        console.error('Ошибка сохранения API:', e);
        alert('Не удалось сохранить источник. Проверьте консоль.');
    }
};
// --- ФУНКЦИЯ ДЛЯ СВОРАЧИВАНИЯ ПАПОК ПРИ ПУБЛИКАЦИИ ---
window.toggleFolderNode = function(param1, param2) {
    let container, icon;
    
    // Подстраиваемся под любую вашу вкладку (Склад или Автопубликацию)
    if (param2 === undefined) {
        container = document.getElementById('folder-children-' + param1);
        icon = document.getElementById('folder-icon-' + param1);
    } else {
        container = document.getElementById(param1);
        icon = document.getElementById(param2);
    }
    
    if (container && icon) {
        // Проверяем текст самой иконки! Это 100% надежный визуальный маркер.
        if (icon.textContent.includes('▼')) {
            // Если стрелочка вниз (папка выглядит открытой) -> скрываем содержимое
            container.style.display = 'none';
            icon.textContent = '▶';
        } else {
            // Если стрелочка вправо (папка выглядит закрытой) -> показываем содержимое
            container.style.display = 'block';
            icon.textContent = '▼';
        }
    }
};


// Запускаем восстановление
restoreUIState();
    </script>
</body>
</html>
"""


@app.route('/api/fix_db')
@login_required
def fix_db():
    db = get_db()
    try:
        # Отключаем проверки ключей на секунду, чтобы SQLite разрешил замену
        db.execute("PRAGMA foreign_keys = OFF")
        db.executescript('''
            CREATE TABLE IF NOT EXISTS messages_new (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                telegram_message_id INTEGER,
                type TEXT,
                text TEXT,
                date TIMESTAMP,
                chat_id INTEGER,
                chat_title TEXT,
                is_blocked INTEGER DEFAULT 0,
                sender_name TEXT,
                FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE,
                UNIQUE(chat_id, telegram_message_id)
            );
            INSERT OR IGNORE INTO messages_new (id, user_id, telegram_message_id, type, text, date, chat_id, chat_title, is_blocked, sender_name)
            SELECT id, user_id, telegram_message_id, type, text, date, chat_id, chat_title, is_blocked, sender_name FROM messages;
            DROP TABLE messages;
            ALTER TABLE messages_new RENAME TO messages;
        ''')
        db.execute("PRAGMA foreign_keys = ON")
        db.commit()
        return "<h1>✅ База данных успешно исправлена!</h1><p>Глобальная блокировка сообщений снята. Теперь парсер будет выкачивать всё.</p><br><a href='/'>Вернуться в парсер</a>"
    except Exception as e:
        return f"<h1>❌ Ошибка:</h1><p>{e}</p>"

@app.route('/api/cleanup_old_bug')
@login_required
def cleanup_old_bug():
    user_id = session['user_id']
    db = get_db()
    
    # Считаем, сколько сломанных сообщений мы найдем
    cursor = db.execute("SELECT COUNT(*) FROM messages WHERE telegram_message_id IS NULL AND user_id = ?", (user_id,))
    bad_count = cursor.fetchone()[0]
    
    if bad_count > 0:
        # Удаляем их (каскадное удаление само почистит привязки к товарам, если они были)
        db.execute("DELETE FROM messages WHERE telegram_message_id IS NULL AND user_id = ?", (user_id,))
        db.commit()
        notify_clients()
        return f"<h1>Уборка завершена! 🧹</h1><p>Удалено сломанных сообщений: <b>{bad_count}</b>.</p><br><a href='/'>Вернуться в парсер</a>"
    else:
        return "<h1>Всё чисто! ✨</h1><p>Сломанных сообщений в базе не найдено.</p><br><a href='/'>Вернуться в парсер</a>"






# ================= УПРАВЛЕНИЕ API И КЛИЕНТАМИ =================

# ✅ А ЭТОТ БЛОК ДОЛЖЕН ОСТАТЬСЯ ✅
@app.route('/api/api_clients/<int:client_id>/filters', methods=['POST'])
@login_required
def update_api_client_filters(client_id):
    data = request.json
    access_rules = json.dumps(data.get('access_rules', {})) # Конвертируем в JSON
    db = get_db()
    db.execute("UPDATE api_clients SET access_rules=? WHERE id=? AND user_id=?", 
               (access_rules, client_id, session['user_id']))
    db.commit()
    return jsonify({'success': True})


@app.route('/api/api_clients', methods=['GET', 'POST'])
@login_required
def manage_api_clients():
    user_id = session['user_id']
    db = get_db()
    if request.method == 'GET':
        # Теперь достаем все поля, включая тайминги
        clients = db.execute("SELECT * FROM api_clients WHERE user_id=?", (user_id,)).fetchall()
        return jsonify([dict(c) for c in clients])
    
    # POST: Создаем нового клиента
    data = request.json
    name = data.get('name')
    schedule_enabled = int(data.get('schedule_enabled', 0))
    time_start = data.get('time_start', '00:00')
    time_end = data.get('time_end', '23:59')
    token = secrets.token_hex(24) 
    
    db.execute("""
        INSERT INTO api_clients (user_id, name, token, schedule_enabled, time_start, time_end) 
        VALUES (?, ?, ?, ?, ?, ?)
    """, (user_id, name, token, schedule_enabled, time_start, time_end))
    
    client_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
    
    db.execute("INSERT INTO api_markups (client_id, folder_id, markup_type, markup_value, rounding) VALUES (?, 0, 'percent', 0, 100)", (client_id,))
    db.commit()
    return jsonify({'success': True})

# --- ДОБАВЬ НОВЫЙ МАРШРУТ ДЛЯ РЕДАКТИРОВАНИЯ ТАЙМИНГОВ ---
@app.route('/api/api_clients/<int:client_id>/schedule', methods=['POST'])
@login_required
def update_api_client_schedule(client_id):
    data = request.json
    db = get_db()
    db.execute("""
        UPDATE api_clients 
        SET schedule_enabled=?, time_start=?, time_end=? 
        WHERE id=? AND user_id=?
    """, (int(data['schedule_enabled']), data['time_start'], data['time_end'], client_id, session['user_id']))
    db.commit()
    return jsonify({'success': True})

@app.route('/api/api_clients/<int:client_id>', methods=['DELETE'])
@login_required
def delete_api_client(client_id):
    db = get_db()
    db.execute("DELETE FROM api_clients WHERE id=? AND user_id=?", (client_id, session['user_id']))
    db.commit()
    return jsonify({'success': True})

@app.route('/api/api_clients/<int:client_id>/markups', methods=['GET', 'POST'])
@login_required
def manage_markups(client_id):
    db = get_db()
    if request.method == 'GET':
        markups = db.execute("""
            SELECT am.*, f.name as folder_name 
            FROM api_markups am 
            LEFT JOIN folders f ON am.folder_id = f.id 
            WHERE am.client_id=?
        """, (client_id,)).fetchall()
        return jsonify([dict(m) for m in markups])
    
    data = request.json
    folder_id = data.get('folder_id') or 0
    db.execute("""
        INSERT INTO api_markups (client_id, folder_id, markup_type, markup_value, rounding) 
        VALUES (?, ?, ?, ?, ?)
        ON CONFLICT(client_id, folder_id) DO UPDATE SET 
        markup_type=excluded.markup_type, markup_value=excluded.markup_value, rounding=excluded.rounding
    """, (client_id, folder_id, data['markup_type'], data['markup_value'], data['rounding']))
    db.commit()
    return jsonify({'success': True})

@app.route('/api/api_markups/<int:markup_id>', methods=['DELETE'])
@login_required
def delete_markup(markup_id):
    db = get_db()
    # Запрещаем удалять базовую настройку (folder_id = 0)
    db.execute("DELETE FROM api_markups WHERE id=? AND folder_id != 0 AND client_id IN (SELECT id FROM api_clients WHERE user_id=?)", (markup_id, session['user_id']))
    db.commit()
    return jsonify({'success': True})


# ================= ПУБЛИЧНЫЙ API ДЛЯ ВЫДАЧИ КАТАЛОГА =================
@app.route('/api/v1/catalog', methods=['GET'])
def public_api_catalog():
    token = request.args.get('token')
    if not token:
        return jsonify({'error': 'Токен не предоставлен'}), 401
        
    db = get_db()
    client = db.execute("SELECT * FROM api_clients WHERE token=?", (token,)).fetchone()
    if not client:
        return jsonify({'error': 'Неверный токен'}), 401
        
    if client['schedule_enabled']:
        from datetime import timedelta
        now = (datetime.utcnow() + timedelta(hours=3)).time()
        start_time = datetime.strptime(client['time_start'], '%H:%M').time()
        end_time = datetime.strptime(client['time_end'], '%H:%M').time()
        is_active = start_time <= now <= end_time if start_time <= end_time else (now >= start_time or now <= end_time)
        if not is_active: 
            return jsonify({'categories': [], 'products': []})
            
    user_id = client['user_id']
    client_id = client['id']

    markups_raw = db.execute("SELECT folder_id, markup_type, markup_value, rounding FROM api_markups WHERE client_id=?", (client_id,)).fetchall()
    markups = {}
    default_markup = {'markup_type': 'percent', 'markup_value': 0, 'rounding': 100}
    for m in markups_raw:
        if m['folder_id'] == 0: 
            default_markup = dict(m)
        else: 
            markups[m['folder_id']] = dict(m)

    # 1. Получаем базовую информацию обо всех товарах
    products_raw = db.execute("SELECT id, name, folder_id, price as manual_price FROM products WHERE user_id = ?", (user_id,)).fetchall()
    
    # 2. Получаем ВСЕ актуальные цены, чтобы потом отфильтровать их по нужным поставщикам
    prices_raw = db.execute("""
        SELECT pm.product_id, pm.extracted_price, m.chat_id, m.date
        FROM product_messages pm
        JOIN messages m ON pm.message_id = m.id
        WHERE pm.status = 'confirmed' AND pm.is_actual = 1 AND m.user_id = ?
        ORDER BY m.date DESC
    """, (user_id,)).fetchall()
    
    product_prices = {}
    for pr in prices_raw:
        pid = pr['product_id']
        if pid not in product_prices:
            product_prices[pid] = []
        product_prices[pid].append(pr)

    products = []
    active_folder_ids = set()
    
    try:
        access_rules = json.loads(client['access_rules']) if 'access_rules' in client.keys() and client['access_rules'] else {}
    except:
        access_rules = {}

    for p in products_raw:
        if not access_rules: 
            continue # Если клиенту ничего не настроено - отдаем пустой прайс
            
        pid_str = str(p['id'])
        if pid_str not in access_rules: 
            continue # Товар галочкой не отмечен
            
        allowed_chats = access_rules[pid_str]
        prices = product_prices.get(p['id'], [])
        best_price = None

        if allowed_chats == 'all' or allowed_chats == ['all']:
            if prices:
                best_price = prices[0]['extracted_price']
        else:
            if isinstance(allowed_chats, str):
                allowed_chats = [allowed_chats]
                
            # Ищем самую свежую цену строго от РАЗРЕШЕННОГО поставщика
            for pr in prices:
                chat_id_str = str(pr['chat_id'])
                is_allowed = False
                for ac in allowed_chats:
                    ac_str = str(ac)
                    if chat_id_str == ac_str or chat_id_str == f"api_src_{ac_str}" or ac_str == f"api_src_{chat_id_str}":
                        is_allowed = True
                        break
                if is_allowed:
                    best_price = pr['extracted_price']
                    break # Нашли правильную цену!

        base_price = float(best_price if best_price is not None else (p['manual_price'] or 0))
        
        # Если цены вообще нет — пропускаем
        if base_price == 0:
            continue

        mk = markups.get(p['folder_id'], default_markup)
        if mk['markup_type'] == 'percent':
            final_price = base_price * (1 + mk['markup_value'] / 100)
        else:
            final_price = base_price + mk['markup_value']
            
        rnd = mk['rounding']
        if rnd > 0:
            final_price = math.ceil(final_price / rnd) * rnd
            
        products.append({
            'id': p['id'],
            'name': p['name'],
            'category_id': p['folder_id'],
            'price': int(final_price)
        })
        
        if p['folder_id']:
            active_folder_ids.add(p['folder_id'])

    # Получаем категории
    folders_raw = db.execute("SELECT id, name, parent_id FROM folders WHERE user_id=?", (user_id,)).fetchall()
    parent_map = {f['id']: f['parent_id'] for f in folders_raw}
    folders_to_keep = set(active_folder_ids)
    
    for fid in active_folder_ids:
        current = fid
        while current in parent_map and parent_map[current] is not None:
            current = parent_map[current]
            folders_to_keep.add(current)
            
    categories = [{'id': f['id'], 'name': f['name'], 'parent_id': f['parent_id']} 
                  for f in folders_raw if f['id'] in folders_to_keep]
                  
    # --- СОРТИРОВКА ТОВАРОВ И КАТЕГОРИЙ ПО ИМЕНИ (А-Я) ---
    products.sort(key=lambda x: (x['name'] or '').lower())
    categories.sort(key=lambda x: (x['name'] or '').lower())
    # -----------------------------------------------------
        
    return jsonify({
        'categories': categories,
        'products': products
    })


# ================= ПАРСИНГ EXCEL ТАБЛИЦ =================
async def parse_excel_message(client, message, chat_id, user_id):
    text_to_save = message.text or ""
    if not message.document:
        return text_to_save
        
    ext = getattr(message.file, 'ext', '').lower()
    if ext not in ['.xlsx', '.xls', '.pdf']:
        return text_to_save

    # Функция для приведения ID к единому формату (без -100)
    def clean_id(cid):
        s = str(cid)
        if s.startswith('-100'): return s[4:]
        if s.startswith('-'): return s[1:]
        return s

    with app.app_context():
        db = get_db()
        target_id = clean_id(chat_id)
        
        # Загружаем все конфиги для этого пользователя
        all_configs = db.execute("SELECT * FROM excel_configs WHERE user_id = ?", (user_id,)).fetchall()
        configs = {}
        for c in all_configs:
            if clean_id(c['chat_id']) == target_id:
                configs[c['sheet_name']] = dict(c)
        
        # Если конфигов нет вообще - выходим сразу
        if not configs:
            logger.warning(f"Нет настроек парсинга для чата {chat_id}")
            return text_to_save

        tmp_path = None
        try:
            # Скачиваем файл во временное хранилище
            suffix = ext
            with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
                await client.download_media(message, tmp.name)
                tmp_path = tmp.name
                
            parsed_lines = []
            unknown_sheets = []
            wildcard_config = configs.get('*')

            # --- ОБРАБОТКА PDF ---
            if ext == '.pdf':
                import pdfplumber
                # Для PDF ищем: 'Страница 1', затем '*', если нет - берем первый попавшийся
                config = configs.get('Страница 1') or wildcard_config
                if not config and configs:
                    config = list(configs.values())[0]
                
                if config:
                    with pdfplumber.open(tmp_path) as pdf:
                        for page in pdf.pages:
                            table = page.extract_table()
                            if not table: continue
                            
                            # Превращаем таблицу в безопасный список строк
                            clean_rows = []
                            for row in table:
                                clean_rows.append([str(cell).replace('\n', ' ') if cell is not None else "" for cell in row])
                            
                            step = int(config['block_step'])
                            start = int(config['start_row'])
                            n_col = int(config['name_col'])
                            p_col = int(config['price_col'])
                            n_off = int(config['name_row_offset'])
                            p_off = int(config['price_row_offset'])
                            is_grouped = config.get('is_grouped', 0)
                            sku_col = config.get('sku_col', -1)

                            if is_grouped == 1:
                                current_group = ""
                                for i in range(start, len(clean_rows)):
                                    row = clean_rows[i]
                                    # Проверка: хватает ли колонок в этой строке PDF?
                                    if len(row) <= max(n_col, p_col, sku_col): continue
                                    
                                    col_n_val = row[n_col].strip()
                                    col_sku_val = row[sku_col].strip() if sku_col != -1 else ""
                                    col_p_val = row[p_col].strip()
                                    
                                    if col_n_val and not col_p_val: current_group = col_n_val
                                    if col_p_val and current_group:
                                        name = col_sku_val if col_sku_val else col_n_val
                                        parsed_lines.append(f"{current_group} {name} - {col_p_val}")
                            else:
                                for i in range(start, len(clean_rows), step):
                                    name_idx = i + n_off
                                    price_idx = i + p_off
                                    if name_idx < len(clean_rows) and price_idx < len(clean_rows):
                                        row_n = clean_rows[name_idx]
                                        row_p = clean_rows[price_idx]
                                        if len(row_n) > n_col and len(row_p) > p_col:
                                            n_val = row_n[n_col].strip()
                                            p_val = row_p[p_col].strip()
                                            if n_val and p_val:
                                                parsed_lines.append(f"{n_val} - {p_val}")

            # --- ОБРАБОТКА EXCEL ---
            else:
                xls = pd.ExcelFile(tmp_path)
                for sheet_name in xls.sheet_names:
                    config = configs.get(sheet_name) or wildcard_config
                    if not config:
                        unknown_sheets.append(sheet_name)
                        continue            
                
                        
                    df = pd.read_excel(xls, sheet_name=sheet_name, header=None)
                    df = df.fillna("")
                    
                    step = config['block_step']
                    start = config['start_row']
                    n_col = config['name_col']
                    p_col = config['price_col']
                    n_off = config['name_row_offset']
                    p_off = config['price_row_offset']
                    is_grouped = config.get('is_grouped', 0)
                    sku_col = config.get('sku_col', -1)
                    
                    if is_grouped == 1:
                        current_group = ""
                        for i in range(start, len(df)):
                            try:
                                col_n_val = str(df.iat[i, n_col]).strip()
                                col_sku_val = str(df.iat[i, sku_col]).strip() if sku_col != -1 else ""
                                col_p_val = str(df.iat[i, p_col]).strip()
                                
                                is_empty_name = (not col_n_val or col_n_val == 'nan')
                                has_price = (col_p_val and col_p_val != 'nan')
                                
                                if n_col == sku_col:
                                    if col_n_val and not has_price: current_group = col_n_val
                                    elif col_n_val and has_price and current_group:
                                        parsed_lines.append(f"{current_group} {col_n_val} - {col_p_val}")
                                else:
                                    if col_n_val and not is_empty_name: current_group = col_n_val
                                    if col_sku_val and col_sku_val != 'nan' and has_price and current_group:
                                        parsed_lines.append(f"{current_group} {col_sku_val} - {col_p_val}")
                            except Exception:
                                continue
                    else:
                        for i in range(start, len(df), step):
                            try:
                                name_row = i + n_off
                                price_row = i + p_off
                                if name_row < len(df) and price_row < len(df):
                                    name_val = str(df.iat[name_row, n_col]).strip()
                                    price_val = str(df.iat[price_row, p_col]).strip()
                                    
                                    if name_val and name_val != 'nan' and price_val and price_val != 'nan':
                                        parsed_lines.append(f"{name_val} - {price_val}")
                            except Exception:
                                continue
            
            # --- ФОРМИРОВАНИЕ ИТОГОВОГО ТЕКСТА ---
            if parsed_lines:
                doc_type = "PDF" if ext == '.pdf' else "EXCEL"
                text_to_save = f"📊 [{doc_type} ДАННЫЕ]:\n" + "\n".join(parsed_lines)
                
                if unknown_sheets and ext != '.pdf':
                    with app.app_context():
                        db = get_db()
                        all_chats = db.execute("SELECT chat_id, chat_title, custom_name FROM tracked_chats WHERE user_id = ?", (user_id,)).fetchall()
                        chat_title = str(chat_id)
                        for tc in all_chats:
                            if clean_id(tc['chat_id']) == target_id:
                                chat_title = tc['custom_name'] if tc['custom_name'] else tc['chat_title']
                                break
                                
                        for us in unknown_sheets:
                            db.execute("INSERT OR IGNORE INTO excel_missing_sheets (user_id, chat_id, chat_title, sheet_name) VALUES (?, ?, ?, ?)", (user_id, chat_id, chat_title, us))
                        db.commit()
                        
                    warning_text = f"\n\n⚠️ Внимание! В файле найдены ненастроенные листы: {', '.join(unknown_sheets)}. Перейдите в раздел «Парсинг Excel», чтобы их настроить."
                    if text_to_save:
                        text_to_save += warning_text
                    else:
                        text_to_save = warning_text.strip()
                        
        except Exception as e:
            logger.error(f"Ошибка парсинга файла: {e}")
        finally:
            if 'tmp_path' in locals() and os.path.exists(tmp_path):
                os.remove(tmp_path)
                
    return text_to_save


@app.route('/api/excel/preview_latest/<chat_id>', methods=['GET'])
@login_required
def preview_latest_excel(chat_id):
    chat_id = int(chat_id) # Добавляем преобразование здесь
    user_id = session['user_id']
    
    # Ищем активную сессию юзербота
    client_to_use = None
    client_loop_to_use = None
    for sid, (thread, client, uid) in user_clients.items():
        if uid == user_id:
            client_to_use = client
            client_loop_to_use = user_loops.get(sid)
            break
            
    if not client_to_use:
        return jsonify({'error': 'Юзербот не запущен. Запустите его во вкладке "Юзерботы".'}), 400

    async def fetch_and_preview():
        try:
            entity = await client_to_use.get_entity(chat_id)
            # Ищем последние 50 сообщений
            async for message in client_to_use.iter_messages(entity, limit=50):
                if message.document and getattr(message.file, 'ext', '').lower() in ['.xlsx', '.xls', '.pdf']:
                    ext = getattr(message.file, 'ext', '').lower()
                    
                    # Скачиваем файл во временную директорию
                    with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
                        await client_to_use.download_media(message, tmp.name)
                        tmp_path = tmp.name
                        
                    try:
                        xls = pd.ExcelFile(tmp_path)
                        data = {}
                        # Читаем по 15 строк для предпросмотра
                        for sheet in xls.sheet_names:
                            df = pd.read_excel(xls, sheet_name=sheet, header=None, nrows=15)
                            df = df.fillna("")
                            data[sheet] = df.values.tolist()
                        return {'success': True, 'sheets_data': data}
                    finally:
                        if os.path.exists(tmp_path):
                            os.remove(tmp_path)
                            
            return {'success': False, 'error': 'В последних 50 сообщениях чата не найден Excel файл'}
        except Exception as e:
            return {'success': False, 'error': str(e)}

    # Запускаем асинхронную задачу в цикле юзербота и ждем результат
    future = asyncio.run_coroutine_threadsafe(fetch_and_preview(), client_loop_to_use)
    try:
        result = future.result(timeout=30)
        if result.get('success'):
            return jsonify(result)
        else:
            return jsonify({'error': result.get('error')}), 400
    except Exception as e:
        return jsonify({'error': 'Таймаут или ошибка при скачивании файла: ' + str(e)}), 500


@app.route('/api/excel/preview', methods=['POST'])
@login_required
def preview_excel():
    if 'file' not in request.files:
        return jsonify({'error': 'Нет файла'}), 400
    file = request.files['file']
    try:
        xls = pd.ExcelFile(file)
        data = {}
        # Читаем по 15 строк с КАЖДОГО листа
        for sheet in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet, header=None, nrows=15)
            df = df.fillna("")
            data[sheet] = df.values.tolist()
        return jsonify({'success': True, 'sheets_data': data})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/excel/save_config', methods=['POST'])
@login_required
def save_excel_config():
    data = request.json
    db = get_db()
    
    chat_id = int(data['chat_id'])
    sheet_name = data['sheet_name']
    is_grouped = data.get('is_grouped', 0)
    sku_col = data.get('sku_col', -1)
    
    db.execute("""
        INSERT INTO excel_configs (user_id, chat_id, sheet_name, name_col, name_row_offset, price_col, price_row_offset, block_step, start_row, is_grouped, sku_col) 
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(chat_id, sheet_name) DO UPDATE SET 
        name_col=excluded.name_col, name_row_offset=excluded.name_row_offset, price_col=excluded.price_col, 
        price_row_offset=excluded.price_row_offset, block_step=excluded.block_step, start_row=excluded.start_row,
        is_grouped=excluded.is_grouped, sku_col=excluded.sku_col
    """, (session['user_id'], chat_id, sheet_name, data['name_col'], data['name_row_offset'], data['price_col'], data['price_row_offset'], data['block_step'], data['start_row'], is_grouped, sku_col))

    cid_str = str(chat_id)
    core_id = cid_str[4:] if cid_str.startswith('-100') else (cid_str[1:] if cid_str.startswith('-') else cid_str)
    possible_ids = [core_id, f"-{core_id}", f"-100{core_id}"]
    
    for pid in possible_ids:
        if sheet_name == '*':
            db.execute("DELETE FROM excel_missing_sheets WHERE chat_id = ? AND user_id = ?", (pid, session['user_id']))
        else:
            db.execute("DELETE FROM excel_missing_sheets WHERE chat_id = ? AND sheet_name = ? AND user_id = ?", (pid, sheet_name, session['user_id']))
    
    db.commit()
    return jsonify({'success': True})



@app.route('/api/excel/configs', methods=['GET'])
@login_required
def get_excel_configs():
    db = get_db()
    # Используем LEFT JOIN, чтобы записи не пропадали, если чат вдруг не найден
    configs = db.execute("""
        SELECT e.*, tc.chat_title 
        FROM excel_configs e 
        LEFT JOIN tracked_chats tc ON CAST(e.chat_id AS TEXT) = CAST(tc.chat_id AS TEXT)
        WHERE e.user_id = ?
    """, (session['user_id'],)).fetchall()
    
    return jsonify([dict(c) for c in configs])

@app.route('/api/excel/configs/<int:config_id>', methods=['DELETE'])
@login_required
def delete_excel_config(config_id):
    db = get_db()
    db.execute("DELETE FROM excel_configs WHERE id = ? AND user_id = ?", (config_id, session['user_id']))
    db.commit()
    return jsonify({'success': True})


import sqlite3

def update_database():
    conn = sqlite3.connect('app.db')
    cursor = conn.cursor()
    
    print("Начинаю проверку базы данных...")

    # --- ДОБАВЛЯЕМ КОЛОНКУ IS_DELAYED ---
    try:
        cursor.execute("ALTER TABLE messages ADD COLUMN is_delayed INTEGER DEFAULT 0")
        print("✅ Колонка is_delayed успешно добавлена в таблицу messages!")
    except sqlite3.OperationalError:
        pass

    # Полный список всех колонок для работы парсера (Google, PDF, Excel)
    columns_to_add = {
        "source_type": "TEXT DEFAULT 'excel'",
        "sheet_url": "TEXT",
        "parse_interval": "INTEGER DEFAULT 60",
        "is_grouped": "INTEGER DEFAULT 0",
        "sku_col": "INTEGER DEFAULT -1",
        "name_col": "INTEGER DEFAULT -1",
        "name_row_offset": "INTEGER DEFAULT 0",
        "price_row_offset": "INTEGER DEFAULT 0",
        "block_step": "INTEGER DEFAULT 1",
        "sheet_name": "TEXT DEFAULT '*'"
    }

    for col, col_type in columns_to_add.items():
        try:
            cursor.execute(f"ALTER TABLE excel_configs ADD COLUMN {col} {col_type}")
            print(f"✅ Колонка {col} успешно добавлена.")
        except sqlite3.OperationalError:
            print(f"⏩ Колонка {col} уже существует, пропускаем.")

    # Устанавливаем тип 'excel' для всех записей, где он еще не задан
    cursor.execute("UPDATE excel_configs SET source_type = 'excel' WHERE source_type IS NULL OR source_type = ''")

    # --- ДОБАВЛЯЕМ КОЛОНКИ FOLDER_ID ---
    try:
        cursor.execute("ALTER TABLE target_chats ADD COLUMN folder_id INTEGER;")
        print("✅ Колонка folder_id добавлена в таблицу target_chats")
    except sqlite3.OperationalError:
        pass

    try:
        cursor.execute("ALTER TABLE excel_configs ADD COLUMN folder_id INTEGER;")
        print("✅ Колонка folder_id добавлена в таблицу excel_configs")
    except sqlite3.OperationalError:
        pass

    # Фиксируем все изменения ОДИН РАЗ и закрываем базу
    conn.commit()
    conn.close()
    
    print("✅ Данные обновлены, старые конфиги теперь помечены как excel")
    print("🎉 Полное обновление базы данных завершено!")




@app.route('/api/fix_delayed')
def fix_delayed():
    try:
        db = get_db()
        db.execute("ALTER TABLE messages ADD COLUMN is_delayed INTEGER DEFAULT 0")
        db.commit()
        return "<h1>✅ База данных успешно обновлена! Колонка is_delayed добавлена.</h1>"
    except Exception as e:
        return f"<h1>Уже добавлена или произошла ошибка:</h1><p>{e}</p>"

from datetime import datetime

def is_parsing_allowed(session_id):
    """
    Проверяет, попадает ли текущее время в разрешенный интервал тайминга.
    """
    try:
        with app.app_context():
            db = get_db()
            # Добавили проверку schedule_enabled
            row = db.execute("SELECT schedule_enabled, time_start, time_end FROM user_telegram_sessions WHERE id = ?", (session_id,)).fetchone()
            
            # Если галочка "Расписание" не стоит или данных нет — разрешаем 24/7
            if not row or not row['schedule_enabled']:
                return True 
            
            # Сдвигаем время сервера на +3 часа (Москва).
            current_time = (datetime.utcnow() + timedelta(hours=3)).time()
            
            start_time = datetime.strptime(row['time_start'], '%H:%M').time()
            end_time = datetime.strptime(row['time_end'], '%H:%M').time()

            if start_time <= end_time:
                return start_time <= current_time <= end_time
            else:
                return current_time >= start_time or current_time <= end_time

    except Exception as e:
        logger.error(f"Ошибка проверки тайминга для сессии {session_id}: {e}")
        return True

def delayed_messages_scheduler():
    while True:
        try:
            with app.app_context():
                db = get_db()
                # Проверяем все активные аккаунты
                cursor = db.execute("SELECT id as session_id, user_id FROM user_telegram_sessions WHERE status = 'active'")
                for session_row in cursor.fetchall():
                    session_id = session_row['session_id']
                    user_id = session_row['user_id']
                    
                    # Если для этого аккаунта СЕЙЧАС рабочее время
                    if is_parsing_allowed(session_id):
                        # Достаем все его скрытые сообщения
                        msg_cursor = db.execute("SELECT * FROM messages WHERE is_delayed = 1 AND user_id = ? ORDER BY date ASC", (user_id,))
                        delayed_messages = msg_cursor.fetchall()
                        
                        for msg in delayed_messages:
                            msg_id = msg['id']
                            parsed_text = msg['text']
                            chat_id = msg['chat_id']
                            
                            # 1. Снимаем флаг 'скрыто'
                            db.execute("UPDATE messages SET is_delayed = 0 WHERE id = ?", (msg_id,))
                            
                            # 2. ЗАПУСКАЕМ ОБРАБОТКУ (миграция привязок к товарам)
                            old_bindings = db.execute("""
                                SELECT pm.id, pm.line_index, p.synonyms
                                FROM product_messages pm
                                JOIN products p ON pm.product_id = p.id
                                JOIN messages m ON pm.message_id = m.id
                                WHERE m.chat_id = ? AND m.user_id = ? AND m.id != ?
                            """, (chat_id, user_id, msg_id)).fetchall()

                            if old_bindings and parsed_text:
                                lines = parsed_text.split('\n')
                                text_lower = parsed_text.lower()
                                
                                for b in old_bindings:
                                    synonyms = [s.strip().lower() for s in (b['synonyms'] or '').split(',') if s.strip()]
                                    line_idx = b['line_index']
                                    match_found = False
                                    new_price = None
                                    
                                    if line_idx == -1:
                                        if any(syn in text_lower for syn in synonyms):
                                            match_found = True
                                            new_price = extract_price(parsed_text)
                                    else:
                                        for i, line in enumerate(lines):
                                            if any(syn in line.lower() for syn in synonyms):
                                                match_found = True
                                                new_price = extract_price(line)
                                                if not new_price and i + 1 < len(lines):
                                                    new_price = extract_price(lines[i+1])
                                                if not new_price and i + 2 < len(lines):
                                                    new_price = extract_price(lines[i+2])
                                                line_idx = i
                                                break
                                     
                                    if match_found:
                                        db.execute("""
                                            UPDATE product_messages 
                                            SET message_id = ?, line_index = ?, extracted_price = ?, is_actual = 1 
                                            WHERE id = ?
                                        """, (msg_id, line_idx, new_price, b['id']))
                            
                            # Очистка старых сообщений (зазор 5 минут)
                            
                            # Зазор 0 минут: как только пришел новый прайс, старый удаляется МГНОВЕННО
                            cutoff_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                            db.execute("""
                                DELETE FROM messages 
                                WHERE chat_id = ? AND user_id = ? AND id != ? AND date < ?
                                  AND id NOT IN (SELECT message_id FROM product_messages)
                            """, (chat_id, user_id, msg_id, cutoff_time))

                            db.commit()
                            
                            # 3. Отправляем уведомление на фронтенд
                            notify_clients()
                            
        except Exception as e:
            print(f"Ошибка в планировщике отложенных сообщений: {e}")
            
        time.sleep(60) # Проверяем тайминги каждую минуту

# ---------- Запуск приложения ----------
update_database()
if __name__ == '__main__':
    update_database()
    threading.Thread(target=publish_scheduler, daemon=True).start()
    threading.Thread(target=api_parser_scheduler, daemon=True).start()
    threading.Thread(target=delayed_messages_scheduler, daemon=True).start()
    
    with app.app_context():
        db = get_db()
        cursor = db.execute("SELECT id, user_id, api_id, api_hash FROM user_telegram_sessions WHERE status = 'active'")
        for row in cursor.fetchall():
            start_message_listener(row['id'], row['user_id'], row['api_id'], row['api_hash'])
    # Запускаем планировщик взаимодействия с ботами
    threading.Thread(target=bot_interaction_scheduler, daemon=True).start()
    socketio.run(app, debug=True, host='0.0.0.0', port=8080, use_reloader=False, allow_unsafe_werkzeug=True)
